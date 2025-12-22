"""
Script para crear el archivo Excel de configuración inicial
Ejecutar una sola vez para crear config/empresas_config.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_config_excel():
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # ═══════════════════════════════════════════════════════════════
    # HOJA FIDI
    # ═══════════════════════════════════════════════════════════════
    ws = wb.create_sheet("FIDI")
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    section_font = Font(bold=True, size=12, color="006400")
    
    row = 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: TENANT
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="TENANT").font = section_font
    row += 1
    
    headers = ["Campo", "Valor"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    tenant_data = [
        ("key", "FIDI"),
        ("rut", "77285542-7"),
        ("nombre", "Fidi SpA"),
    ]
    for campo, valor in tenant_data:
        ws.cell(row=row, column=1, value=campo)
        ws.cell(row=row, column=2, value=valor)
        row += 1
    
    row += 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: PERIODOS
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="PERIODOS").font = section_font
    row += 1
    
    headers = ["Campo", "Valor"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    periodos_data = [
        ("actual", "202511"),
        ("fecha_corte", "2025-11-30"),
        ("tasa_impuesto", "0.27"),
    ]
    for campo, valor in periodos_data:
        ws.cell(row=row, column=1, value=campo)
        ws.cell(row=row, column=2, value=valor)
        row += 1
    
    row += 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: PERIODOS COMPARATIVOS
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="PERIODOS_COMPARATIVOS").font = section_font
    row += 1
    
    headers = ["ID Periodo", "Nombre"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    comparativos = [
        ("202503", "Mar 2025"),
        ("202506", "Jun 2025"),
        ("202511", "Nov 2025"),
    ]
    for id_p, nombre in comparativos:
        ws.cell(row=row, column=1, value=id_p)
        ws.cell(row=row, column=2, value=nombre)
        row += 1
    
    row += 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: BALANCE CLASIFICADO
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="BALANCE_CLASIFICADO").font = section_font
    row += 1
    
    headers = ["Categoría", "Nombre Mostrar", "Prefijos", "Excluir Cuentas", "Cuentas Específicas", "Descripción"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    balance_data = [
        ("activo_corriente", "Activo Corriente", "11", "1109009", "", ""),
        ("activo_no_corriente", "Activo No Corriente", "12", "1301001", "", ""),
        ("intangibles", "Intangibles", "", "", "1109009,1301001", "Depósito + Inversiones"),
        ("pasivo_corriente", "Pasivo Corriente", "21", "", "", ""),
        ("pasivo_no_corriente", "Pasivo No Corriente", "22", "", "", ""),
        ("patrimonio", "Patrimonio", "31", "", "", ""),
    ]
    for cat, nombre, pref, excl, especif, desc in balance_data:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=nombre)
        ws.cell(row=row, column=3, value=pref)
        ws.cell(row=row, column=4, value=excl)
        ws.cell(row=row, column=5, value=especif)
        ws.cell(row=row, column=6, value=desc)
        row += 1
    
    row += 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: ESTADO DE RESULTADOS
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="ESTADO_RESULTADOS").font = section_font
    row += 1
    
    headers = ["Tipo", "Key", "Nombre", "Cuentas", "Descripción"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    eerr_data = [
        ("ingresos", "ingresos", "Ingresos por Ventas", "4101001", "Ventas Del Giro"),
        ("costo_ventas", "costo_ventas", "Costo de Ventas", "5101001", "Costo De Venta"),
        ("gastos_operacionales", "remuneraciones", "Remuneraciones y Salarios", "5201001,5201999", "Remuneraciones + Otras"),
        ("gastos_operacionales", "honorarios", "Honorarios Profesionales", "5205002", "Honorarios"),
        ("gastos_operacionales", "arriendo", "Arriendo", "5304001", "Arriendo Inmueble"),
        ("gastos_operacionales", "marketing", "Marketing y Publicidad", "5307001,5307002", "Publicidad + Medios"),
        ("gastos_operacionales", "legales", "Gastos Legales y Contables", "5308001,5308002,5308003", "Notariales + Legal + Contable"),
        ("gastos_operacionales", "otros", "Otros Gastos Operativos", "5309016,5310999", "Suscripciones + Generales"),
        ("otros_gastos", "multas", "Multas y Sanciones", "5314005", "Multas Al Fisco"),
        ("otros_gastos", "financieros", "Gastos Financieros", "5403001,5403003,5403005", "Intereses + Comisiones + Factoring"),
    ]
    for tipo, key, nombre, cuentas, desc in eerr_data:
        ws.cell(row=row, column=1, value=tipo)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value=nombre)
        ws.cell(row=row, column=4, value=cuentas)
        ws.cell(row=row, column=5, value=desc)
        row += 1
    
    row += 1
    
    # ─────────────────────────────────────────────────────────
    # SECCIÓN: OUTPUT
    # ─────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="OUTPUT").font = section_font
    row += 1
    
    headers = ["Campo", "Valor"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    output_data = [
        ("carpeta", "generados"),
        ("prefijo_archivo", "Balance_PorCuenta"),
    ]
    for campo, valor in output_data:
        ws.cell(row=row, column=1, value=campo)
        ws.cell(row=row, column=2, value=valor)
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    
    # ═══════════════════════════════════════════════════════════════
    # HOJA INSTRUCCIONES
    # ═══════════════════════════════════════════════════════════════
    ws_inst = wb.create_sheet("_INSTRUCCIONES", 0)
    
    instrucciones = [
        ("INSTRUCCIONES DE USO", ""),
        ("", ""),
        ("1. AGREGAR NUEVA EMPRESA", ""),
        ("   - Clic derecho en la pestaña FIDI → 'Mover o copiar'", ""),
        ("   - Marcar 'Crear una copia'", ""),
        ("   - Renombrar la hoja con el KEY de la empresa (ej: CISI)", ""),
        ("   - Modificar los datos según corresponda", ""),
        ("", ""),
        ("2. EJECUTAR", ""),
        ("   python3 balance_excel_v2.py FIDI", ""),
        ("   python3 balance_excel_v2.py CISI", ""),
        ("", ""),
        ("3. SECCIONES", ""),
        ("   TENANT: Datos de la empresa", ""),
        ("   PERIODOS: Período actual y tasa de impuesto", ""),
        ("   PERIODOS_COMPARATIVOS: Para hoja de EEFF Comparativos", ""),
        ("   BALANCE_CLASIFICADO: Cómo agrupar cuentas en el balance", ""),
        ("   ESTADO_RESULTADOS: Cómo agrupar cuentas en el EERR", ""),
        ("   OUTPUT: Carpeta y prefijo del archivo generado", ""),
        ("", ""),
        ("4. FORMATO DE CUENTAS", ""),
        ("   - Separar múltiples cuentas con coma: 5201001,5201999", ""),
        ("   - Prefijos: usar solo el inicio del código (ej: 11 para 11xxxxx)", ""),
        ("", ""),
        ("5. NOTAS", ""),
        ("   - No cambiar los nombres de las secciones (TENANT, PERIODOS, etc)", ""),
        ("   - No cambiar el orden de las columnas", ""),
        ("   - Pueden agregar más filas en cada sección", ""),
    ]
    
    for row_idx, (col1, col2) in enumerate(instrucciones, 1):
        ws_inst.cell(row=row_idx, column=1, value=col1)
        if row_idx == 1:
            ws_inst.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color="006400")
    
    ws_inst.column_dimensions['A'].width = 60
    
    # Guardar
    wb.save("config/empresas_config.xlsx")
    print("✅ Archivo creado: config/empresas_config.xlsx")


if __name__ == "__main__":
    crear_config_excel()


