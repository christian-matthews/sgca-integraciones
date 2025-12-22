#!/usr/bin/env python3
"""
Balance Tributario + Estado de Resultados a Excel - Odoo/FactorIT
==================================================================

Genera un Excel con:
- Resumen: Balance Clasificado + Estado de Resultados + KPIs
- Balance: Todas las cuentas con saldos
- Hojas por cuenta: Detalle de movimientos con hipervínculos

IMPORTANTE:
- Incluye Resultado del Período en Patrimonio
- Verifica cuadratura: Activos = Pasivos + Patrimonio

Uso:
    python -m odoo.balance_excel FactorIT
    python -m odoo.balance_excel FactorIT2
"""

import os
import sys
from datetime import datetime
from dotenv import load_dotenv
import psycopg2
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()


def get_env_clean(key, default=None):
    val = os.getenv(key)
    return val.strip() if val else default


DB_CONFIG = {
    'host': get_env_clean('SERVER'),
    'port': get_env_clean('PORT', '5432'),
    'user': get_env_clean('DB_USER'),
    'password': get_env_clean('PASSWORD'),
}

DATABASES = {
    'FactorIT': 'FactorIT SpA',
    'FactorIT2': 'FactorIT Ltda',
}

# Estilos
ESTILOS = {
    "font_header": Font(bold=True, size=10, color="FFFFFF"),
    "fill_header": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "font_titulo": Font(bold=True, size=14, color="1F4E79"),
    "font_section": Font(bold=True, size=12, color="1F4E79"),
    "fill_section": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
    "font_category": Font(bold=True, size=11),
    "font_subtotal": Font(bold=True, size=10),
    "font_total": Font(bold=True, size=11),
    "font_total_final": Font(bold=True, size=12, color="FFFFFF"),
    "fill_total_final": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "fill_subtotal": PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid"),
    "thin_border": Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    ),
    "formato_miles": '#,##0'
}

# Clasificación de cuentas por prefijo de código
CLASIFICACION_CUENTAS = {
    "activo_corriente": {"nombre": "Activo Corriente", "prefijos": ["11"]},
    "activo_no_corriente": {"nombre": "Activo No Corriente", "prefijos": ["12", "13", "14"]},
    "pasivo_corriente": {"nombre": "Pasivo Corriente", "prefijos": ["21"]},
    "pasivo_no_corriente": {"nombre": "Pasivo No Corriente", "prefijos": ["22", "23"]},
    "patrimonio": {"nombre": "Patrimonio", "prefijos": ["3"]},
    "ingresos": {"nombre": "Ingresos Operacionales", "prefijos": ["4"]},
    "costos": {"nombre": "Costos de Venta", "prefijos": ["51"]},
    "gastos_operacionales": {"nombre": "Gastos Operacionales", "prefijos": ["52", "53", "54", "55"]},
    "otros_ingresos": {"nombre": "Otros Ingresos", "prefijos": ["6"]},  # No operacionales
    "otros_gastos": {"nombre": "Otros Gastos", "prefijos": ["7"]},      # No operacionales
    "apertura": {"nombre": "Saldos Apertura", "prefijos": ["8"]},       # Ajustes de apertura
}


def clasificar_cuenta(codigo):
    """Clasifica una cuenta según su código."""
    for categoria, config in CLASIFICACION_CUENTAS.items():
        for prefijo in config["prefijos"]:
            if codigo.startswith(prefijo):
                return categoria
    return None


def sanitize_sheet_name(codigo, nombre):
    """Limpia nombre para hoja Excel (max 31 chars)."""
    name = f"{codigo} {nombre}"
    for char in ['\\', '/', '*', '?', '[', ']', ':']:
        name = name.replace(char, '')
    return name[:31]


def obtener_balance(cursor, fecha_hasta=None):
    """Obtiene el balance de saldos por cuenta."""
    
    if fecha_hasta is None:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')
    
    query = """
    SELECT 
        aa.code as codigo,
        aa.name as cuenta,
        aa.user_type_id as tipo,
        SUM(aml.debit) as debe,
        SUM(aml.credit) as haber,
        SUM(aml.debit) - SUM(aml.credit) as saldo
    FROM account_move_line aml
    JOIN account_account aa ON aml.account_id = aa.id
    JOIN account_move am ON aml.move_id = am.id
    WHERE am.state = 'posted'
      AND aml.date <= %s
    GROUP BY aa.id, aa.code, aa.name, aa.user_type_id
    HAVING SUM(aml.debit) != 0 OR SUM(aml.credit) != 0
    ORDER BY aa.code
    """
    
    cursor.execute(query, (fecha_hasta,))
    
    cuentas = []
    for row in cursor.fetchall():
        codigo, nombre, tipo, debe, haber, saldo = row
        debe = float(debe or 0)
        haber = float(haber or 0)
        saldo = float(saldo or 0)
        
        # Determinar clasificación
        clasificacion = clasificar_cuenta(codigo)
        
        # Lógica de signos según tipo de cuenta:
        # - Activos (1): saldo deudor (+) es positivo
        # - Pasivos (2): saldo acreedor (-) se invierte a positivo
        # - Patrimonio (3): saldo acreedor (-) se invierte a positivo
        # - Ingresos (4): saldo acreedor (-) se invierte a positivo
        # - Gastos (5): saldo deudor (+) es positivo
        # - Otros Ingresos (6): saldo acreedor (-) se invierte a positivo
        # - Otros Gastos (7): saldo deudor (+) es positivo
        # - Apertura (8): saldo acreedor (-) se invierte a positivo
        
        if clasificacion in ["activo_corriente", "activo_no_corriente"]:
            valor_balance = saldo  # Deudor positivo
        elif clasificacion in ["pasivo_corriente", "pasivo_no_corriente", "patrimonio"]:
            valor_balance = -saldo  # Acreedor: invertir signo
        elif clasificacion in ["ingresos", "otros_ingresos"]:
            valor_balance = -saldo  # Acreedor: invertir signo
        elif clasificacion in ["costos", "gastos_operacionales", "otros_gastos"]:
            valor_balance = saldo  # Deudor positivo
        elif clasificacion == "apertura":
            valor_balance = -saldo  # Acreedor: invertir signo
        else:
            valor_balance = saldo
        
        cuentas.append({
            'codigo': codigo,
            'cuenta': nombre,
            'tipo': tipo,
            'debe': debe,
            'haber': haber,
            'saldo': saldo,
            'clasificacion': clasificacion,
            'valor_balance': valor_balance
        })
    
    return cuentas


def obtener_movimientos_cuenta(cursor, codigo_cuenta, fecha_hasta=None):
    """Obtiene los movimientos de una cuenta específica."""
    
    if fecha_hasta is None:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')
    
    query = """
    SELECT 
        aml.date as fecha,
        am.name as asiento,
        aml.name as descripcion,
        rp.name as tercero,
        aml.debit as debe,
        aml.credit as haber
    FROM account_move_line aml
    JOIN account_account aa ON aml.account_id = aa.id
    JOIN account_move am ON aml.move_id = am.id
    LEFT JOIN res_partner rp ON aml.partner_id = rp.id
    WHERE aa.code = %s
      AND am.state = 'posted'
      AND aml.date <= %s
    ORDER BY aml.date DESC, am.name
    """
    
    cursor.execute(query, (codigo_cuenta, fecha_hasta))
    return cursor.fetchall()


def generar_balance_excel(db_name, fecha_hasta=None):
    """Genera el Excel de Balance para una empresa."""
    
    empresa_nombre = DATABASES.get(db_name, db_name)
    
    if fecha_hasta is None:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')
    
    print("=" * 70)
    print(f"   GENERANDO BALANCE - {empresa_nombre}")
    print("=" * 70)
    print(f"   Fecha de corte: {fecha_hasta}")
    
    # Conectar
    conn = psycopg2.connect(
        host=DB_CONFIG['host'],
        port=DB_CONFIG['port'],
        database=db_name,
        user=DB_CONFIG['user'],
        password=DB_CONFIG['password'],
    )
    cursor = conn.cursor()
    
    # Obtener balance
    print("\n📊 Obteniendo balance...")
    cuentas = obtener_balance(cursor, fecha_hasta)
    print(f"   {len(cuentas)} cuentas con movimientos")
    
    # Clasificar cuentas
    categorias = {}
    for cat in CLASIFICACION_CUENTAS.keys():
        categorias[cat] = []
    
    for c in cuentas:
        if c['clasificacion'] and c['valor_balance'] != 0:
            categorias[c['clasificacion']].append(c)
    
    # Calcular totales
    total_activo_corriente = sum(c['valor_balance'] for c in categorias.get('activo_corriente', []))
    total_activo_no_corriente = sum(c['valor_balance'] for c in categorias.get('activo_no_corriente', []))
    total_activos = total_activo_corriente + total_activo_no_corriente
    
    total_pasivo_corriente = sum(c['valor_balance'] for c in categorias.get('pasivo_corriente', []))
    total_pasivo_no_corriente = sum(c['valor_balance'] for c in categorias.get('pasivo_no_corriente', []))
    total_pasivos = total_pasivo_corriente + total_pasivo_no_corriente
    
    patrimonio_sin_resultado = sum(c['valor_balance'] for c in categorias.get('patrimonio', []))
    
    # Estado de Resultados
    # Los valores ya están con signo correcto (positivo = favorable)
    ingresos = sum(c['valor_balance'] for c in categorias.get('ingresos', []))
    costos = sum(c['valor_balance'] for c in categorias.get('costos', []))
    utilidad_bruta = ingresos - costos
    
    gastos_op = sum(c['valor_balance'] for c in categorias.get('gastos_operacionales', []))
    resultado_operacional = utilidad_bruta - gastos_op
    
    # Otros ingresos y gastos no operacionales
    otros_ingresos = sum(c['valor_balance'] for c in categorias.get('otros_ingresos', []))
    otros_gastos = sum(c['valor_balance'] for c in categorias.get('otros_gastos', []))
    resultado_no_operacional = otros_ingresos - otros_gastos
    
    resultado_antes_impuestos = resultado_operacional + resultado_no_operacional
    
    # Ajustes de apertura (van al patrimonio, no afectan resultado)
    ajustes_apertura = sum(c['valor_balance'] for c in categorias.get('apertura', []))
    
    # El resultado neto es el resultado antes de impuestos
    # (el impuesto ya está incluido en otros_gastos como "Impuesto a la Renta")
    resultado_neto = resultado_antes_impuestos
    
    # Para mostrar el impuesto separado, lo buscamos
    impuesto_contabilizado = sum(c['valor_balance'] for c in categorias.get('otros_gastos', []) 
                                  if 'impuesto' in c['cuenta'].lower() and 'renta' in c['cuenta'].lower())
    
    # Patrimonio total (incluye resultado del período y ajustes de apertura)
    total_patrimonio = patrimonio_sin_resultado + resultado_neto + ajustes_apertura
    
    # Verificación de cuadratura
    diferencia = total_activos - (total_pasivos + total_patrimonio)
    cuadra = abs(diferencia) < 1
    
    print(f"\n📋 Resumen:")
    print(f"   Total Activos:      ${total_activos:>18,.0f}")
    print(f"   Total Pasivos:      ${total_pasivos:>18,.0f}")
    print(f"   Patrimonio base:    ${patrimonio_sin_resultado:>18,.0f}")
    print(f"   Resultado Período:  ${resultado_neto:>18,.0f}")
    print(f"   Ajustes Apertura:   ${ajustes_apertura:>18,.0f}")
    print(f"   Total Patrimonio:   ${total_patrimonio:>18,.0f}")
    print(f"   Pasivos+Patrimonio: ${total_pasivos + total_patrimonio:>18,.0f}")
    print(f"   {'✅ CUADRA' if cuadra else f'⚠️ DIFERENCIA: ${diferencia:,.0f}'}")
    
    # Crear Excel
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'generados')
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    periodo = fecha_hasta.replace('-', '')[:6]
    filename = os.path.join(output_dir, f'Balance_Odoo_{db_name}_{periodo}_{timestamp}.xlsx')
    
    print(f"\n📝 Generando Excel...")
    
    cuenta_a_hoja = {}
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # ═══════════════════════════════════════════════════════════════════
        # HOJA 1: RESUMEN
        # ═══════════════════════════════════════════════════════════════════
        rows = []
        row_types = []
        
        # Título
        rows.append([f"BALANCE GENERAL - {empresa_nombre}", "", "", "", f"Fecha: {fecha_hasta}"])
        row_types.append("titulo")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # BALANCE CLASIFICADO
        rows.append(["BALANCE CLASIFICADO", "", "", "", ""])
        row_types.append("section")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # ACTIVOS
        rows.append(["ACTIVOS", "", "", "", ""])
        row_types.append("category")
        
        rows.append(["  Activo Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('activo_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Activo Corriente", total_activo_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["  Activo No Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('activo_no_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Activo No Corriente", total_activo_no_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["TOTAL ACTIVOS", total_activos, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # PASIVOS
        rows.append(["PASIVOS", "", "", "", ""])
        row_types.append("category")
        
        rows.append(["  Pasivo Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('pasivo_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Pasivo Corriente", total_pasivo_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["  Pasivo No Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('pasivo_no_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Pasivo No Corriente", total_pasivo_no_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["TOTAL PASIVOS", total_pasivos, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # PATRIMONIO
        rows.append(["PATRIMONIO", "", "", "", ""])
        row_types.append("category")
        for c in categorias.get('patrimonio', []):
            rows.append([f"  {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Resultado del Período", resultado_neto, "", "", "(calculado)"])
        row_types.append("item")
        if ajustes_apertura != 0:
            rows.append(["  Ajustes de Apertura", ajustes_apertura, "", "", ""])
            row_types.append("item")
        rows.append(["TOTAL PATRIMONIO", total_patrimonio, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["TOTAL PASIVOS + PATRIMONIO", total_pasivos + total_patrimonio, "", "", ""])
        row_types.append("total_final")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Verificación
        if cuadra:
            rows.append(["✅ CUADRATURA OK: Activos = Pasivos + Patrimonio", "", "", "", ""])
            row_types.append("verification_ok")
        else:
            rows.append([f"⚠️ DESCUADRE: Diferencia = ${diferencia:,.0f}", "", "", "", ""])
            row_types.append("verification_error")
        
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # ESTADO DE RESULTADOS
        rows.append(["ESTADO DE RESULTADOS", "", "", "", ""])
        row_types.append("section")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Ingresos Operacionales", ingresos, "", "", ""])
        row_types.append("item")
        rows.append(["Costo de Ventas", -costos, "", "", ""])
        row_types.append("item")
        rows.append(["UTILIDAD BRUTA", utilidad_bruta, "", "", ""])
        row_types.append("subtotal")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Gastos Operacionales", -gastos_op, "", "", ""])
        row_types.append("item")
        rows.append(["RESULTADO OPERACIONAL", resultado_operacional, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Otros Ingresos No Operacionales", otros_ingresos, "", "", ""])
        row_types.append("item")
        rows.append(["Otros Gastos No Operacionales", -otros_gastos, "", "", "(incluye impuestos)"])
        row_types.append("item")
        if impuesto_contabilizado > 0:
            rows.append([f"  (Impuesto Renta incluido)", -impuesto_contabilizado, "", "", ""])
            row_types.append("item")
        rows.append(["RESULTADO NETO", resultado_neto, "", "", ""])
        row_types.append("total_final")
        
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # KPIs
        rows.append(["INDICADORES FINANCIEROS", "", "", "", ""])
        row_types.append("section")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        margen_bruto = (utilidad_bruta / ingresos * 100) if ingresos else 0
        margen_op = (resultado_operacional / ingresos * 100) if ingresos else 0
        margen_neto = (resultado_neto / ingresos * 100) if ingresos else 0
        roa = (resultado_neto / total_activos * 100) if total_activos else 0
        roe = (resultado_neto / total_patrimonio * 100) if total_patrimonio else 0
        
        rows.append(["Margen Bruto", f"{margen_bruto:.1f}%", "Utilidad Bruta / Ingresos", "", ""])
        row_types.append("kpi")
        rows.append(["Margen Operacional", f"{margen_op:.1f}%", "Resultado Op / Ingresos", "", ""])
        row_types.append("kpi")
        rows.append(["Margen Neto", f"{margen_neto:.1f}%", "Resultado Neto / Ingresos", "", ""])
        row_types.append("kpi")
        rows.append(["ROA", f"{roa:.1f}%", "Resultado / Activos", "", ""])
        row_types.append("kpi")
        rows.append(["ROE", f"{roe:.1f}%", "Resultado / Patrimonio", "", ""])
        row_types.append("kpi")
        
        # Escribir hoja Resumen
        df_resumen = pd.DataFrame(rows)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
        ws_resumen = writer.sheets['Resumen']
        
        # Aplicar formatos
        for row_idx, row_type in enumerate(row_types, start=1):
            cell_a = ws_resumen.cell(row=row_idx, column=1)
            cell_b = ws_resumen.cell(row=row_idx, column=2)
            
            if row_type == "titulo":
                cell_a.font = ESTILOS["font_titulo"]
            elif row_type == "section":
                cell_a.font = ESTILOS["font_section"]
                cell_a.fill = ESTILOS["fill_section"]
            elif row_type == "category":
                cell_a.font = ESTILOS["font_category"]
            elif row_type == "subcategory":
                cell_a.font = Font(bold=True, italic=True)
            elif row_type == "subtotal":
                cell_a.font = ESTILOS["font_subtotal"]
                cell_b.font = ESTILOS["font_subtotal"]
                cell_a.fill = ESTILOS["fill_subtotal"]
                cell_b.fill = ESTILOS["fill_subtotal"]
            elif row_type == "total":
                cell_a.font = ESTILOS["font_total"]
                cell_b.font = ESTILOS["font_total"]
            elif row_type == "total_final":
                cell_a.font = ESTILOS["font_total_final"]
                cell_a.fill = ESTILOS["fill_total_final"]
                cell_b.font = ESTILOS["font_total_final"]
                cell_b.fill = ESTILOS["fill_total_final"]
            elif row_type == "verification_ok":
                cell_a.font = Font(bold=True, color="006400")
            elif row_type == "verification_error":
                cell_a.font = Font(bold=True, color="FF0000")
            elif row_type == "kpi":
                cell_b.font = Font(bold=True, color="1F4E79")
            
            # Formato numérico
            if isinstance(rows[row_idx-1][1], (int, float)):
                cell_b.number_format = ESTILOS["formato_miles"]
        
        # Anchos de columna
        ws_resumen.column_dimensions['A'].width = 45
        ws_resumen.column_dimensions['B'].width = 18
        ws_resumen.column_dimensions['C'].width = 30
        ws_resumen.column_dimensions['E'].width = 20
        
        # ═══════════════════════════════════════════════════════════════════
        # HOJA 2: BALANCE DETALLADO
        # ═══════════════════════════════════════════════════════════════════
        df_balance = pd.DataFrame(cuentas)
        df_balance = df_balance[['codigo', 'cuenta', 'debe', 'haber', 'saldo', 'clasificacion']]
        df_balance.columns = ['Código', 'Cuenta', 'Debe', 'Haber', 'Saldo', 'Clasificación']
        df_balance['Ver Detalle'] = ''
        
        df_balance.to_excel(writer, sheet_name='Balance', index=False, startrow=1)
        
        ws_balance = writer.sheets['Balance']
        ws_balance.cell(row=1, column=1).value = f"BALANCE DETALLADO - {empresa_nombre} - {fecha_hasta}"
        ws_balance.cell(row=1, column=1).font = ESTILOS["font_titulo"]
        
        # Formato encabezados
        for col in range(1, 8):
            cell = ws_balance.cell(row=2, column=col)
            cell.font = ESTILOS["font_header"]
            cell.fill = ESTILOS["fill_header"]
        
        # ═══════════════════════════════════════════════════════════════════
        # HOJAS DE DETALLE POR CUENTA
        # ═══════════════════════════════════════════════════════════════════
        print("   Generando hojas de detalle...")
        
        cuentas_con_detalle = 0
        for c in cuentas:
            if abs(c['saldo']) < 1:
                continue
            
            movimientos = obtener_movimientos_cuenta(cursor, c['codigo'], fecha_hasta)
            
            if not movimientos:
                continue
            
            # Crear hoja
            sheet_name = sanitize_sheet_name(c['codigo'], c['cuenta'])
            
            df_mov = pd.DataFrame(movimientos, columns=['Fecha', 'Asiento', 'Descripción', 'Tercero', 'Debe', 'Haber'])
            df_mov['Fecha'] = df_mov['Fecha'].astype(str)
            
            # Calcular saldo acumulado
            saldo_acum = []
            acum = 0
            for _, row in df_mov.iterrows():
                acum += float(row['Debe'] or 0) - float(row['Haber'] or 0)
                saldo_acum.append(acum)
            df_mov['Saldo'] = saldo_acum
            
            df_mov.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            
            ws = writer.sheets[sheet_name]
            ws.cell(row=1, column=1).value = "← Volver al Balance"
            ws.cell(row=1, column=1).hyperlink = "#'Balance'!A1"
            ws.cell(row=1, column=1).style = "Hyperlink"
            ws.cell(row=2, column=1).value = f"{c['codigo']} - {c['cuenta']}"
            ws.cell(row=2, column=1).font = ESTILOS["font_titulo"]
            
            cuenta_a_hoja[c['codigo']] = sheet_name
            cuentas_con_detalle += 1
        
        print(f"   {cuentas_con_detalle} cuentas con detalle")
        
        # Agregar hipervínculos en Balance
        for row_idx, c in enumerate(cuentas, start=3):
            if c['codigo'] in cuenta_a_hoja:
                cell = ws_balance.cell(row=row_idx, column=7)
                cell.value = "→ Ver"
                cell.hyperlink = f"#'{cuenta_a_hoja[c['codigo']]}'!A1"
                cell.style = "Hyperlink"
    
    cursor.close()
    conn.close()
    
    print(f"\n✅ Archivo generado: {filename}")
    print("=" * 70)
    
    return filename


def main():
    if len(sys.argv) < 2:
        print("Uso: python -m odoo.balance_excel <DATABASE> [FECHA]")
        print("Ejemplo: python -m odoo.balance_excel FactorIT")
        print("         python -m odoo.balance_excel FactorIT2 2025-11-30")
        print(f"\nBases disponibles: {', '.join(DATABASES.keys())}")
        sys.exit(1)
    
    db_name = sys.argv[1]
    fecha = sys.argv[2] if len(sys.argv) > 2 else None
    
    if db_name not in DATABASES:
        print(f"❌ Base de datos '{db_name}' no encontrada")
        print(f"   Disponibles: {', '.join(DATABASES.keys())}")
        sys.exit(1)
    
    generar_balance_excel(db_name, fecha)


if __name__ == '__main__':
    main()

