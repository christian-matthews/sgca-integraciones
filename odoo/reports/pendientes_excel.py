#!/usr/bin/env python3
"""
ART-002: Reporte de Pendientes - Implementación Odoo
=====================================================

Genera reporte Excel con pendientes operativos según especificación ART-002.

Uso:
    python pendientes_excel.py FactorIT
    python pendientes_excel.py FactorIT2
    python pendientes_excel.py  # Todas las empresas
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Agregar path para importar pendientes
sys.path.insert(0, str(Path(__file__).parent.parent))
from pendientes import obtener_pendientes_empresa, DATABASES

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN - PALETA ART-002 (igual que Balance)
# ═══════════════════════════════════════════════════════════════════════════════

COLORS = {
    "header_bg": "006400",      # Verde oscuro
    "header_text": "FFFFFF",    # Blanco
    "section_bg": "E8F5E9",     # Verde muy claro
    "subtotal_bg": "C8E6C9",    # Verde claro
    "alert_bg": "FFCDD2",       # Rojo claro
    "warning_bg": "FFF9C4",     # Amarillo claro
    "border": "CCCCCC",         # Gris claro
}

OUTPUT_DIR = Path(__file__).parent / "generados"
OUTPUT_DIR.mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════════

def get_styles():
    """Retorna diccionario de estilos reutilizables."""
    thin_border = Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    )
    
    return {
        "title": Font(bold=True, size=14, color=COLORS["header_bg"]),
        "subtitle": Font(bold=True, size=12, color=COLORS["header_bg"]),
        "header": Font(bold=True, size=10, color=COLORS["header_text"]),
        "header_fill": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
        "section": Font(bold=True, size=11),
        "section_fill": PatternFill(start_color=COLORS["section_bg"], end_color=COLORS["section_bg"], fill_type="solid"),
        "subtotal": Font(bold=True, size=10),
        "subtotal_fill": PatternFill(start_color=COLORS["subtotal_bg"], end_color=COLORS["subtotal_bg"], fill_type="solid"),
        "alert_fill": PatternFill(start_color=COLORS["alert_bg"], end_color=COLORS["alert_bg"], fill_type="solid"),
        "warning_fill": PatternFill(start_color=COLORS["warning_bg"], end_color=COLORS["warning_bg"], fill_type="solid"),
        "hyperlink": Font(color="0000FF", underline="single"),
        "border": thin_border,
        "number_format": '#,##0',
    }


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: RESUMEN PENDIENTES
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_resumen(writer, data: dict, empresa: str, fecha: str):
    """Crea la hoja Resumen Pendientes."""
    
    styles = get_styles()
    
    # Obtener datos - Odoo estructura ligeramente diferente
    sii = data.get("pendientes_sii", {})
    contab = data.get("pendientes_contabilizar", {})
    concil = data.get("pendientes_conciliar", {})
    
    # SII puede tener estructura con accionables/tacitos
    if "accionables" in sii:
        sii_docs = sii.get("accionables", {}).get("documentos", [])
        sii_cantidad = sii.get("accionables", {}).get("cantidad", 0)
        sii_total = sii.get("accionables", {}).get("total", 0)
    else:
        sii_docs = sii.get("documentos", [])
        sii_cantidad = sii.get("cantidad", 0)
        sii_total = sii.get("total", 0)
    
    contab_docs = contab.get("documentos", [])
    concil_movs = concil.get("movimientos", [])
    
    # Calcular stats
    hoy = datetime.now().date()
    
    def calcular_stats_fecha(items, campo_fecha):
        if not items:
            return 0, None, 0
        dias_list = []
        for i in items:
            fecha_str = i.get(campo_fecha)
            if fecha_str:
                try:
                    f = datetime.strptime(str(fecha_str)[:10], "%Y-%m-%d").date()
                    dias = (hoy - f).days
                    dias_list.append(dias)
                    i["dias_pendiente"] = dias
                except:
                    i["dias_pendiente"] = 0
            else:
                i["dias_pendiente"] = 0
        
        if not dias_list:
            return 0, None, 0
        promedio = sum(dias_list) / len(dias_list)
        max_dias = max(dias_list)
        mas_antiguo = None
        for i in items:
            if i.get("dias_pendiente") == max_dias:
                mas_antiguo = i.get(campo_fecha)
                break
        return promedio, mas_antiguo, max_dias
    
    # Stats SII
    sii_prom = sii_max = 0
    sii_antiguo = None
    if sii_docs:
        for doc in sii_docs:
            dias = doc.get("dias_documento", 0)
            doc["dias_pendiente"] = dias
        dias_list = [d.get("dias_pendiente", 0) for d in sii_docs]
        sii_prom = sum(dias_list) / len(dias_list) if dias_list else 0
        sii_max = max(dias_list) if dias_list else 0
        for d in sii_docs:
            if d.get("dias_pendiente") == sii_max:
                sii_antiguo = d.get("fecha")
                break
    
    contab_prom, contab_antiguo, contab_max = calcular_stats_fecha(contab_docs, "fecha")
    concil_prom, concil_antiguo, concil_max = calcular_stats_fecha(concil_movs, "fecha")
    
    # Construir datos
    rows = [
        ["← Navegación", "", "", "", "", ""],
        [f"RESUMEN DE PENDIENTES - {empresa}", "", "", "", "", ""],
        [f"Fecha consulta: {fecha}", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["Categoría", "Cantidad", "Monto Total", "Días Promedio", "Más Antiguo", "Ver Detalle"],
        [
            "Pendientes SII",
            sii_cantidad,
            sii_total,
            round(sii_prom, 1),
            f"{sii_antiguo} ({sii_max} días)" if sii_antiguo else "-",
            "→ Ver"
        ],
        [
            "Pendientes Contabilizar",
            contab.get("cantidad", 0),
            sum(d.get("monto", 0) or 0 for d in contab_docs),
            round(contab_prom, 1),
            f"{str(contab_antiguo)[:10]} ({contab_max} días)" if contab_antiguo else "-",
            "→ Ver"
        ],
        [
            "Pendientes Conciliar",
            concil.get("cantidad", 0),
            abs(concil.get("total_cargos", 0)) + abs(concil.get("total_abonos", 0)),
            round(concil_prom, 1),
            f"{str(concil_antiguo)[:10]} ({concil_max} días)" if concil_antiguo else "-",
            "→ Ver"
        ],
        ["", "", "", "", "", ""],
        ["", "Totales:", "", "", "", ""],
        ["", "Total Cargos (Conciliar):", abs(concil.get("total_cargos", 0)), "", "", ""],
        ["", "Total Abonos (Conciliar):", abs(concil.get("total_abonos", 0)), "", "", ""],
    ]
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Resumen Pendientes", index=False, header=False)
    
    ws = writer.sheets["Resumen Pendientes"]
    
    # Aplicar estilos
    ws.cell(row=2, column=1).font = styles["title"]
    ws.cell(row=3, column=1).font = Font(italic=True, size=10)
    
    for col in range(1, 7):
        cell = ws.cell(row=5, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    for row in range(6, 9):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            if col in [2, 3, 4]:
                cell.number_format = styles["number_format"]
                cell.alignment = Alignment(horizontal='right')
    
    # Hipervínculos
    ws.cell(row=6, column=6).hyperlink = "#'Pendientes SII'!A1"
    ws.cell(row=6, column=6).font = styles["hyperlink"]
    ws.cell(row=7, column=6).hyperlink = "#'Pendientes Contabilizar'!A1"
    ws.cell(row=7, column=6).font = styles["hyperlink"]
    ws.cell(row=8, column=6).hyperlink = "#'Pendientes Conciliar'!A1"
    ws.cell(row=8, column=6).font = styles["hyperlink"]
    
    ws.cell(row=10, column=2).font = Font(bold=True)
    ws.cell(row=11, column=3).number_format = styles["number_format"]
    ws.cell(row=12, column=3).number_format = styles["number_format"]
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES SII
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_sii(writer, data: dict):
    """Crea la hoja Pendientes SII."""
    
    styles = get_styles()
    sii = data.get("pendientes_sii", {})
    
    # Odoo puede tener estructura con accionables/tacitos
    if "accionables" in sii:
        docs = sii.get("accionables", {}).get("documentos", [])
    else:
        docs = sii.get("documentos", [])
    
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES SII - Documentos por Aceptar/Rechazar"],
        [""],
        ["Fecha", "Tipo", "Folio", "Proveedor RUT", "Proveedor", "Monto", "Días", "Estado"],
    ]
    
    for doc in docs:
        dias = doc.get("dias_documento", doc.get("dias_pendiente", 0)) or 0
        estado = "Accionable" if dias <= 8 else "Tácito"
        rows.append([
            str(doc.get("fecha", ""))[:10],
            doc.get("tipo", doc.get("tipo_documento", "")),
            doc.get("folio", doc.get("number", "")),
            doc.get("proveedor_rut", doc.get("rut_emisor", "")),
            doc.get("proveedor", doc.get("nombre_emisor", "")),
            doc.get("monto", doc.get("amount_total", 0)),
            dias,
            estado
        ])
    
    if not docs:
        rows.append(["", "Sin pendientes", "", "", "", "", "", ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes SII", index=False, header=False)
    
    ws = writer.sheets["Pendientes SII"]
    
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    for row_idx, doc in enumerate(docs, start=5):
        dias = doc.get("dias_documento", doc.get("dias_pendiente", 0)) or 0
        
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            if dias >= 6:
                cell.fill = styles["alert_fill"]
            elif dias >= 4:
                cell.fill = styles["warning_fill"]
            
            if col == 6:
                cell.number_format = styles["number_format"]
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES CONTABILIZAR
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_contabilizar(writer, data: dict):
    """Crea la hoja Pendientes Contabilizar."""
    
    styles = get_styles()
    contab = data.get("pendientes_contabilizar", {})
    docs = contab.get("documentos", [])
    
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES CONTABILIZAR - Documentos Aceptados sin Contabilizar"],
        [""],
        ["Fecha", "Tipo", "Folio", "Proveedor RUT", "Proveedor", "Monto", "Días Pendiente"],
    ]
    
    for doc in docs:
        rows.append([
            str(doc.get("fecha", ""))[:10],
            doc.get("tipo", doc.get("tipo_documento", "")),
            doc.get("folio", doc.get("number", "")),
            doc.get("proveedor_rut", doc.get("rut_emisor", "")),
            doc.get("proveedor", doc.get("nombre_emisor", "")),
            doc.get("monto", doc.get("amount_total", 0)),
            doc.get("dias_pendiente", 0)
        ])
    
    if not docs:
        rows.append(["", "Sin pendientes", "", "", "", "", ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes Contabilizar", index=False, header=False)
    
    ws = writer.sheets["Pendientes Contabilizar"]
    
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    for col in range(1, 8):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    for row_idx, doc in enumerate(docs, start=5):
        dias = doc.get("dias_pendiente", 0)
        
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            if dias >= 30:
                cell.fill = styles["alert_fill"]
            elif dias >= 15:
                cell.fill = styles["warning_fill"]
            
            if col == 6:
                cell.number_format = styles["number_format"]
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES CONCILIAR
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_conciliar(writer, data: dict):
    """Crea la hoja Pendientes Conciliar."""
    
    styles = get_styles()
    concil = data.get("pendientes_conciliar", {})
    movs = concil.get("movimientos", [])
    
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES CONCILIAR - Movimientos Bancarios sin Conciliar"],
        [""],
        ["Fecha", "Banco", "Código", "Descripción", "Número Doc", "Cargo", "Abono", "Días"],
    ]
    
    for mov in movs:
        # Odoo: amount puede ser positivo o negativo
        monto = mov.get("amount", mov.get("monto", 0)) or 0
        cargo = abs(monto) if monto < 0 else 0
        abono = abs(monto) if monto > 0 else 0
        
        rows.append([
            str(mov.get("fecha", mov.get("date", "")))[:10],
            mov.get("banco", mov.get("journal_name", "")),
            mov.get("banco_codigo", mov.get("journal_code", "")),
            mov.get("descripcion", mov.get("name", mov.get("payment_ref", ""))),
            mov.get("numero_doc", mov.get("ref", "")),
            cargo,
            abono,
            mov.get("dias_pendiente", 0)
        ])
    
    if not movs:
        rows.append(["", "Sin pendientes", "", "", "", "", "", ""])
    
    # Subtotales
    total_cargos = abs(concil.get("total_cargos", 0))
    total_abonos = abs(concil.get("total_abonos", 0))
    rows.append(["", "", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "TOTALES:", total_cargos, total_abonos, ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes Conciliar", index=False, header=False)
    
    ws = writer.sheets["Pendientes Conciliar"]
    
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    for row_idx, mov in enumerate(movs, start=5):
        dias = mov.get("dias_pendiente", 0)
        
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            if dias >= 30:
                cell.fill = styles["alert_fill"]
            elif dias >= 15:
                cell.fill = styles["warning_fill"]
            
            if col in [6, 7]:
                cell.number_format = styles["number_format"]
    
    total_row = 5 + len(movs) + 1
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = styles["number_format"]
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = styles["number_format"]
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def generar_reporte(db_name: str) -> str:
    """
    Genera el reporte Excel de pendientes para una empresa Odoo.
    
    Args:
        db_name: Nombre de base de datos (FactorIT, FactorIT2)
    
    Returns:
        Path del archivo generado
    """
    print(f"\n📊 Obteniendo pendientes de {db_name}...")
    data = obtener_pendientes_empresa(db_name)
    
    empresa = data.get("empresa", db_name)
    fecha = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    filename = OUTPUT_DIR / f"Pendientes_{db_name}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        print("   📋 Generando Resumen...")
        crear_hoja_resumen(writer, data, empresa, fecha)
        
        print("   📄 Generando Pendientes SII...")
        crear_hoja_sii(writer, data)
        
        print("   📝 Generando Pendientes Contabilizar...")
        crear_hoja_contabilizar(writer, data)
        
        print("   🏦 Generando Pendientes Conciliar...")
        crear_hoja_conciliar(writer, data)
    
    print(f"\n✅ Guardado: {filename}")
    return str(filename)


def main():
    """CLI principal."""
    print("═" * 60)
    print("   ART-002: REPORTE DE PENDIENTES (Odoo)")
    print("═" * 60)
    
    db_name = sys.argv[1] if len(sys.argv) > 1 else None
    
    if db_name:
        generar_reporte(db_name)
    else:
        for db in DATABASES.keys():
            generar_reporte(db)
    
    print("\n" + "═" * 60)


if __name__ == "__main__":
    main()
