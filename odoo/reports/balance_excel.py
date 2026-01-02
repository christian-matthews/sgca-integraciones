#!/usr/bin/env python3
"""
ART-001: Balance + Análisis - Implementación Odoo
==================================================

Genera reporte Excel con Balance, EERR y análisis por cuenta.

Uso:
    python -m odoo.reports.balance_excel FactorIT 2025-12-31
    python -m odoo.reports.balance_excel FactorIT2
"""

import os
import sys
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
import psycopg2
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Paths
SCRIPT_DIR = Path(__file__).parent
ODOO_DIR = SCRIPT_DIR.parent
INTEGRACIONES_DIR = ODOO_DIR.parent
OUTPUT_DIR = SCRIPT_DIR / "generados"
OUTPUT_DIR.mkdir(exist_ok=True)

load_dotenv()


def get_env_clean(key, default=None):
    val = os.getenv(key)
    return val.strip() if val else default


DB_CONFIG = {
    'host': get_env_clean('SERVER'),
    'port': get_env_clean('PORT', '5432'),
    'user': get_env_clean('DB_USER'),
    'password': get_env_clean('PASSWORD'),
}

DATABASES = {
    'FactorIT': 'FactorIT SpA',
    'FactorIT2': 'FactorIT Ltda',
}

# Estilos - Paleta verde SGCA (consistente con ART-001)
COLORS = {
    "header_bg": "006400",
    "header_text": "FFFFFF",
    "section_bg": "E8F5E9",
    "subtotal_bg": "C8E6C9",
    "border": "CCCCCC",
}

ESTILOS = {
    "font_header": Font(bold=True, size=10, color=COLORS["header_text"]),
    "fill_header": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
    "font_titulo": Font(bold=True, size=14, color=COLORS["header_bg"]),
    "font_section": Font(bold=True, size=12, color=COLORS["header_bg"]),
    "fill_section": PatternFill(start_color=COLORS["section_bg"], end_color=COLORS["section_bg"], fill_type="solid"),
    "font_category": Font(bold=True, size=11),
    "font_subtotal": Font(bold=True, size=10),
    "font_total": Font(bold=True, size=11),
    "font_total_final": Font(bold=True, size=12, color=COLORS["header_text"]),
    "fill_total_final": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
    "fill_subtotal": PatternFill(start_color=COLORS["subtotal_bg"], end_color=COLORS["subtotal_bg"], fill_type="solid"),
    "thin_border": Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    ),
    "formato_miles": '#,##0'
}

# Fecha de inicio para reportes (coherencia con período de control)
FECHA_INICIO_REPORTE = '2025-01-01'

CLASIFICACION_CUENTAS = {
    "activo_corriente": {"nombre": "Activo Corriente", "prefijos": ["11"]},
    "activo_no_corriente": {"nombre": "Activo No Corriente", "prefijos": ["12", "13", "14"]},
    "pasivo_corriente": {"nombre": "Pasivo Corriente", "prefijos": ["21"]},
    "pasivo_no_corriente": {"nombre": "Pasivo No Corriente", "prefijos": ["22", "23"]},
    "patrimonio": {"nombre": "Patrimonio", "prefijos": ["3"]},
    "ingresos": {"nombre": "Ingresos Operacionales", "prefijos": ["4"]},
    "costos": {"nombre": "Costos de Venta", "prefijos": ["51"]},
    "gastos_operacionales": {"nombre": "Gastos Operacionales", "prefijos": ["52", "53", "54", "55"]},
    "otros_ingresos": {"nombre": "Otros Ingresos", "prefijos": ["6"]},
    "otros_gastos": {"nombre": "Otros Gastos", "prefijos": ["7"]},
    "apertura": {"nombre": "Saldos Apertura", "prefijos": ["8"]},
}


def clasificar_cuenta(codigo):
    for categoria, config in CLASIFICACION_CUENTAS.items():
        for prefijo in config["prefijos"]:
            if codigo.startswith(prefijo):
                return categoria
    return None


def sanitize_sheet_name(codigo, nombre):
    name = f"{codigo} {nombre}"
    for char in ['\\', '/', '*', '?', '[', ']', ':']:
        name = name.replace(char, '')
    return name[:31]


def obtener_balance(cursor, fecha_hasta):
    query = """
    SELECT 
        aa.code as codigo,
        aa.name as cuenta,
        aa.user_type_id as tipo,
        SUM(aml.debit) as debe,
        SUM(aml.credit) as haber,
        SUM(aml.debit) - SUM(aml.credit) as saldo
    FROM account_move_line aml
    JOIN account_account aa ON aml.account_id = aa.id
    JOIN account_move am ON aml.move_id = am.id
    WHERE am.state = 'posted'
      AND aml.date >= %s
      AND aml.date <= %s
    GROUP BY aa.id, aa.code, aa.name, aa.user_type_id
    HAVING SUM(aml.debit) != 0 OR SUM(aml.credit) != 0
    ORDER BY aa.code
    """
    cursor.execute(query, (FECHA_INICIO_REPORTE, fecha_hasta))
    
    cuentas = []
    for row in cursor.fetchall():
        codigo, nombre, tipo, debe, haber, saldo = row
        debe = float(debe or 0)
        haber = float(haber or 0)
        saldo = float(saldo or 0)
        clasificacion = clasificar_cuenta(codigo)
        
        if clasificacion in ["activo_corriente", "activo_no_corriente"]:
            valor_balance = saldo
        elif clasificacion in ["pasivo_corriente", "pasivo_no_corriente", "patrimonio"]:
            valor_balance = -saldo
        elif clasificacion in ["ingresos", "otros_ingresos"]:
            valor_balance = -saldo
        elif clasificacion in ["costos", "gastos_operacionales", "otros_gastos"]:
            valor_balance = saldo
        elif clasificacion == "apertura":
            valor_balance = -saldo
        else:
            valor_balance = saldo
        
        cuentas.append({
            'codigo': codigo,
            'cuenta': nombre,
            'tipo': tipo,
            'debe': debe,
            'haber': haber,
            'saldo': saldo,
            'clasificacion': clasificacion,
            'valor_balance': valor_balance
        })
    
    return cuentas


def obtener_movimientos_cuenta(cursor, codigo_cuenta, fecha_hasta):
    query = """
    SELECT 
        aml.date as fecha,
        am.name as asiento,
        aml.name as descripcion,
        rp.name as tercero,
        aml.debit as debe,
        aml.credit as haber
    FROM account_move_line aml
    JOIN account_account aa ON aml.account_id = aa.id
    JOIN account_move am ON aml.move_id = am.id
    LEFT JOIN res_partner rp ON aml.partner_id = rp.id
    WHERE aa.code = %s
      AND am.state = 'posted'
      AND aml.date >= %s
      AND aml.date <= %s
    ORDER BY aml.date DESC, am.name
    """
    cursor.execute(query, (codigo_cuenta, FECHA_INICIO_REPORTE, fecha_hasta))
    return cursor.fetchall()


def generar_reporte(db_name: str, fecha_corte: str) -> str:
    """
    Genera el reporte de Balance + Análisis para una empresa Odoo.
    
    Args:
        db_name: Nombre de base de datos (FactorIT, FactorIT2)
        fecha_corte: Fecha de corte en formato YYYY-MM-DD
    
    Returns:
        Path del archivo generado
    """
    empresa_nombre = DATABASES.get(db_name, db_name)
    
    print("=" * 70)
    print(f"   GENERANDO BALANCE - {empresa_nombre}")
    print("=" * 70)
    print(f"   Fecha de corte: {fecha_corte}")
    
    conn = psycopg2.connect(
        host=DB_CONFIG['host'],
        port=DB_CONFIG['port'],
        database=db_name,
        user=DB_CONFIG['user'],
        password=DB_CONFIG['password'],
    )
    cursor = conn.cursor()
    
    print("\n📊 Obteniendo balance...")
    cuentas = obtener_balance(cursor, fecha_corte)
    print(f"   {len(cuentas)} cuentas con movimientos")
    
    # Clasificar
    categorias = {cat: [] for cat in CLASIFICACION_CUENTAS.keys()}
    for c in cuentas:
        if c['clasificacion'] and c['valor_balance'] != 0:
            categorias[c['clasificacion']].append(c)
    
    # Calcular totales
    total_activo_corriente = sum(c['valor_balance'] for c in categorias.get('activo_corriente', []))
    total_activo_no_corriente = sum(c['valor_balance'] for c in categorias.get('activo_no_corriente', []))
    total_activos = total_activo_corriente + total_activo_no_corriente
    
    total_pasivo_corriente = sum(c['valor_balance'] for c in categorias.get('pasivo_corriente', []))
    total_pasivo_no_corriente = sum(c['valor_balance'] for c in categorias.get('pasivo_no_corriente', []))
    total_pasivos = total_pasivo_corriente + total_pasivo_no_corriente
    
    patrimonio_sin_resultado = sum(c['valor_balance'] for c in categorias.get('patrimonio', []))
    
    # Estado de Resultados
    ingresos = sum(c['valor_balance'] for c in categorias.get('ingresos', []))
    costos = sum(c['valor_balance'] for c in categorias.get('costos', []))
    utilidad_bruta = ingresos - costos
    
    gastos_op = sum(c['valor_balance'] for c in categorias.get('gastos_operacionales', []))
    resultado_operacional = utilidad_bruta - gastos_op
    
    otros_ingresos = sum(c['valor_balance'] for c in categorias.get('otros_ingresos', []))
    otros_gastos = sum(c['valor_balance'] for c in categorias.get('otros_gastos', []))
    resultado_no_operacional = otros_ingresos - otros_gastos
    
    resultado_neto = resultado_operacional + resultado_no_operacional
    
    ajustes_apertura = sum(c['valor_balance'] for c in categorias.get('apertura', []))
    total_patrimonio = patrimonio_sin_resultado + resultado_neto + ajustes_apertura
    
    diferencia = total_activos - (total_pasivos + total_patrimonio)
    cuadra = abs(diferencia) < 1
    
    print(f"\n📋 Resumen:")
    print(f"   Total Activos:      ${total_activos:>18,.0f}")
    print(f"   Total Pasivos:      ${total_pasivos:>18,.0f}")
    print(f"   Total Patrimonio:   ${total_patrimonio:>18,.0f}")
    print(f"   {'✅ CUADRA' if cuadra else f'⚠️ DIFERENCIA: ${diferencia:,.0f}'}")
    
    # Crear Excel
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    periodo = fecha_corte.replace('-', '')[:6]
    filename = OUTPUT_DIR / f'Balance_Odoo_{db_name}_{periodo}_{timestamp}.xlsx'
    
    print(f"\n📝 Generando Excel...")
    
    cuenta_a_hoja = {}
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # HOJA RESUMEN
        rows = []
        row_types = []
        
        rows.append([f"BALANCE GENERAL - {empresa_nombre}", "", "", "", f"Fecha: {fecha_corte}"])
        row_types.append("titulo")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Balance
        rows.append(["BALANCE CLASIFICADO", "", "", "", ""])
        row_types.append("section")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Activos
        rows.append(["ACTIVOS", "", "", "", ""])
        row_types.append("category")
        rows.append(["  Activo Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('activo_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Activo Corriente", total_activo_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["  Activo No Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('activo_no_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Activo No Corriente", total_activo_no_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["TOTAL ACTIVOS", total_activos, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Pasivos
        rows.append(["PASIVOS", "", "", "", ""])
        row_types.append("category")
        rows.append(["  Pasivo Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('pasivo_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Pasivo Corriente", total_pasivo_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["  Pasivo No Corriente", "", "", "", ""])
        row_types.append("subcategory")
        for c in categorias.get('pasivo_no_corriente', [])[:10]:
            rows.append([f"    {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Total Pasivo No Corriente", total_pasivo_no_corriente, "", "", ""])
        row_types.append("subtotal")
        
        rows.append(["TOTAL PASIVOS", total_pasivos, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Patrimonio
        rows.append(["PATRIMONIO", "", "", "", ""])
        row_types.append("category")
        for c in categorias.get('patrimonio', []):
            rows.append([f"  {c['cuenta']}", c['valor_balance'], "", "", ""])
            row_types.append("item")
        rows.append(["  Resultado del Período", resultado_neto, "", "", "(calculado)"])
        row_types.append("item")
        if ajustes_apertura != 0:
            rows.append(["  Ajustes de Apertura", ajustes_apertura, "", "", ""])
            row_types.append("item")
        rows.append(["TOTAL PATRIMONIO", total_patrimonio, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["TOTAL PASIVOS + PATRIMONIO", total_pasivos + total_patrimonio, "", "", ""])
        row_types.append("total_final")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        if cuadra:
            rows.append(["✅ CUADRATURA OK", "", "", "", ""])
            row_types.append("verification_ok")
        else:
            rows.append([f"⚠️ DESCUADRE: ${diferencia:,.0f}", "", "", "", ""])
            row_types.append("verification_error")
        
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        # Estado de Resultados
        rows.append(["ESTADO DE RESULTADOS", "", "", "", ""])
        row_types.append("section")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Ingresos Operacionales", ingresos, "", "", ""])
        row_types.append("item")
        rows.append(["Costo de Ventas", -costos, "", "", ""])
        row_types.append("item")
        rows.append(["UTILIDAD BRUTA", utilidad_bruta, "", "", ""])
        row_types.append("subtotal")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Gastos Operacionales", -gastos_op, "", "", ""])
        row_types.append("item")
        rows.append(["RESULTADO OPERACIONAL", resultado_operacional, "", "", ""])
        row_types.append("total")
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        
        rows.append(["Otros Ingresos", otros_ingresos, "", "", ""])
        row_types.append("item")
        rows.append(["Otros Gastos", -otros_gastos, "", "", ""])
        row_types.append("item")
        rows.append(["RESULTADO NETO", resultado_neto, "", "", ""])
        row_types.append("total_final")
        
        # KPIs
        rows.append(["", "", "", "", ""])
        row_types.append("empty")
        rows.append(["INDICADORES", "", "", "", ""])
        row_types.append("section")
        
        margen_bruto = (utilidad_bruta / ingresos * 100) if ingresos else 0
        margen_neto = (resultado_neto / ingresos * 100) if ingresos else 0
        
        rows.append(["Margen Bruto", f"{margen_bruto:.1f}%", "", "", ""])
        row_types.append("kpi")
        rows.append(["Margen Neto", f"{margen_neto:.1f}%", "", "", ""])
        row_types.append("kpi")
        
        # Escribir Resumen
        df_resumen = pd.DataFrame(rows)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
        ws_resumen = writer.sheets['Resumen']
        
        for row_idx, row_type in enumerate(row_types, start=1):
            cell_a = ws_resumen.cell(row=row_idx, column=1)
            cell_b = ws_resumen.cell(row=row_idx, column=2)
            
            if row_type == "titulo":
                cell_a.font = ESTILOS["font_titulo"]
            elif row_type == "section":
                cell_a.font = ESTILOS["font_section"]
                cell_a.fill = ESTILOS["fill_section"]
            elif row_type == "category":
                cell_a.font = ESTILOS["font_category"]
            elif row_type == "subtotal":
                cell_a.font = ESTILOS["font_subtotal"]
                cell_b.font = ESTILOS["font_subtotal"]
                cell_a.fill = ESTILOS["fill_subtotal"]
                cell_b.fill = ESTILOS["fill_subtotal"]
            elif row_type == "total":
                cell_a.font = ESTILOS["font_total"]
                cell_b.font = ESTILOS["font_total"]
            elif row_type == "total_final":
                cell_a.font = ESTILOS["font_total_final"]
                cell_a.fill = ESTILOS["fill_total_final"]
                cell_b.font = ESTILOS["font_total_final"]
                cell_b.fill = ESTILOS["fill_total_final"]
            elif row_type == "verification_ok":
                cell_a.font = Font(bold=True, color="006400")
            elif row_type == "verification_error":
                cell_a.font = Font(bold=True, color="FF0000")
            elif row_type == "kpi":
                cell_b.font = Font(bold=True, color=COLORS["header_bg"])
            
            if isinstance(rows[row_idx-1][1], (int, float)):
                cell_b.number_format = ESTILOS["formato_miles"]
        
        ws_resumen.column_dimensions['A'].width = 45
        ws_resumen.column_dimensions['B'].width = 18
        ws_resumen.column_dimensions['E'].width = 20
        
        # HOJA BALANCE DETALLADO
        df_balance = pd.DataFrame(cuentas)
        df_balance = df_balance[['codigo', 'cuenta', 'debe', 'haber', 'saldo', 'clasificacion']]
        df_balance.columns = ['Código', 'Cuenta', 'Debe', 'Haber', 'Saldo', 'Clasificación']
        df_balance['Ver Detalle'] = ''
        
        df_balance.to_excel(writer, sheet_name='Balance', index=False, startrow=1)
        
        ws_balance = writer.sheets['Balance']
        ws_balance.cell(row=1, column=1).value = f"BALANCE DETALLADO - {empresa_nombre} - {fecha_corte}"
        ws_balance.cell(row=1, column=1).font = ESTILOS["font_titulo"]
        
        for col in range(1, 8):
            cell = ws_balance.cell(row=2, column=col)
            cell.font = ESTILOS["font_header"]
            cell.fill = ESTILOS["fill_header"]
        
        # HOJAS DE DETALLE
        print("   Generando hojas de detalle...")
        cuentas_con_detalle = 0
        
        for c in cuentas:
            if abs(c['saldo']) < 1:
                continue
            
            movimientos = obtener_movimientos_cuenta(cursor, c['codigo'], fecha_corte)
            if not movimientos:
                continue
            
            sheet_name = sanitize_sheet_name(c['codigo'], c['cuenta'])
            
            df_mov = pd.DataFrame(movimientos, columns=['Fecha', 'Asiento', 'Descripción', 'Tercero', 'Debe', 'Haber'])
            df_mov['Fecha'] = df_mov['Fecha'].astype(str)
            
            # Convertir Debe y Haber a float (vienen como Decimal de PostgreSQL)
            df_mov['Debe'] = df_mov['Debe'].apply(lambda x: float(x) if x else 0.0)
            df_mov['Haber'] = df_mov['Haber'].apply(lambda x: float(x) if x else 0.0)
            
            saldo_acum = []
            acum = 0.0
            for _, row in df_mov.iterrows():
                acum += row['Debe'] - row['Haber']
                saldo_acum.append(acum)
            df_mov['Saldo'] = saldo_acum
            
            df_mov.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
            
            ws = writer.sheets[sheet_name]
            ws.cell(row=1, column=1).value = "← Volver al Balance"
            ws.cell(row=1, column=1).hyperlink = "#'Balance'!A1"
            ws.cell(row=1, column=1).style = "Hyperlink"
            ws.cell(row=2, column=1).value = f"{c['codigo']} - {c['cuenta']}"
            ws.cell(row=2, column=1).font = ESTILOS["font_titulo"]
            
            cuenta_a_hoja[c['codigo']] = sheet_name
            cuentas_con_detalle += 1
        
        print(f"   {cuentas_con_detalle} cuentas con detalle")
        
        # Hipervínculos
        for row_idx, c in enumerate(cuentas, start=3):
            if c['codigo'] in cuenta_a_hoja:
                cell = ws_balance.cell(row=row_idx, column=7)
                cell.value = "→ Ver"
                cell.hyperlink = f"#'{cuenta_a_hoja[c['codigo']]}'!A1"
                cell.style = "Hyperlink"
    
    cursor.close()
    conn.close()
    
    print(f"\n✅ Archivo generado: {filename}")
    print("=" * 70)
    
    return str(filename)


def main():
    db_name = sys.argv[1] if len(sys.argv) > 1 else "FactorIT"
    fecha = sys.argv[2] if len(sys.argv) > 2 else datetime.now().strftime('%Y-%m-%d')
    
    if db_name not in DATABASES:
        print(f"❌ Base de datos '{db_name}' no encontrada")
        print(f"   Disponibles: {', '.join(DATABASES.keys())}")
        sys.exit(1)
    
    generar_reporte(db_name, fecha)


if __name__ == '__main__':
    main()
