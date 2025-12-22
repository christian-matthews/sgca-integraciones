#!/usr/bin/env python3
"""
Skualo Control - Sistema de Control y Reportes
===============================================

Sistema escalable para gestionar múltiples empresas.

Uso:
    # Setup inicial de empresa (solo necesita RUT)
    python skualo_control.py setup 77285542-7
    
    # Controles de pendientes
    python skualo_control.py bancos 77285542-7
    python skualo_control.py aprobar 77285542-7
    python skualo_control.py contabilizar 77285542-7
    python skualo_control.py reporte 77285542-7
    
    # Reportes contables
    python skualo_control.py balance 77285542-7 [periodo]
    
    # Administración
    python skualo_control.py listar
"""

import os
import sys
import json
import requests
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

load_dotenv()
TOKEN = os.getenv('SKUALO_API_TOKEN')
BASE_URL = 'https://api.skualo.cl'
DIAS_ACEPTACION_TACITA = 8

# Directorio de configuraciones
CONFIG_DIR = Path(__file__).parent / 'config' / 'empresas'
CONFIG_DIR.mkdir(parents=True, exist_ok=True)

# Mapeo de tipos DTE del SII a tipos internos Skualo
TIPO_DTE_A_INTERNO = {
    33: 'FACE',   # Factura Electrónica
    34: 'FXCE',   # Factura No Afecta o Exenta Electrónica
    61: 'NCCE',   # Nota de Crédito Electrónica
    56: 'NDCE',   # Nota de Débito Electrónica
    52: 'GDES',   # Guía de Despacho Electrónica
    110: 'FEXP',  # Factura de Exportación Electrónica
}

# Palabras clave para detectar cuentas bancarias
PALABRAS_BANCO = [
    'banco', 'santander', 'chile', 'estado', 'bci', 'scotiabank', 
    'itau', 'itaú', 'security', 'bice', 'falabella', 'ripley',
    'consorcio', 'internacional', 'corpbanca', 'tapp', 'tenpo',
    'mercado pago', 'cuenta corriente', 'cta cte', 'cta. cte',
    'coopeuch', 'bancoestado', 'bco.', 'bco '
]


# ═══════════════════════════════════════════════════════════════════════════════
# UTILIDADES API
# ═══════════════════════════════════════════════════════════════════════════════

def get_headers():
    return {
        'Authorization': f'Bearer {TOKEN}',
        'accept': 'application/json'
    }


def api_get(rut, endpoint, params=None):
    """Realiza una llamada GET a la API."""
    url = f'{BASE_URL}/{rut}{endpoint}'
    try:
        r = requests.get(url, headers=get_headers(), params=params, timeout=30)
        if r.ok:
            return r.json()
    except Exception as e:
        print(f'   ⚠️ Error API: {e}')
    return None


def api_get_all(rut, endpoint, params=None):
    """Obtiene todos los registros paginados de un endpoint."""
    all_items = []
    page = 1
    base_params = params or {}
    
    while True:
        paged_params = {**base_params, 'PageSize': 100, 'Page': page}
        data = api_get(rut, endpoint, paged_params)
        if not data:
            break
        items = data.get('items', data) if isinstance(data, dict) else data
        if isinstance(items, list):
            all_items.extend(items)
        if not data.get('next'):
            break
        page += 1
    
    return all_items


# ═══════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════════

def get_config_path(rut):
    """Obtiene la ruta del archivo de configuración de una empresa."""
    return CONFIG_DIR / f'{rut}.json'


def config_existe(rut):
    """Verifica si existe configuración para una empresa."""
    return get_config_path(rut).exists()


def cargar_config(rut):
    """Carga la configuración de una empresa."""
    path = get_config_path(rut)
    if not path.exists():
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def guardar_config(rut, config):
    """Guarda la configuración de una empresa."""
    path = get_config_path(rut)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    print(f'\n   💾 Configuración guardada: {path}')


def listar_empresas_configuradas():
    """Lista todas las empresas configuradas."""
    configs = list(CONFIG_DIR.glob('*.json'))
    if not configs:
        print('\n   No hay empresas configuradas.')
        print('   Usa: python skualo_control.py setup <RUT>')
        return []
    
    print('\n   Empresas configuradas:')
    print('   ' + '-' * 60)
    empresas = []
    for config_path in configs:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            empresas.append(config)
            print(f"   {config['rut']}: {config['nombre']}")
            print(f"      Cuentas bancarias: {len(config.get('cuentas_bancarias', []))}")
            print(f"      Configurado: {config.get('configurado_el', '?')}")
    
    return empresas


# ═══════════════════════════════════════════════════════════════════════════════
# SETUP DE EMPRESA
# ═══════════════════════════════════════════════════════════════════════════════

def setup_empresa(rut):
    """
    Configura una nueva empresa.
    Solo necesita el RUT, extrae todo lo demás automáticamente.
    """
    print('=' * 80)
    print('SETUP DE EMPRESA')
    print('=' * 80)
    print(f'\n   RUT: {rut}')
    
    # Verificar si ya existe configuración
    if config_existe(rut):
        config_actual = cargar_config(rut)
        print(f'\n   ⚠️ Ya existe configuración para {config_actual["nombre"]}')
        respuesta = input('   ¿Desea reconfigurar? (s/N): ').strip().lower()
        if respuesta != 's':
            print('   Operación cancelada.')
            return None
    
    config = {
        'rut': rut,
        'configurado_el': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'cuentas_bancarias': [],
        'cuenta_clientes': None,
        'cuenta_proveedores': None,
    }
    
    # ───────────────────────────────────────────────────────────────────────────
    # 1. OBTENER INFORMACIÓN DE LA EMPRESA
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   1. Obteniendo información de la empresa...')
    
    empresa = api_get(rut, '/empresa')
    if not empresa:
        print('   ❌ No se pudo conectar con la API. Verifique el RUT y el token.')
        return None
    
    config['nombre'] = empresa.get('nombre', rut)
    config['razon_social'] = empresa.get('razonSocial', '')
    config['giro'] = empresa.get('giro', '')
    
    print(f'      ✅ {config["nombre"]}')
    print(f'      Razón Social: {config["razon_social"]}')
    
    # ───────────────────────────────────────────────────────────────────────────
    # 2. DETECTAR CUENTAS BANCARIAS
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   2. Detectando cuentas bancarias...')
    
    # Obtener balance del período actual
    periodo = datetime.now().strftime('%Y%m')
    balance = api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
    
    if not balance:
        # Intentar con el mes anterior
        from datetime import timedelta
        fecha_ant = datetime.now().replace(day=1) - timedelta(days=1)
        periodo = fecha_ant.strftime('%Y%m')
        balance = api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
    
    if not balance:
        print('      ⚠️ No se pudo obtener el balance. Deberá configurar cuentas manualmente.')
    else:
        # Detectar cuentas bancarias
        cuentas_detectadas = []
        for cuenta in balance:
            codigo = cuenta.get('idCuenta', '')
            nombre = cuenta.get('cuenta', '').lower()
            
            es_banco = False
            
            # Método 1: Prefijo 1102 (estándar)
            if codigo.startswith('1102'):
                es_banco = True
            
            # Método 2: Prefijo 1103 (moneda extranjera)
            elif codigo.startswith('1103'):
                es_banco = True
            
            # Método 3: Nombre contiene palabra clave
            elif any(palabra in nombre for palabra in PALABRAS_BANCO):
                if codigo.startswith('1'):  # Solo cuentas de activo
                    es_banco = True
            
            if es_banco:
                cuentas_detectadas.append({
                    'codigo': codigo,
                    'nombre': cuenta.get('cuenta', ''),
                    'activa': True,
                    'saldo': cuenta.get('saldo', 0)
                })
        
        # Mostrar cuentas detectadas y pedir confirmación
        if cuentas_detectadas:
            print(f'\n      Se detectaron {len(cuentas_detectadas)} cuentas bancarias:')
            print('      ' + '-' * 60)
            for i, c in enumerate(cuentas_detectadas, 1):
                saldo_str = f'${c["saldo"]:,.0f}' if c["saldo"] else '$0'
                print(f'      {i}. [{c["codigo"]}] {c["nombre"]} (Saldo: {saldo_str})')
            
            print('\n      ¿Son correctas estas cuentas bancarias?')
            print('      - Presione ENTER para confirmar todas')
            print('      - Escriba los números a EXCLUIR separados por coma (ej: 2,3)')
            print('      - Escriba "manual" para ingresar códigos manualmente')
            
            respuesta = input('\n      Su selección: ').strip().lower()
            
            if respuesta == 'manual':
                print('\n      Ingrese los códigos de cuenta separados por coma:')
                codigos = input('      Códigos: ').strip()
                cuentas_manuales = []
                for cod in codigos.split(','):
                    cod = cod.strip()
                    # Buscar la cuenta en el balance
                    for cuenta in balance:
                        if cuenta.get('idCuenta', '') == cod:
                            cuentas_manuales.append({
                                'codigo': cod,
                                'nombre': cuenta.get('cuenta', ''),
                                'activa': True,
                                'saldo': cuenta.get('saldo', 0)
                            })
                            break
                    else:
                        cuentas_manuales.append({
                            'codigo': cod,
                            'nombre': f'Cuenta {cod}',
                            'activa': True,
                            'saldo': 0
                        })
                cuentas_detectadas = cuentas_manuales
            elif respuesta:
                # Excluir las seleccionadas
                excluir = set()
                for num in respuesta.split(','):
                    try:
                        excluir.add(int(num.strip()))
                    except:
                        pass
                cuentas_detectadas = [c for i, c in enumerate(cuentas_detectadas, 1) 
                                     if i not in excluir]
            
            config['cuentas_bancarias'] = cuentas_detectadas
            print(f'\n      ✅ {len(cuentas_detectadas)} cuentas bancarias configuradas')
        else:
            print('      ⚠️ No se detectaron cuentas bancarias automáticamente.')
            print('      Ingrese los códigos de cuenta separados por coma:')
            codigos = input('      Códigos: ').strip()
            if codigos:
                for cod in codigos.split(','):
                    cod = cod.strip()
                    config['cuentas_bancarias'].append({
                        'codigo': cod,
                        'nombre': f'Cuenta {cod}',
                        'activa': True
                    })
    
    # ───────────────────────────────────────────────────────────────────────────
    # 3. DETECTAR CUENTAS DE CLIENTES Y PROVEEDORES
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   3. Detectando cuentas de clientes y proveedores...')
    
    if balance:
        # Buscar cuenta de clientes (1107xxx o 1108xxx)
        for cuenta in balance:
            codigo = cuenta.get('idCuenta', '')
            nombre = cuenta.get('cuenta', '').lower()
            
            if not config['cuenta_clientes']:
                if codigo.startswith('1107') or codigo.startswith('1108'):
                    config['cuenta_clientes'] = codigo
                    print(f'      ✅ Clientes: [{codigo}] {cuenta.get("cuenta", "")}')
                elif 'cliente' in nombre or 'por cobrar' in nombre or 'deudor' in nombre:
                    config['cuenta_clientes'] = codigo
                    print(f'      ✅ Clientes: [{codigo}] {cuenta.get("cuenta", "")}')
            
            if not config['cuenta_proveedores']:
                if codigo.startswith('2110') or codigo.startswith('2111'):
                    config['cuenta_proveedores'] = codigo
                    print(f'      ✅ Proveedores: [{codigo}] {cuenta.get("cuenta", "")}')
                elif 'proveedor' in nombre or 'por pagar' in nombre or 'acreedor' in nombre:
                    config['cuenta_proveedores'] = codigo
                    print(f'      ✅ Proveedores: [{codigo}] {cuenta.get("cuenta", "")}')
    
    # ───────────────────────────────────────────────────────────────────────────
    # 4. VERIFICAR ACCESO A ENDPOINTS NECESARIOS
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   4. Verificando acceso a endpoints...')
    
    endpoints_check = [
        ('/sii/dte/recibidos', 'DTEs Recibidos'),
        ('/sii/dte', 'DTEs Emitidos'),
    ]
    
    config['endpoints_disponibles'] = {}
    
    for endpoint, nombre in endpoints_check:
        data = api_get(rut, endpoint, {'PageSize': 1})
        disponible = data is not None
        config['endpoints_disponibles'][endpoint] = disponible
        status = '✅' if disponible else '❌'
        print(f'      {status} {nombre}')
    
    # Verificar cuentas bancarias
    if config['cuentas_bancarias']:
        cuenta_test = config['cuentas_bancarias'][0]['codigo']
        data = api_get(rut, f'/bancos/{cuenta_test}', {'PageSize': 1})
        disponible = data is not None
        config['endpoints_disponibles']['/bancos'] = disponible
        status = '✅' if disponible else '❌'
        print(f'      {status} Movimientos Bancarios')
    
    # ───────────────────────────────────────────────────────────────────────────
    # 5. GUARDAR CONFIGURACIÓN
    # ───────────────────────────────────────────────────────────────────────────
    guardar_config(rut, config)
    
    # Resumen
    print('\n' + '=' * 80)
    print('RESUMEN DE CONFIGURACIÓN')
    print('=' * 80)
    print(f'''
   Empresa: {config['nombre']}
   RUT: {config['rut']}
   
   Cuentas Bancarias: {len(config['cuentas_bancarias'])}''')
    for c in config['cuentas_bancarias']:
        print(f'      • [{c["codigo"]}] {c["nombre"]}')
    print(f'''
   Cuenta Clientes: {config['cuenta_clientes'] or 'No detectada'}
   Cuenta Proveedores: {config['cuenta_proveedores'] or 'No detectada'}
   
   ✅ Configuración completada
''')
    print('=' * 80)
    
    return config


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN 1: MOVIMIENTOS BANCARIOS PENDIENTES DE CONCILIAR
# ═══════════════════════════════════════════════════════════════════════════════

def movimientos_bancarios_pendientes(rut, mostrar=True):
    """
    Obtiene movimientos bancarios pendientes de conciliar.
    Usa la configuración guardada de la empresa.
    """
    config = cargar_config(rut)
    if not config:
        print(f'\n   ❌ No hay configuración para {rut}')
        print(f'   Ejecute primero: python skualo_control.py setup {rut}')
        return None
    
    if mostrar:
        print('\n' + '─' * 80)
        print(f'🏦 MOVIMIENTOS BANCARIOS SIN CONCILIAR - {config["nombre"]}')
        print('─' * 80)
    
    resultado = {
        'empresa': config['nombre'],
        'rut': rut,
        'fecha': datetime.now().isoformat(),
        'cuentas': [],
        'total_sin_conciliar': 0
    }
    
    if not config.get('cuentas_bancarias'):
        if mostrar:
            print('\n   ⚠️ No hay cuentas bancarias configuradas')
        return resultado
    
    for cuenta in config['cuentas_bancarias']:
        if not cuenta.get('activa', True):
            continue
        
        codigo = cuenta['codigo']
        nombre = cuenta['nombre']
        
        # Obtener todos los movimientos
        movimientos = api_get_all(rut, f'/bancos/{codigo}')
        
        # Filtrar no conciliados
        sin_conciliar = [m for m in movimientos if not m.get('conciliado', True)]
        
        cuenta_resultado = {
            'codigo': codigo,
            'nombre': nombre,
            'total_movimientos': len(movimientos),
            'sin_conciliar': len(sin_conciliar),
            'movimientos': sin_conciliar
        }
        resultado['cuentas'].append(cuenta_resultado)
        resultado['total_sin_conciliar'] += len(sin_conciliar)
        
        if mostrar and sin_conciliar:
            print(f'\n   📌 {nombre} ({codigo})')
            print(f'      Total: {len(movimientos)} | Sin conciliar: {len(sin_conciliar)}')
            print(f'\n      {"Fecha":<12} {"NumDoc":<12} {"Cargo":>12} {"Abono":>12} {"Glosa":<25}')
            print(f'      {"-"*73}')
            
            for m in sin_conciliar[:10]:
                fecha = str(m.get('fecha', ''))[:10]
                num = str(m.get('numDoc', ''))[:12]
                cargo = m.get('montoCargo', 0) or 0
                abono = m.get('montoAbono', 0) or 0
                glosa = str(m.get('glosa', ''))[:25]
                cargo_s = f'{cargo:>12,.0f}' if cargo else ' ' * 12
                abono_s = f'{abono:>12,.0f}' if abono else ' ' * 12
                print(f'      {fecha:<12} {num:<12} {cargo_s} {abono_s} {glosa}')
            
            if len(sin_conciliar) > 10:
                print(f'      ... y {len(sin_conciliar) - 10} más')
    
    if mostrar:
        if resultado['total_sin_conciliar'] == 0:
            print('\n   ✅ Todos los movimientos están conciliados')
        print(f'\n   📊 TOTAL SIN CONCILIAR: {resultado["total_sin_conciliar"]}')
    
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN 2: DOCUMENTOS PENDIENTES DE APROBAR EN SII
# ═══════════════════════════════════════════════════════════════════════════════

def documentos_por_aprobar_sii(rut, mostrar=True):
    """
    Obtiene documentos pendientes de aprobar en el SII.
    Regla: DTEs recibidos con menos de 8 días sin respuesta.
    """
    config = cargar_config(rut)
    if not config:
        print(f'\n   ❌ No hay configuración para {rut}')
        print(f'   Ejecute primero: python skualo_control.py setup {rut}')
        return None
    
    if mostrar:
        print('\n' + '─' * 80)
        print(f'📄 DOCUMENTOS PENDIENTES DE APROBAR EN SII - {config["nombre"]}')
        print('─' * 80)
    
    resultado = {
        'empresa': config['nombre'],
        'rut': rut,
        'fecha': datetime.now().isoformat(),
        'pendientes': [],
        'total_pendientes': 0,
        'monto_total': 0
    }
    
    # Obtener DTEs recibidos
    dtes = api_get_all(rut, '/sii/dte/recibidos')
    hoy = datetime.now()
    
    for dte in dtes:
        # Si ya tiene respuesta, no está pendiente
        if dte.get('fechaRespuesta'):
            continue
        
        # Calcular días desde recepción
        fecha_recep_str = dte.get('creadoEl', '')
        if not fecha_recep_str:
            continue
        
        try:
            fecha_recep = datetime.fromisoformat(fecha_recep_str.split('.')[0])
            dias = (hoy - fecha_recep).days
        except:
            continue
        
        # Si tiene más de 8 días, ya está aceptado tácitamente
        if dias > DIAS_ACEPTACION_TACITA:
            continue
        
        # Es un documento pendiente de aprobar
        doc = {
            'rut_emisor': dte.get('rutEmisor', ''),
            'emisor': dte.get('emisor', ''),
            'tipo_documento': dte.get('tipoDocumento', ''),
            'tipo_dte': dte.get('idTipoDocumento'),
            'folio': dte.get('folio'),
            'fecha_emision': str(dte.get('fechaEmision', ''))[:10],
            'monto': dte.get('montoTotal', 0),
            'dias_restantes': DIAS_ACEPTACION_TACITA - dias,
            'fecha_limite': (fecha_recep + __import__('datetime').timedelta(days=DIAS_ACEPTACION_TACITA)).strftime('%Y-%m-%d')
        }
        resultado['pendientes'].append(doc)
        resultado['monto_total'] += doc['monto']
    
    resultado['total_pendientes'] = len(resultado['pendientes'])
    
    if mostrar:
        if resultado['pendientes']:
            print(f'\n   {"Emisor":<30} {"Tipo":<15} {"Folio":<10} {"Días":>5} {"Total":>15}')
            print(f'   {"-"*77}')
            
            for doc in sorted(resultado['pendientes'], key=lambda x: x['dias_restantes']):
                print(f'   {doc["emisor"][:30]:<30} {doc["tipo_documento"][:15]:<15} {doc["folio"]:<10} {doc["dias_restantes"]:>5} {doc["monto"]:>15,.0f}')
            
            print(f'   {"-"*77}')
            print(f'   {"TOTAL":>62} {resultado["monto_total"]:>15,.0f}')
        else:
            print('\n   ✅ No hay documentos pendientes de aprobar')
        
        print(f'\n   📊 TOTAL PENDIENTES: {resultado["total_pendientes"]}')
    
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN 3: DOCUMENTOS PENDIENTES DE CONTABILIZAR
# ═══════════════════════════════════════════════════════════════════════════════

def documentos_por_contabilizar(rut, mostrar=True):
    """
    Obtiene documentos pendientes de contabilizar.
    Son DTEs aceptados (> 8 días o con respuesta) que NO existen en /documentos.
    """
    config = cargar_config(rut)
    if not config:
        print(f'\n   ❌ No hay configuración para {rut}')
        print(f'   Ejecute primero: python skualo_control.py setup {rut}')
        return None
    
    if mostrar:
        print('\n' + '─' * 80)
        print(f'📋 DOCUMENTOS PENDIENTES DE CONTABILIZAR - {config["nombre"]}')
        print('─' * 80)
    
    resultado = {
        'empresa': config['nombre'],
        'rut': rut,
        'fecha': datetime.now().isoformat(),
        'pendientes': [],
        'ya_contabilizados': 0,
        'total_pendientes': 0,
        'monto_total': 0
    }
    
    # Obtener DTEs recibidos
    dtes = api_get_all(rut, '/sii/dte/recibidos')
    hoy = datetime.now()
    
    if mostrar:
        print(f'\n   Verificando {len(dtes)} DTEs recibidos...')
    
    for dte in dtes:
        # Verificar si está aceptado
        fecha_respuesta = dte.get('fechaRespuesta')
        
        if not fecha_respuesta:
            # Verificar aceptación tácita (> 8 días)
            fecha_recep_str = dte.get('creadoEl', '')
            try:
                fecha_recep = datetime.fromisoformat(fecha_recep_str.split('.')[0])
                dias = (hoy - fecha_recep).days
                if dias <= DIAS_ACEPTACION_TACITA:
                    continue  # Aún pendiente de aprobar
            except:
                continue
        
        # El DTE está aceptado, verificar si está contabilizado
        tipo_dte = dte.get('idTipoDocumento')
        folio = dte.get('folio')
        tipo_interno = TIPO_DTE_A_INTERNO.get(tipo_dte, 'FACE')
        
        # Verificar si existe en documentos
        doc = api_get(rut, f'/documentos/{tipo_interno}/{folio}')
        
        if doc:
            resultado['ya_contabilizados'] += 1
        else:
            # Pendiente de contabilizar
            pendiente = {
                'rut_emisor': dte.get('rutEmisor', ''),
                'emisor': dte.get('emisor', ''),
                'tipo_documento': dte.get('tipoDocumento', ''),
                'tipo_dte': tipo_dte,
                'tipo_interno': tipo_interno,
                'folio': folio,
                'fecha_emision': str(dte.get('fechaEmision', ''))[:10],
                'monto': dte.get('montoTotal', 0),
            }
            resultado['pendientes'].append(pendiente)
            resultado['monto_total'] += pendiente['monto']
    
    resultado['total_pendientes'] = len(resultado['pendientes'])
    
    if mostrar:
        if resultado['pendientes']:
            print(f'\n   {"Emisor":<30} {"Tipo":<15} {"Folio":<10} {"Fecha":<12} {"Total":>15}')
            print(f'   {"-"*84}')
            
            for doc in sorted(resultado['pendientes'], key=lambda x: x['fecha_emision'], reverse=True):
                print(f'   {doc["emisor"][:30]:<30} {doc["tipo_documento"][:15]:<15} {doc["folio"]:<10} {doc["fecha_emision"]:<12} {doc["monto"]:>15,.0f}')
            
            print(f'   {"-"*84}')
            print(f'   {"TOTAL":>69} {resultado["monto_total"]:>15,.0f}')
        else:
            print('\n   ✅ Todos los documentos están contabilizados')
        
        print(f'\n   📊 PENDIENTES: {resultado["total_pendientes"]} | CONTABILIZADOS: {resultado["ya_contabilizados"]}')
    
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN 4: GENERAR BALANCE CONTABLE (EXCEL)
# ═══════════════════════════════════════════════════════════════════════════════

def generar_balance_excel(rut, periodo=None):
    """
    Genera un Excel con el Balance Tributario y Análisis por Cuenta.
    
    Incluye:
    - Hoja "Balance Tributario" con todas las cuentas
    - Una hoja por cada cuenta con movimientos
    - Hipervínculos entre hojas
    """
    try:
        import pandas as pd
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    except ImportError:
        print('\n   ❌ Requiere pandas y openpyxl')
        print('   Instalar: pip install pandas openpyxl')
        return None
    
    config = cargar_config(rut)
    if not config:
        print(f'\n   ❌ No hay configuración para {rut}')
        print(f'   Ejecute primero: python skualo_control.py setup {rut}')
        return None
    
    # Determinar período
    if not periodo:
        periodo = datetime.now().strftime('%Y%m')
    
    # Calcular fecha de corte (último día del mes)
    año = int(periodo[:4])
    mes = int(periodo[4:6])
    if mes == 12:
        fecha_corte = f'{año}-12-31'
    else:
        from datetime import date
        import calendar
        ultimo_dia = calendar.monthrange(año, mes)[1]
        fecha_corte = f'{año}-{mes:02d}-{ultimo_dia:02d}'
    
    print('\n' + '=' * 80)
    print(f'GENERANDO BALANCE CONTABLE - {config["nombre"]}')
    print('=' * 80)
    print(f'\n   RUT: {rut}')
    print(f'   Período: {periodo}')
    print(f'   Fecha Corte: {fecha_corte}')
    
    # ───────────────────────────────────────────────────────────────────────────
    # 1. OBTENER BALANCE TRIBUTARIO
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   1. Obteniendo Balance Tributario...')
    
    balance = api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
    if not balance:
        print('      ❌ No se pudo obtener el balance')
        return None
    
    # Filtrar cuentas con movimientos (debe o haber > 0)
    cuentas_con_movimiento = [c for c in balance if c.get('debe', 0) != 0 or c.get('haber', 0) != 0 or c.get('saldo', 0) != 0]
    cuentas_con_saldo = [c for c in balance if c.get('saldo', 0) != 0]
    print(f'      ✅ {len(balance)} cuentas ({len(cuentas_con_movimiento)} con movimiento, {len(cuentas_con_saldo)} con saldo)')
    
    # ───────────────────────────────────────────────────────────────────────────
    # 2. CREAR EXCEL
    # ───────────────────────────────────────────────────────────────────────────
    print('\n   2. Creando Excel...')
    
    # Crear directorio de salida
    output_dir = Path(__file__).parent / 'generados'
    output_dir.mkdir(exist_ok=True)
    
    # Nombre del archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_empresa = config['nombre'].replace(' ', '_')[:20]
    filename = output_dir / f'Balance_{nombre_empresa}_{periodo}_{timestamp}.xlsx'
    
    # Estilos
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def sanitize_sheet_name(codigo, nombre):
        """Limpiar nombre para hoja Excel (max 31 chars)"""
        name = f"{codigo} {nombre}"
        for char in ['\\', '/', '*', '?', '[', ']', ':']:
            name = name.replace(char, '')
        return name[:31]
    
    # Crear DataFrame del balance
    df_balance = pd.DataFrame(balance)
    
    # Renombrar columnas
    columnas = {
        'idCuenta': 'Código',
        'cuenta': 'Cuenta',
        'tipo': 'Tipo',
        'debitos': 'Débitos',
        'creditos': 'Créditos',
        'debe': 'Debe',
        'haber': 'Haber',
        'saldo': 'Saldo'
    }
    df_balance = df_balance.rename(columns=columnas)
    
    # Seleccionar columnas relevantes
    cols_mostrar = [c for c in ['Código', 'Cuenta', 'Tipo', 'Debe', 'Haber', 'Saldo'] if c in df_balance.columns]
    df_balance = df_balance[cols_mostrar]
    
    # Agregar columna para hipervínculo
    df_balance['Ver Detalle'] = ''
    
    # Guardar en Excel
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Hoja principal: Balance Tributario
        df_balance.to_excel(writer, sheet_name='Balance Tributario', index=False)
        
        ws_balance = writer.sheets['Balance Tributario']
        
        # Formato de encabezados
        for col in range(1, len(df_balance.columns) + 1):
            cell = ws_balance.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Formato de datos
        for row in range(2, len(df_balance) + 2):
            for col in range(1, len(df_balance.columns) + 1):
                cell = ws_balance.cell(row=row, column=col)
                cell.border = border
                
                # Formato numérico
                if col >= 4 and col <= 6:  # Debe, Haber, Saldo
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
        
        # Ancho de columnas
        ws_balance.column_dimensions['A'].width = 12  # Código
        ws_balance.column_dimensions['B'].width = 40  # Cuenta
        ws_balance.column_dimensions['C'].width = 12  # Tipo
        ws_balance.column_dimensions['D'].width = 15  # Debe
        ws_balance.column_dimensions['E'].width = 15  # Haber
        ws_balance.column_dimensions['F'].width = 15  # Saldo
        ws_balance.column_dimensions['G'].width = 12  # Ver Detalle
        
        # ───────────────────────────────────────────────────────────────────────
        # 3. OBTENER ANÁLISIS POR CUENTA
        # ───────────────────────────────────────────────────────────────────────
        print('\n   3. Obteniendo análisis por cuenta...')
        
        cuentas_procesadas = 0
        cuentas_a_procesar = cuentas_con_movimiento if cuentas_con_movimiento else balance
        
        for idx, cuenta in enumerate(cuentas_a_procesar):
            codigo = cuenta.get('idCuenta', '')
            nombre = cuenta.get('cuenta', '')
            
            # Obtener análisis
            analisis = api_get(rut, f'/contabilidad/reportes/analisisporcuenta/{codigo}?fechaCorte={fecha_corte}&soloPendientes=false')
            
            if not analisis:
                continue
            
            # Filtrar movimientos con valor
            movimientos = [m for m in analisis if m.get('saldo', 0) != 0 or m.get('valor', 0) != 0]
            
            if not movimientos:
                continue
            
            # Crear hoja para esta cuenta
            sheet_name = sanitize_sheet_name(codigo, nombre)
            
            df_cuenta = pd.DataFrame(movimientos)
            
            # Renombrar columnas
            col_rename = {
                'comprobante': 'Comp',
                'idTipoDoc': 'Tipo',
                'numDoc': 'N° Doc',
                'auxiliar': 'Auxiliar',
                'emision': 'Emisión',
                'vencimiento': 'Vencimiento',
                'valor': 'Valor',
                'saldo': 'Saldo',
                'glosa': 'Glosa'
            }
            df_cuenta = df_cuenta.rename(columns=col_rename)
            
            # Seleccionar columnas
            cols_cuenta = [c for c in ['Comp', 'Tipo', 'N° Doc', 'Auxiliar', 'Emisión', 'Vencimiento', 'Valor', 'Saldo', 'Glosa'] 
                          if c in df_cuenta.columns]
            df_cuenta = df_cuenta[cols_cuenta]
            
            # Formatear fechas
            for col in ['Emisión', 'Vencimiento']:
                if col in df_cuenta.columns:
                    df_cuenta[col] = df_cuenta[col].apply(lambda x: str(x)[:10] if x else '')
            
            # Guardar hoja
            df_cuenta.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formato de la hoja
            ws_cuenta = writer.sheets[sheet_name]
            
            for col in range(1, len(df_cuenta.columns) + 1):
                cell = ws_cuenta.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            # Agregar hipervínculo en Balance Tributario
            row_balance = None
            for row in range(2, len(df_balance) + 2):
                if ws_balance.cell(row=row, column=1).value == codigo:
                    row_balance = row
                    break
            
            if row_balance:
                cell_link = ws_balance.cell(row=row_balance, column=7)
                cell_link.value = 'Ver Detalle'
                cell_link.hyperlink = f"#'{sheet_name}'!A1"
                cell_link.font = Font(color='0563C1', underline='single')
            
            cuentas_procesadas += 1
            
            # Progreso
            if (cuentas_procesadas % 10 == 0):
                print(f'      Procesadas: {cuentas_procesadas}/{len(cuentas_a_procesar)}')
    
    print(f'\n   ✅ Excel generado: {filename}')
    print(f'      Cuentas con detalle: {cuentas_procesadas}')
    
    return str(filename)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTE COMPLETO
# ═══════════════════════════════════════════════════════════════════════════════

def reporte_completo(rut):
    """Genera un reporte completo con los 3 controles."""
    config = cargar_config(rut)
    if not config:
        print(f'\n   ❌ No hay configuración para {rut}')
        print(f'   Ejecute primero: python skualo_control.py setup {rut}')
        return None
    
    print('=' * 80)
    print(f'REPORTE DE CONTROL - {config["nombre"]}')
    print(f'RUT: {rut}')
    print(f'Fecha: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    print('=' * 80)
    
    # Ejecutar los 3 controles
    r1 = movimientos_bancarios_pendientes(rut, mostrar=True)
    r2 = documentos_por_aprobar_sii(rut, mostrar=True)
    r3 = documentos_por_contabilizar(rut, mostrar=True)
    
    # Resumen
    print('\n' + '=' * 80)
    print('RESUMEN EJECUTIVO')
    print('=' * 80)
    print(f'''
   🏦 Movimientos sin conciliar:     {r1['total_sin_conciliar'] if r1 else 0:>5}
   📄 Documentos por aprobar SII:    {r2['total_pendientes'] if r2 else 0:>5}
   📋 Documentos por contabilizar:   {r3['total_pendientes'] if r3 else 0:>5}
   ✅ Documentos contabilizados:     {r3['ya_contabilizados'] if r3 else 0:>5}
''')
    print('=' * 80)
    
    return {'bancos': r1, 'aprobar': r2, 'contabilizar': r3}


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def mostrar_ayuda():
    print('''
Skualo Control - Sistema de Control y Reportes
===============================================

Uso:
    python skualo_control.py <comando> [rut] [opciones]

CONFIGURACIÓN:
    setup <rut>              Configura una nueva empresa
    listar                   Lista empresas configuradas

CONTROLES DE PENDIENTES:
    bancos <rut>             Movimientos bancarios sin conciliar
    aprobar <rut>            Documentos pendientes de aprobar en SII
    contabilizar <rut>       Documentos pendientes de contabilizar
    reporte <rut>            Reporte completo (los 3 controles)

REPORTES CONTABLES:
    balance <rut> [periodo]  Genera Balance en Excel con análisis por cuenta
                             Período opcional: YYYYMM (ej: 202511)

Ejemplos:
    python skualo_control.py setup 77285542-7
    python skualo_control.py reporte 77949039-4
    python skualo_control.py balance 77285542-7 202511
''')


def main():
    if len(sys.argv) < 2:
        mostrar_ayuda()
        sys.exit(0)
    
    comando = sys.argv[1].lower()
    
    if comando == 'listar':
        listar_empresas_configuradas()
    
    elif comando in ['setup', 'bancos', 'aprobar', 'contabilizar', 'reporte', 'balance']:
        if len(sys.argv) < 3:
            print(f'Error: El comando "{comando}" requiere un RUT')
            print(f'Uso: python skualo_control.py {comando} <RUT>')
            sys.exit(1)
        
        rut = sys.argv[2]
        
        if comando == 'setup':
            setup_empresa(rut)
        elif comando == 'bancos':
            movimientos_bancarios_pendientes(rut)
        elif comando == 'aprobar':
            documentos_por_aprobar_sii(rut)
        elif comando == 'contabilizar':
            documentos_por_contabilizar(rut)
        elif comando == 'reporte':
            reporte_completo(rut)
        elif comando == 'balance':
            periodo = sys.argv[3] if len(sys.argv) > 3 else None
            generar_balance_excel(rut, periodo)
    
    else:
        print(f'Comando desconocido: {comando}')
        mostrar_ayuda()
        sys.exit(1)


if __name__ == '__main__':
    main()

