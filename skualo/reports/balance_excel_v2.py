"""
Balance Tributario + Análisis por Cuenta a Excel - VERSIÓN 2 (Parametrizable)

Uso:
    python balance_excel_v2.py FIDI
    python balance_excel_v2.py CISI
    
Configuración en: config/empresas_config.xlsx (una hoja por empresa)
"""

import os
import sys
import json
import requests
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

API_BASE = "https://api.skualo.cl"
TOKEN = os.getenv("SKUALO_API_TOKEN")
CONFIG_EXCEL = os.path.join(os.path.dirname(__file__), "config", "empresas_config.xlsx")


# ═══════════════════════════════════════════════════════════════════════════════
# API
# ═══════════════════════════════════════════════════════════════════════════════

def api_get(tenant_rut, path):
    """Llamada GET a la API"""
    url = f"{API_BASE}/{tenant_rut}{path}"
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "accept": "application/json"
    }
    response = requests.get(url, headers=headers)
    if response.ok:
        return response.json()
    return None


def get_balance(tenant_rut, id_periodo):
    """Obtener Balance Tributario"""
    return api_get(tenant_rut, f"/contabilidad/reportes/balancetributario/{id_periodo}")


def get_analisis_cuenta(tenant_rut, id_cuenta, fecha_corte):
    """Obtener Análisis por Cuenta"""
    return api_get(tenant_rut, f"/contabilidad/reportes/analisisporcuenta/{id_cuenta}?fechaCorte={fecha_corte}&soloPendientes=false")


# ═══════════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_sheet_name(codigo, nombre):
    """Limpiar nombre para hoja Excel (max 31 chars)"""
    name = f"{codigo} {nombre}"
    for char in ['\\', '/', '*', '?', '[', ']', ':']:
        name = name.replace(char, '')
    return name[:31]


def cargar_config_desde_excel(empresa_key):
    """
    Carga la configuración de una empresa desde el archivo Excel.
    Cada hoja es una empresa diferente.
    """
    wb = load_workbook(CONFIG_EXCEL, data_only=True)
    
    if empresa_key not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{empresa_key}' en {CONFIG_EXCEL}")
    
    ws = wb[empresa_key]
    
    # Leer todas las celdas en un diccionario por sección
    config = {
        "tenant": {},
        "periodos": {"comparativos": []},
        "balance_clasificado": {},
        "estado_resultados": {
            "gastos_operacionales": {},
            "otros_gastos": {}
        },
        "impuesto_renta": {},
        "output": {}
    }
    
    current_section = None
    header_row = None
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        # Obtener valor de primera celda
        first_cell = row[0].value
        
        if first_cell is None:
            continue
        
        first_cell_str = str(first_cell).strip()
        
        # Detectar sección
        if first_cell_str in ["TENANT", "PERIODOS", "PERIODOS_COMPARATIVOS", 
                              "BALANCE_CLASIFICADO", "ESTADO_RESULTADOS", "OUTPUT"]:
            current_section = first_cell_str
            header_row = None
            continue
        
        # Detectar header (fila después de sección)
        if current_section and header_row is None:
            if first_cell_str in ["Campo", "Categoría", "Tipo", "ID Periodo"]:
                header_row = [c.value for c in row]
                continue
        
        # Procesar datos según sección
        if current_section == "TENANT" and header_row:
            campo = row[0].value
            valor = row[1].value
            if campo and valor:
                config["tenant"][campo] = str(valor)
        
        elif current_section == "PERIODOS" and header_row:
            campo = row[0].value
            valor = row[1].value
            if campo and valor:
                if campo == "tasa_impuesto":
                    config["impuesto_renta"]["tasa"] = float(valor)
                else:
                    config["periodos"][campo] = str(valor)
        
        elif current_section == "PERIODOS_COMPARATIVOS" and header_row:
            id_periodo = row[0].value
            nombre = row[1].value
            if id_periodo and nombre:
                config["periodos"]["comparativos"].append({
                    "id": str(id_periodo),
                    "nombre": str(nombre)
                })
        
        elif current_section == "BALANCE_CLASIFICADO" and header_row:
            categoria = row[0].value
            if not categoria:
                continue
            
            nombre = row[1].value or ""
            prefijos = row[2].value or ""
            excluir = row[3].value or ""
            especificas = row[4].value or ""
            desc = row[5].value if len(row) > 5 else ""
            
            cat_config = {"nombre": str(nombre)}
            
            if especificas:
                cat_config["cuentas_especificas"] = [c.strip() for c in str(especificas).split(",")]
                if desc:
                    cat_config["descripcion"] = str(desc)
            else:
                if prefijos:
                    cat_config["prefijos"] = [p.strip() for p in str(prefijos).split(",")]
                if excluir:
                    cat_config["excluir_cuentas"] = [c.strip() for c in str(excluir).split(",")]
            
            config["balance_clasificado"][str(categoria)] = cat_config
        
        elif current_section == "ESTADO_RESULTADOS" and header_row:
            tipo = row[0].value
            key = row[1].value
            nombre = row[2].value
            cuentas = row[3].value
            desc = row[4].value if len(row) > 4 else ""
            
            if not tipo or not key:
                continue
            
            item = {
                "nombre": str(nombre) if nombre else "",
                "cuentas": [c.strip() for c in str(cuentas).split(",")] if cuentas else [],
                "descripcion": str(desc) if desc else ""
            }
            
            if tipo == "ingresos":
                config["estado_resultados"]["ingresos"] = item
            elif tipo == "costo_ventas":
                config["estado_resultados"]["costo_ventas"] = item
            elif tipo == "gastos_operacionales":
                config["estado_resultados"]["gastos_operacionales"][str(key)] = item
            elif tipo == "otros_gastos":
                config["estado_resultados"]["otros_gastos"][str(key)] = item
        
        elif current_section == "OUTPUT" and header_row:
            campo = row[0].value
            valor = row[1].value
            if campo and valor:
                config["output"][str(campo)] = str(valor)
    
    wb.close()
    return config


def clasificar_cuenta(codigo, config_balance):
    """
    Determina a qué categoría pertenece una cuenta según la configuración.
    Retorna el nombre de la categoría o None si no aplica.
    """
    for categoria, reglas in config_balance.items():
        # Primero verificar cuentas específicas
        if "cuentas_especificas" in reglas:
            if codigo in reglas["cuentas_especificas"]:
                return categoria
    
    for categoria, reglas in config_balance.items():
        if "cuentas_especificas" in reglas:
            continue  # Ya procesadas arriba
            
        # Verificar si está excluida
        excluidas = reglas.get("excluir_cuentas", [])
        if codigo in excluidas:
            continue
            
        # Verificar prefijos
        prefijos = reglas.get("prefijos", [])
        for prefijo in prefijos:
            if codigo.startswith(prefijo):
                return categoria
    
    return None


def calcular_eerr(balance, config_eerr):
    """Calcula los valores del Estado de Resultados según la configuración"""
    cuentas_dict = {c["idCuenta"]: c for c in balance}
    
    def get_valor(cuenta_id):
        if cuenta_id in cuentas_dict:
            c = cuentas_dict[cuenta_id]
            return c.get("ganancias", 0) - c.get("perdidas", 0)
        return 0
    
    def suma_cuentas(cuentas_list):
        return sum(get_valor(c) for c in cuentas_list)
    
    resultado = {}
    
    # Ingresos y Costo de Ventas
    resultado["ingresos"] = suma_cuentas(config_eerr["ingresos"]["cuentas"])
    resultado["costo_ventas"] = suma_cuentas(config_eerr["costo_ventas"]["cuentas"])
    resultado["utilidad_bruta"] = resultado["ingresos"] + resultado["costo_ventas"]
    
    # Gastos Operacionales
    resultado["gastos_op"] = {}
    total_gastos_op = 0
    for key, cfg in config_eerr.get("gastos_operacionales", {}).items():
        valor = suma_cuentas(cfg["cuentas"])
        resultado["gastos_op"][key] = {"nombre": cfg["nombre"], "valor": valor}
        total_gastos_op += valor
    resultado["total_gastos_op"] = total_gastos_op
    
    resultado["resultado_operacional"] = resultado["utilidad_bruta"] + total_gastos_op
    
    # Otros Gastos
    resultado["otros_gastos"] = {}
    total_otros = 0
    for key, cfg in config_eerr.get("otros_gastos", {}).items():
        valor = suma_cuentas(cfg["cuentas"])
        resultado["otros_gastos"][key] = {"nombre": cfg["nombre"], "valor": valor}
        total_otros += valor
    resultado["total_otros_gastos"] = total_otros
    
    resultado["resultado_antes_impuestos"] = resultado["resultado_operacional"] + total_otros
    
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS GLOBALES
# ═══════════════════════════════════════════════════════════════════════════════

ESTILOS = {
    "font_header": Font(bold=True, size=10, color="FFFFFF"),
    "fill_header": PatternFill(start_color="006400", end_color="006400", fill_type="solid"),
    "font_titulo": Font(bold=True, size=14, color="006400"),
    "font_section": Font(bold=True, size=13, color="006400"),
    "fill_section": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "font_category": Font(bold=True, size=11),
    "font_subcategory": Font(bold=True, italic=True, size=10),
    "font_subtotal": Font(bold=True, size=10),
    "font_total": Font(bold=True, size=11),
    "font_total_final": Font(bold=True, size=12, color="FFFFFF"),
    "fill_total_final": PatternFill(start_color="006400", end_color="006400", fill_type="solid"),
    "fill_subtotal": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "thin_border": Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    ),
    "formato_miles": '#,##0'
}


# ═══════════════════════════════════════════════════════════════════════════════
# GENERADORES DE HOJAS
# ═══════════════════════════════════════════════════════════════════════════════

def crear_resumen(balance, writer, config):
    """Crea la hoja Resumen con Balance Clasificado, Estado de Resultados y KPIs"""
    
    tenant_name = config["tenant"]["nombre"]
    periodo = config["periodos"]["actual"]
    config_balance = config["balance_clasificado"]
    config_eerr = config["estado_resultados"]
    tasa_impuesto = config["impuesto_renta"]["tasa"]
    
    cuentas_dict = {c["idCuenta"]: c for c in balance}
    
    # ═══════════════════════════════════════════════════════════
    # CLASIFICAR CUENTAS
    # ═══════════════════════════════════════════════════════════
    
    categorias = {cat: [] for cat in config_balance.keys()}
    
    for c in balance:
        codigo = c["idCuenta"]
        categoria = clasificar_cuenta(codigo, config_balance)
        
        if categoria:
            # Determinar si es activo o pasivo
            if categoria in ["activo_corriente", "activo_no_corriente", "intangibles"]:
                if c["activos"] != 0:
                    categorias[categoria].append((c["cuenta"], c["activos"]))
            elif categoria == "patrimonio":
                valor = c["pasivos"] if c["pasivos"] != 0 else -c["activos"]
                if valor != 0:
                    categorias[categoria].append((c["cuenta"], valor))
            else:  # pasivos
                if c["pasivos"] != 0:
                    categorias[categoria].append((c["cuenta"], c["pasivos"]))
    
    # Calcular totales
    totales = {cat: sum(v for _, v in items) for cat, items in categorias.items()}
    
    total_activos = totales.get("activo_corriente", 0) + totales.get("activo_no_corriente", 0) + totales.get("intangibles", 0)
    total_pasivos = totales.get("pasivo_corriente", 0) + totales.get("pasivo_no_corriente", 0)
    patrimonio_sin_resultado = totales.get("patrimonio", 0)
    
    # ═══════════════════════════════════════════════════════════
    # ESTADO DE RESULTADOS
    # ═══════════════════════════════════════════════════════════
    
    eerr = calcular_eerr(balance, config_eerr)
    
    impuesto = -eerr["resultado_antes_impuestos"] * tasa_impuesto if eerr["resultado_antes_impuestos"] > 0 else 0
    resultado_neto = eerr["resultado_antes_impuestos"] + impuesto
    
    # ═══════════════════════════════════════════════════════════
    # PATRIMONIO TOTAL (incluye resultado del período)
    # ═══════════════════════════════════════════════════════════
    total_patrimonio = patrimonio_sin_resultado + resultado_neto
    
    # Verificación de cuadratura
    diferencia_cuadratura = total_activos - (total_pasivos + total_patrimonio)
    cuadra = abs(diferencia_cuadratura) < 1  # Tolerancia de $1 por redondeos
    
    # ═══════════════════════════════════════════════════════════
    # KPIs
    # ═══════════════════════════════════════════════════════════
    
    ingresos = eerr["ingresos"] if eerr["ingresos"] != 0 else 1
    kpis = {
        "margen_bruto": (eerr["utilidad_bruta"] / ingresos * 100) if eerr["ingresos"] != 0 else 0,
        "margen_operacional": (eerr["resultado_operacional"] / ingresos * 100) if eerr["ingresos"] != 0 else 0,
        "margen_neto": (resultado_neto / ingresos * 100) if eerr["ingresos"] != 0 else 0,
        "roa": (resultado_neto / total_activos * 100) if total_activos != 0 else 0,
        "roe": (resultado_neto / total_patrimonio * 100) if total_patrimonio != 0 else 0,
        "ratio_deuda": (total_pasivos / total_activos * 100) if total_activos != 0 else 0,
    }
    
    # ═══════════════════════════════════════════════════════════
    # CONSTRUIR HOJA
    # ═══════════════════════════════════════════════════════════
    
    rows = []
    row_types = []
    
    # Fila 1: Navegación
    rows.append(["", "", "", "", ""])
    row_types.append("nav_row")
    
    # Header
    periodo_texto = f"{periodo[:4]}-{periodo[4:]}"
    rows.append([f"RESUMEN FINANCIERO - {tenant_name}", "", "", "", f"Período: {periodo_texto}"])
    row_types.append("header_main")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # BALANCE CLASIFICADO
    # ─────────────────────────────────────────────────────────
    rows.append(["BALANCE CLASIFICADO", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    # ACTIVOS
    rows.append(["ACTIVOS", "", "", "", ""])
    row_types.append("category")
    
    for cat_key in ["activo_corriente", "activo_no_corriente", "intangibles"]:
        if cat_key in config_balance:
            cat_nombre = config_balance[cat_key]["nombre"]
            items = categorias.get(cat_key, [])
            
            rows.append([f"  {cat_nombre}", "", "", "", ""])
            row_types.append("subcategory")
            
            for nombre, valor in items[:5]:
                rows.append([f"    {nombre}", valor, "", "", ""])
                row_types.append("item")
            
            if len(items) > 5:
                otros = sum(v for _, v in items[5:])
                rows.append([f"    Otros {cat_nombre.lower()}", otros, "", "", ""])
                row_types.append("item")
            
            rows.append([f"  Total {cat_nombre}", totales.get(cat_key, 0), "", "", ""])
            row_types.append("subtotal")
    
    rows.append(["TOTAL ACTIVOS", total_activos, "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # PASIVOS
    rows.append(["PASIVOS", "", "", "", ""])
    row_types.append("category")
    
    for cat_key in ["pasivo_corriente", "pasivo_no_corriente"]:
        if cat_key in config_balance:
            cat_nombre = config_balance[cat_key]["nombre"]
            items = categorias.get(cat_key, [])
            
            rows.append([f"  {cat_nombre}", "", "", "", ""])
            row_types.append("subcategory")
            
            for nombre, valor in items[:5]:
                rows.append([f"    {nombre}", valor, "", "", ""])
                row_types.append("item")
            
            if len(items) > 5:
                otros = sum(v for _, v in items[5:])
                rows.append([f"    Otros {cat_nombre.lower()}", otros, "", "", ""])
                row_types.append("item")
            
            rows.append([f"  Total {cat_nombre}", totales.get(cat_key, 0), "", "", ""])
            row_types.append("subtotal")
    
    rows.append(["TOTAL PASIVOS", total_pasivos, "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # PATRIMONIO
    rows.append(["PATRIMONIO", "", "", "", ""])
    row_types.append("category")
    for nombre, valor in categorias.get("patrimonio", []):
        rows.append([f"  {nombre}", valor, "", "", ""])
        row_types.append("item")
    # Agregar resultado del período
    rows.append([f"  Resultado del Período", resultado_neto, "", "", ""])
    row_types.append("item")
    rows.append([f"  (Patrimonio sin resultado)", patrimonio_sin_resultado, "", "", "(referencia)"])
    row_types.append("item")
    rows.append(["TOTAL PATRIMONIO", total_patrimonio, "", "", ""])
    row_types.append("total")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["TOTAL PASIVOS + PATRIMONIO", total_pasivos + total_patrimonio, "", "", ""])
    row_types.append("total_final")
    
    # Verificación de cuadratura
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    if cuadra:
        rows.append(["✅ CUADRATURA OK: Activos = Pasivos + Patrimonio", "", "", "", ""])
        row_types.append("verification_ok")
    else:
        rows.append([f"⚠️ DESCUADRE: Diferencia = ${diferencia_cuadratura:,.0f}", "", "", "", ""])
        row_types.append("verification_error")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # ESTADO DE RESULTADOS
    # ─────────────────────────────────────────────────────────
    rows.append(["ESTADO DE RESULTADOS", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    rows.append([config_eerr["ingresos"]["nombre"], eerr["ingresos"], "", "", ""])
    row_types.append("item")
    rows.append([config_eerr["costo_ventas"]["nombre"], eerr["costo_ventas"], "", "", ""])
    row_types.append("item")
    rows.append(["UTILIDAD BRUTA", eerr["utilidad_bruta"], "", "", ""])
    row_types.append("subtotal")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["Gastos Operacionales:", "", "", "", ""])
    row_types.append("subcategory")
    for key, data in eerr["gastos_op"].items():
        rows.append([f"  {data['nombre']}", data["valor"], "", "", ""])
        row_types.append("item")
    rows.append(["Total Gastos Operacionales", eerr["total_gastos_op"], "", "", ""])
    row_types.append("subtotal")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["RESULTADO OPERACIONAL (EBIT)", eerr["resultado_operacional"], "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["Otros Gastos:", "", "", "", ""])
    row_types.append("subcategory")
    for key, data in eerr["otros_gastos"].items():
        rows.append([f"  {data['nombre']}", data["valor"], "", "", ""])
        row_types.append("item")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["RESULTADO ANTES DE IMPUESTOS", eerr["resultado_antes_impuestos"], "", "", ""])
    row_types.append("total")
    rows.append([f"Impuesto a la Renta ({int(tasa_impuesto*100)}%)", impuesto, "", "", ""])
    row_types.append("item")
    rows.append(["RESULTADO NETO", resultado_neto, "", "", ""])
    row_types.append("total_final")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # KPIs
    # ─────────────────────────────────────────────────────────
    rows.append(["INDICADORES FINANCIEROS (KPIs)", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    rows.append(["", "Valor", "Interpretación", "", ""])
    row_types.append("header_kpi")
    
    rows.append(["Margen Bruto", f"{kpis['margen_bruto']:.1f}%", "Eficiencia en costos directos", "", ""])
    row_types.append("kpi")
    rows.append(["Margen Operacional", f"{kpis['margen_operacional']:.1f}%", "Rentabilidad operativa", "", ""])
    row_types.append("kpi")
    rows.append(["Margen Neto", f"{kpis['margen_neto']:.1f}%", "Rentabilidad final", "", ""])
    row_types.append("kpi")
    rows.append(["ROA", f"{kpis['roa']:.1f}%", "Retorno sobre activos", "", ""])
    row_types.append("kpi")
    rows.append(["ROE", f"{kpis['roe']:.1f}%", "Retorno sobre patrimonio", "", ""])
    row_types.append("kpi")
    rows.append(["Ratio de Endeudamiento", f"{kpis['ratio_deuda']:.1f}%", "Nivel de apalancamiento", "", ""])
    row_types.append("kpi")
    
    # Escribir a Excel
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Resumen", index=False, header=False)
    
    ws = writer.sheets["Resumen"]
    
    # Aplicar formatos
    for row_idx, row_type in enumerate(row_types, start=1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)
        cell_e = ws.cell(row=row_idx, column=5)
        
        if row_type == "header_main":
            cell_a.font = Font(bold=True, size=16, color="006400")
        elif row_type == "section_title":
            cell_a.font = ESTILOS["font_section"]
            cell_a.fill = ESTILOS["fill_section"]
            cell_e.hyperlink = "#'Documentación'!A1"
            cell_e.style = "Hyperlink"
        elif row_type == "category":
            cell_a.font = ESTILOS["font_category"]
        elif row_type == "subcategory":
            cell_a.font = Font(bold=True, italic=True)
        elif row_type == "subtotal":
            cell_a.font = ESTILOS["font_subtotal"]
            cell_b.font = ESTILOS["font_subtotal"]
        elif row_type == "total":
            cell_a.font = ESTILOS["font_total"]
            cell_b.font = ESTILOS["font_total"]
        elif row_type == "total_final":
            cell_a.font = ESTILOS["font_total_final"]
            cell_a.fill = ESTILOS["fill_total_final"]
            cell_b.font = ESTILOS["font_total_final"]
            cell_b.fill = ESTILOS["fill_total_final"]
        elif row_type == "verification_ok":
            cell_a.font = Font(bold=True, color="006400")  # Verde
        elif row_type == "verification_error":
            cell_a.font = Font(bold=True, color="FF0000")  # Rojo
        elif row_type == "header_kpi":
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
            ws.cell(row=row_idx, column=3).font = Font(bold=True)
        elif row_type == "kpi":
            cell_b.font = Font(bold=True, color="006400")
        
        if isinstance(rows[row_idx-1][1], (int, float)):
            cell_b.number_format = ESTILOS["formato_miles"]
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['E'].width = 22
    
    # Navegación en A1
    ws.cell(row=1, column=1).value = "📊 DASHBOARD PRINCIPAL"
    ws.cell(row=1, column=1).font = Font(bold=True, size=10, color="666666")
    
    return {
        "ingresos": eerr["ingresos"],
        "utilidad_bruta": eerr["utilidad_bruta"],
        "resultado_operacional": eerr["resultado_operacional"],
        "resultado_neto": resultado_neto,
        "total_activos": total_activos,
        "total_pasivos": total_pasivos,
        "total_patrimonio": total_patrimonio
    }


def crear_eeff_comparativos(tenant_rut, config, writer):
    """Crea hoja de Estados Financieros Comparativos"""
    
    periodos = config["periodos"]["comparativos"]
    config_balance = config["balance_clasificado"]
    config_eerr = config["estado_resultados"]
    tasa_impuesto = config["impuesto_renta"]["tasa"]
    
    # Obtener balances
    balances = {}
    for periodo in periodos:
        print(f"   📊 Obteniendo balance {periodo['nombre']}...")
        balance = get_balance(tenant_rut, periodo["id"])
        if balance:
            balances[periodo["nombre"]] = {c["idCuenta"]: c for c in balance}
    
    if not balances:
        return
    
    todas_cuentas = set()
    for balance_dict in balances.values():
        todas_cuentas.update(balance_dict.keys())
    cuentas_ordenadas = sorted(todas_cuentas)
    
    nombres_periodos = [p["nombre"] for p in periodos]
    
    # Construir datos
    rows = []
    row_types = []
    
    # Navegación y título
    rows.append(["← Volver al Resumen", ""] + [""] * len(periodos))
    row_types.append("nav_row")
    rows.append(["ESTADOS FINANCIEROS COMPARATIVOS", ""] + [""] * len(periodos))
    row_types.append("titulo_principal")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Header
    header = ["Código", "Cuenta"] + nombres_periodos
    rows.append(header)
    row_types.append("header")
    
    # Totales por período y categoría
    totales = {nombre: {cat: 0 for cat in config_balance.keys()} for nombre in nombres_periodos}
    
    # ─────────────────────────────────────────────────────────
    # BALANCE GENERAL
    # ─────────────────────────────────────────────────────────
    rows.append(["", "BALANCE GENERAL"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ACTIVOS
    rows.append(["", "ACTIVOS"] + [""] * len(periodos))
    row_types.append("category")
    
    for cat_key in ["activo_corriente", "activo_no_corriente", "intangibles"]:
        if cat_key not in config_balance:
            continue
            
        cat_cfg = config_balance[cat_key]
        cat_nombre = cat_cfg["nombre"]
        
        rows.append(["", f"  {cat_nombre}"] + [""] * len(periodos))
        row_types.append("subcategory")
        
        for codigo in cuentas_ordenadas:
            categoria = clasificar_cuenta(codigo, config_balance)
            if categoria != cat_key:
                continue
            
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for nombre_periodo in nombres_periodos:
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["activos"]
                    valores.append(val)
                    totales[nombre_periodo][cat_key] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
        
        rows.append(["", f"  Total {cat_nombre}"] + [totales[n][cat_key] for n in nombres_periodos])
        row_types.append("subtotal")
    
    # Total Activos
    total_activos_por_periodo = [
        totales[n].get("activo_corriente", 0) + 
        totales[n].get("activo_no_corriente", 0) + 
        totales[n].get("intangibles", 0) 
        for n in nombres_periodos
    ]
    rows.append(["", "TOTAL ACTIVOS"] + total_activos_por_periodo)
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # PASIVOS
    rows.append(["", "PASIVOS"] + [""] * len(periodos))
    row_types.append("category")
    
    for cat_key in ["pasivo_corriente", "pasivo_no_corriente"]:
        if cat_key not in config_balance:
            continue
            
        cat_cfg = config_balance[cat_key]
        cat_nombre = cat_cfg["nombre"]
        
        rows.append(["", f"  {cat_nombre}"] + [""] * len(periodos))
        row_types.append("subcategory")
        
        for codigo in cuentas_ordenadas:
            categoria = clasificar_cuenta(codigo, config_balance)
            if categoria != cat_key:
                continue
            
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for nombre_periodo in nombres_periodos:
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["pasivos"]
                    valores.append(val)
                    totales[nombre_periodo][cat_key] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
        
        rows.append(["", f"  Total {cat_nombre}"] + [totales[n][cat_key] for n in nombres_periodos])
        row_types.append("subtotal")
    
    # Total Pasivos
    total_pasivos_por_periodo = [
        totales[n].get("pasivo_corriente", 0) + totales[n].get("pasivo_no_corriente", 0)
        for n in nombres_periodos
    ]
    rows.append(["", "TOTAL PASIVOS"] + total_pasivos_por_periodo)
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # PATRIMONIO
    rows.append(["", "PATRIMONIO"] + [""] * len(periodos))
    row_types.append("category")
    
    for codigo in cuentas_ordenadas:
        categoria = clasificar_cuenta(codigo, config_balance)
        if categoria != "patrimonio":
            continue
        
        nombre_cuenta = ""
        for balance_dict in balances.values():
            if codigo in balance_dict:
                nombre_cuenta = balance_dict[codigo]["cuenta"]
                break
        
        valores = []
        tiene_valor = False
        for nombre_periodo in nombres_periodos:
            if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                c = balances[nombre_periodo][codigo]
                val = c["pasivos"] if c["pasivos"] != 0 else -c["activos"]
                valores.append(val)
                totales[nombre_periodo]["patrimonio"] += val
                if val != 0:
                    tiene_valor = True
            else:
                valores.append(0)
        
        if tiene_valor:
            rows.append([codigo, f"  {nombre_cuenta}"] + valores)
            row_types.append("item")
    
    rows.append(["", "TOTAL PATRIMONIO"] + [totales[n]["patrimonio"] for n in nombres_periodos])
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Total Pasivos + Patrimonio
    total_pp = [total_pasivos_por_periodo[i] + totales[nombres_periodos[i]]["patrimonio"] for i in range(len(periodos))]
    rows.append(["", "TOTAL PASIVOS + PATRIMONIO"] + total_pp)
    row_types.append("total_final")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # ESTADO DE RESULTADOS
    # ─────────────────────────────────────────────────────────
    rows.append(["", "ESTADO DE RESULTADOS"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Calcular EERR por período
    eerr_periodos = {}
    for nombre_periodo in nombres_periodos:
        if nombre_periodo in balances:
            balance_list = [{"idCuenta": k, **v} for k, v in balances[nombre_periodo].items()]
            eerr_periodos[nombre_periodo] = calcular_eerr(balance_list, config_eerr)
        else:
            eerr_periodos[nombre_periodo] = {}
    
    def eerr_row(concepto, campo, tipo="item"):
        valores = [eerr_periodos.get(n, {}).get(campo, 0) for n in nombres_periodos]
        rows.append(["", concepto] + valores)
        row_types.append(tipo)
    
    eerr_row(config_eerr["ingresos"]["nombre"], "ingresos")
    eerr_row(config_eerr["costo_ventas"]["nombre"], "costo_ventas")
    eerr_row("UTILIDAD BRUTA", "utilidad_bruta", "subtotal")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Gastos Operacionales:"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for key, cfg in config_eerr.get("gastos_operacionales", {}).items():
        valores = []
        for n in nombres_periodos:
            eerr = eerr_periodos.get(n, {})
            gastos_op = eerr.get("gastos_op", {})
            valores.append(gastos_op.get(key, {}).get("valor", 0))
        rows.append(["", f"  {cfg['nombre']}"] + valores)
        row_types.append("item")
    
    eerr_row("Total Gastos Operacionales", "total_gastos_op", "subtotal")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    eerr_row("RESULTADO OPERACIONAL (EBIT)", "resultado_operacional", "total")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Otros Gastos:"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for key, cfg in config_eerr.get("otros_gastos", {}).items():
        valores = []
        for n in nombres_periodos:
            eerr = eerr_periodos.get(n, {})
            otros = eerr.get("otros_gastos", {})
            valores.append(otros.get(key, {}).get("valor", 0))
        rows.append(["", f"  {cfg['nombre']}"] + valores)
        row_types.append("item")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    eerr_row("RESULTADO ANTES DE IMPUESTOS", "resultado_antes_impuestos", "total")
    
    # Impuesto y Resultado Neto
    impuestos = []
    resultados_netos = []
    for n in nombres_periodos:
        rai = eerr_periodos.get(n, {}).get("resultado_antes_impuestos", 0)
        imp = -rai * tasa_impuesto if rai > 0 else 0
        impuestos.append(imp)
        resultados_netos.append(rai + imp)
    
    rows.append(["", f"Impuesto a la Renta ({int(tasa_impuesto*100)}%)"] + impuestos)
    row_types.append("item")
    rows.append(["", "RESULTADO NETO"] + resultados_netos)
    row_types.append("total_final")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # KPIs
    # ─────────────────────────────────────────────────────────
    rows.append(["", "INDICADORES FINANCIEROS (KPIs)"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    def kpi_row(nombre, valores_list):
        rows.append(["", nombre] + valores_list)
        row_types.append("kpi")
    
    # Calcular KPIs por período
    kpis_periodos = {n: {} for n in nombres_periodos}
    for i, n in enumerate(nombres_periodos):
        eerr = eerr_periodos.get(n, {})
        ingresos = eerr.get("ingresos", 0) or 1
        t_activos = total_activos_por_periodo[i] or 1
        t_patrimonio = totales[n]["patrimonio"] or 1
        t_pasivos = total_pasivos_por_periodo[i]
        r_neto = resultados_netos[i]
        
        kpis_periodos[n] = {
            "margen_bruto": f"{(eerr.get('utilidad_bruta', 0) / ingresos * 100):.1f}%" if eerr.get("ingresos", 0) else "0.0%",
            "margen_op": f"{(eerr.get('resultado_operacional', 0) / ingresos * 100):.1f}%" if eerr.get("ingresos", 0) else "0.0%",
            "margen_neto": f"{(r_neto / ingresos * 100):.1f}%" if eerr.get("ingresos", 0) else "0.0%",
            "roa": f"{(r_neto / t_activos * 100):.1f}%",
            "roe": f"{(r_neto / t_patrimonio * 100):.1f}%",
            "ratio_deuda": f"{(t_pasivos / t_activos * 100):.1f}%"
        }
    
    rows.append(["", "Márgenes de Rentabilidad"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  Margen Bruto", [kpis_periodos[n]["margen_bruto"] for n in nombres_periodos])
    kpi_row("  Margen Operacional", [kpis_periodos[n]["margen_op"] for n in nombres_periodos])
    kpi_row("  Margen Neto", [kpis_periodos[n]["margen_neto"] for n in nombres_periodos])
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Rentabilidad sobre Inversión"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  ROA", [kpis_periodos[n]["roa"] for n in nombres_periodos])
    kpi_row("  ROE", [kpis_periodos[n]["roe"] for n in nombres_periodos])
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Estructura Financiera"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  Ratio de Endeudamiento", [kpis_periodos[n]["ratio_deuda"] for n in nombres_periodos])
    
    # Escribir a Excel
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="EEFF Comparativos", index=False, header=False)
    
    ws = writer.sheets["EEFF Comparativos"]
    num_cols = 2 + len(periodos)
    
    # Aplicar formatos
    for row_idx, row_type in enumerate(row_types, start=1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if row_type == "nav_row":
                if col_idx == 1:
                    cell.hyperlink = "#'Resumen'!A1"
                    cell.style = "Hyperlink"
            elif row_type == "titulo_principal":
                if col_idx == 1:
                    cell.font = ESTILOS["font_titulo"]
            elif row_type == "header":
                cell.font = ESTILOS["font_header"]
                cell.fill = ESTILOS["fill_header"]
                cell.alignment = Alignment(horizontal='center')
            elif row_type == "section_title":
                cell.font = ESTILOS["font_section"]
                cell.fill = ESTILOS["fill_section"]
            elif row_type == "category":
                cell.font = ESTILOS["font_category"]
            elif row_type == "subcategory":
                cell.font = ESTILOS["font_subcategory"]
            elif row_type == "subtotal":
                cell.font = ESTILOS["font_subtotal"]
                cell.fill = ESTILOS["fill_subtotal"]
            elif row_type == "total":
                cell.font = ESTILOS["font_total"]
            elif row_type == "total_final":
                cell.font = ESTILOS["font_total_final"]
                cell.fill = ESTILOS["fill_total_final"]
            elif row_type == "kpi":
                if col_idx >= 3:
                    cell.font = Font(bold=True, color="006400")
                    cell.alignment = Alignment(horizontal='center')
            
            # Formato miles
            if col_idx >= 3 and row_idx > 1:
                val = rows[row_idx-1][col_idx-1] if col_idx-1 < len(rows[row_idx-1]) else None
                if isinstance(val, (int, float)):
                    cell.number_format = ESTILOS["formato_miles"]
    
    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    for i in range(len(periodos)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16


def cargar_documentacion():
    """Carga la documentación desde el archivo externo"""
    doc_path = os.path.join(os.path.dirname(__file__), "config", "documentacion.json")
    with open(doc_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def crear_documentacion(writer, config):
    """Crea la hoja de documentación combinando archivo externo + config empresa"""
    
    # Cargar documentación base
    doc = cargar_documentacion()
    
    config_balance = config["balance_clasificado"]
    config_eerr = config["estado_resultados"]
    tasa_impuesto = config["impuesto_renta"]["tasa"]
    tenant_name = config["tenant"]["nombre"]
    
    data = [
        ("← Volver al Resumen", "", "", ""),
        ("", "", "", ""),
        (f"{doc['titulo']} - {tenant_name}", "", "", ""),
        ("", "", "", ""),
    ]
    
    # Procesar secciones del archivo externo
    for seccion in doc["secciones"]:
        data.append(("═" * 60, "", "", ""))
        data.append((seccion["titulo"], "", "", ""))
        data.append(("═" * 60, "", "", ""))
        
        if seccion.get("descripcion"):
            data.append((seccion["descripcion"], "", "", ""))
        data.append(("", "", "", ""))
        
        # Secciones dinámicas que leen de config empresa
        if seccion.get("tipo") == "config_dinamica":
            if seccion["id"] == "balance_clasificado":
                for cat_key, cat_cfg in config_balance.items():
                    nombre = cat_cfg["nombre"]
                    if "cuentas_especificas" in cat_cfg:
                        cuentas = ", ".join(cat_cfg["cuentas_especificas"])
                        desc = cat_cfg.get("descripcion", "Cuentas específicas")
                    else:
                        prefijos = ", ".join(cat_cfg.get("prefijos", []))
                        excluidas = cat_cfg.get("excluir_cuentas", [])
                        desc = f"Prefijos: {prefijos}"
                        if excluidas:
                            desc += f" (excluye: {', '.join(excluidas)})"
                        cuentas = prefijos
                    data.append((nombre, cuentas, desc, ""))
                    
            elif seccion["id"] == "estado_resultados":
                data.append((config_eerr["ingresos"]["nombre"], 
                           ", ".join(config_eerr["ingresos"]["cuentas"]), 
                           config_eerr["ingresos"]["descripcion"], ""))
                data.append((config_eerr["costo_ventas"]["nombre"], 
                           ", ".join(config_eerr["costo_ventas"]["cuentas"]), 
                           config_eerr["costo_ventas"]["descripcion"], ""))
                
                for key, cfg in config_eerr.get("gastos_operacionales", {}).items():
                    data.append((cfg["nombre"], ", ".join(cfg["cuentas"]), cfg["descripcion"], ""))
                
                for key, cfg in config_eerr.get("otros_gastos", {}).items():
                    data.append((cfg["nombre"], ", ".join(cfg["cuentas"]), cfg["descripcion"], ""))
        
        # Secciones con items del archivo externo
        elif "items" in seccion:
            if seccion["id"] == "calculos_eerr":
                for item in seccion["items"]:
                    formula = item["formula"]
                    if "Impuesto" in item["concepto"]:
                        formula = f"{int(tasa_impuesto*100)}{formula}"
                    data.append((item["concepto"], "=", formula, ""))
                    
            elif seccion["id"] == "kpis":
                for item in seccion["items"]:
                    data.append((item["nombre"], item["formula"], item["interpretacion"], ""))
                    
            elif seccion["id"] == "notas":
                for item in seccion["items"]:
                    data.append((f"• {item['nota']}", "", "", ""))
        
        data.append(("", "", "", ""))
    
    # Metadata
    data.append(("─" * 60, "", "", ""))
    data.append((f"Versión: {doc['metadata']['version']}", 
                f"Actualizado: {doc['metadata']['ultima_actualizacion']}", 
                f"Fuente: {doc['metadata']['fuente']}", ""))
    
    df = pd.DataFrame(data, columns=["Concepto", "Cuentas/Fórmula", "Descripción", ""])
    df.to_excel(writer, sheet_name="Documentación", index=False, header=False)
    
    ws = writer.sheets["Documentación"]
    
    # Navegación
    ws.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    
    # Título
    ws.cell(row=3, column=1).font = ESTILOS["font_titulo"]
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main(empresa_key):
    print("═" * 60)
    print("   BALANCE + ANÁLISIS POR CUENTA A EXCEL (V2)")
    print("═" * 60)
    
    # Cargar configuración desde Excel
    print(f"\n📁 Cargando configuración: {CONFIG_EXCEL} → Hoja: {empresa_key}")
    config = cargar_config_desde_excel(empresa_key)
    
    tenant = config["tenant"]
    periodo = config["periodos"]["actual"]
    fecha_corte = config["periodos"]["fecha_corte"]
    output_dir = config["output"]["carpeta"]
    prefijo = config["output"]["prefijo_archivo"]
    
    print(f"   Tenant: {tenant['nombre']} ({tenant['rut']})")
    print(f"   Período: {periodo}")
    
    # Obtener Balance
    print("\n📊 Obteniendo Balance Tributario...")
    balance = get_balance(tenant["rut"], periodo)
    
    if not balance:
        print("   ❌ Error obteniendo balance")
        return
    
    balance_filtrado = [
        c for c in balance
        if c["deudor"] != 0 or c["acreedor"] != 0 or
           c["activos"] != 0 or c["pasivos"] != 0 or
           c["perdidas"] != 0 or c["ganancias"] != 0
    ]
    print(f"   ✅ {len(balance_filtrado)} cuentas (de {len(balance)} totales)")
    
    # Crear DataFrame del balance
    df_balance = pd.DataFrame(balance_filtrado)
    df_balance = df_balance.rename(columns={
        "idCuenta": "Código", "cuenta": "Cuenta", "debitos": "Débitos",
        "creditos": "Créditos", "deudor": "Saldo Deudor", "acreedor": "Saldo Acreedor",
        "activos": "Activos", "pasivos": "Pasivos", "perdidas": "Pérdidas", "ganancias": "Ganancias"
    })
    df_balance["Ver Detalle"] = ""
    
    # Crear archivo
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{output_dir}/{prefijo}_{tenant['key']}_{periodo}_{timestamp}.xlsx"
    
    cuenta_a_hoja = {}
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 1. Resumen
        print("\n📈 Generando Resumen...")
        crear_resumen(balance, writer, config)
        
        # 2. EEFF Comparativos
        print("\n📊 Generando Estados Financieros Comparativos...")
        crear_eeff_comparativos(tenant["rut"], config, writer)
        
        # 3. Documentación
        print("\n📝 Generando Documentación...")
        crear_documentacion(writer, config)
        
        # 4. Balance Tributario
        periodo_texto = f"{periodo[:4]}-{periodo[4:]}"
        df_balance.to_excel(writer, sheet_name="Balance Tributario", index=False, startrow=2)
        
        ws_balance = writer.sheets["Balance Tributario"]
        ws_balance.cell(row=1, column=1).value = "← Volver al Resumen"
        ws_balance.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
        ws_balance.cell(row=1, column=1).style = "Hyperlink"
        ws_balance.cell(row=2, column=1).value = f"BALANCE TRIBUTARIO - {tenant['nombre']} - Período: {periodo_texto}"
        ws_balance.cell(row=2, column=1).font = Font(bold=True, size=12, color="006400")
        
        # 5. Hojas de análisis por cuenta
        print("\n📋 Procesando cuentas con movimientos...")
        cuentas_con_datos = 0
        sheet_names = {"Balance Tributario", "Resumen", "EEFF Comparativos", "Documentación"}
        
        for cuenta in balance_filtrado:
            analisis = get_analisis_cuenta(tenant["rut"], cuenta["idCuenta"], fecha_corte)
            
            if analisis and len(analisis) > 0:
                df_cuenta = pd.DataFrame(analisis)
                if "saldo" in df_cuenta.columns:
                    df_cuenta = df_cuenta[df_cuenta["saldo"] != 0]
                
                if len(df_cuenta) == 0:
                    continue
                
                # Seleccionar columnas
                cols_deseadas = ["fecha", "numero", "tipo", "glosa", "debe", "haber", "saldo"]
                cols_disponibles = [c for c in cols_deseadas if c in df_cuenta.columns]
                
                if "auxiliar" in df_cuenta.columns:
                    cols_disponibles.insert(3, "auxiliar")
                elif "idAuxiliar" in df_cuenta.columns:
                    df_cuenta["auxiliar"] = df_cuenta["idAuxiliar"]
                    cols_disponibles.insert(3, "auxiliar")
                
                df_cuenta = df_cuenta[cols_disponibles]
                df_cuenta = df_cuenta.rename(columns={
                    "fecha": "Fecha", "numero": "Comprobante", "tipo": "Tipo",
                    "auxiliar": "Auxiliar", "glosa": "Glosa",
                    "debe": "Debe", "haber": "Haber", "saldo": "Saldo"
                })
                
                if "Fecha" in df_cuenta.columns:
                    df_cuenta["Fecha"] = df_cuenta["Fecha"].str[:10]
                
                sheet_name = sanitize_sheet_name(cuenta["idCuenta"], cuenta["cuenta"])
                suffix = 1
                original = sheet_name
                while sheet_name in sheet_names:
                    sheet_name = f"{original[:28]}_{suffix}"
                    suffix += 1
                
                sheet_names.add(sheet_name)
                df_cuenta.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                cuenta_a_hoja[cuenta["idCuenta"]] = sheet_name
                cuentas_con_datos += 1
                print(f"   ✅ {cuenta['idCuenta']}: {len(df_cuenta)} movimientos")
                
                ws_cuenta = writer.sheets[sheet_name]
                ws_cuenta.cell(row=1, column=1).value = "← Volver al Resumen"
                ws_cuenta.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
                ws_cuenta.cell(row=1, column=1).style = "Hyperlink"
                ws_cuenta.cell(row=2, column=1).value = f"ANÁLISIS DE CUENTA: {cuenta['idCuenta']} - {cuenta['cuenta']} | Período: {periodo_texto}"
                ws_cuenta.cell(row=2, column=1).font = Font(bold=True, size=11, color="006400")
        
        print(f"\n   📊 Cuentas con datos: {cuentas_con_datos}")
        
        # 6. Aplicar formato
        print("\n💰 Aplicando formato...")
        
        # Balance Tributario
        ws_balance = writer.sheets["Balance Tributario"]
        for col in range(1, 12):
            cell = ws_balance.cell(row=3, column=col)
            cell.font = ESTILOS["font_header"]
            cell.fill = ESTILOS["fill_header"]
            cell.alignment = Alignment(horizontal='center')
            cell.border = ESTILOS["thin_border"]
        
        for row in range(4, len(balance_filtrado) + 4):
            for col in range(1, 12):
                cell = ws_balance.cell(row=row, column=col)
                cell.border = ESTILOS["thin_border"]
                if 3 <= col <= 10:
                    cell.number_format = ESTILOS["formato_miles"]
        
        ws_balance.column_dimensions['A'].width = 12
        ws_balance.column_dimensions['B'].width = 35
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_balance.column_dimensions[col].width = 14
        ws_balance.column_dimensions['K'].width = 10
        
        # Hojas de cuentas
        for codigo, sheet_name in cuenta_a_hoja.items():
            ws = writer.sheets[sheet_name]
            num_cols = ws.max_column
            
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = ESTILOS["font_header"]
                cell.fill = ESTILOS["fill_header"]
                cell.alignment = Alignment(horizontal='center')
                cell.border = ESTILOS["thin_border"]
            
            for row in range(4, ws.max_row + 1):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = ESTILOS["thin_border"]
                    if col >= num_cols - 2:
                        cell.number_format = ESTILOS["formato_miles"]
            
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            if num_cols == 8:
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 35
                ws.column_dimensions['F'].width = 14
                ws.column_dimensions['G'].width = 14
                ws.column_dimensions['H'].width = 14
            else:
                ws.column_dimensions['D'].width = 35
                ws.column_dimensions['E'].width = 14
                ws.column_dimensions['F'].width = 14
                ws.column_dimensions['G'].width = 14
        
        # 7. Hipervínculos en Balance Tributario
        print("🔗 Agregando hipervínculos...")
        for row_idx, cuenta in enumerate(balance_filtrado, start=4):
            codigo = cuenta["idCuenta"]
            if codigo in cuenta_a_hoja:
                cell = ws_balance.cell(row=row_idx, column=11)
                cell.value = "→ Ver"
                cell.hyperlink = f"#'{cuenta_a_hoja[codigo]}'!A1"
                cell.style = "Hyperlink"
    
    print(f"\n💾 Guardado: {filename}")
    print("\n" + "═" * 60)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python balance_excel_v2.py <EMPRESA>")
        print("Ejemplo: python balance_excel_v2.py FIDI")
        print("         python balance_excel_v2.py CISI")
        print(f"\nConfiguraciones disponibles en: {CONFIG_EXCEL}")
        sys.exit(1)
    
    empresa_key = sys.argv[1].upper()
    main(empresa_key)

