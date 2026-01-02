#!/usr/bin/env python3
"""
ART-002: Reporte de Pendientes - Implementación Skualo
=======================================================

Genera reporte Excel con pendientes operativos según especificación ART-002.

Uso:
    python pendientes_excel.py FIDI
    python pendientes_excel.py CISI
    python pendientes_excel.py  # Todas las empresas activas
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Agregar path para importar pendientes
sys.path.insert(0, str(Path(__file__).parent.parent))
from pendientes import obtener_pendientes_empresa, _get_tenants

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN - PALETA ART-002 (igual que Balance)
# ═══════════════════════════════════════════════════════════════════════════════

COLORS = {
    "header_bg": "006400",      # Verde oscuro
    "header_text": "FFFFFF",    # Blanco
    "section_bg": "E8F5E9",     # Verde muy claro
    "subtotal_bg": "C8E6C9",    # Verde claro
    "alert_bg": "FFCDD2",       # Rojo claro
    "warning_bg": "FFF9C4",     # Amarillo claro
    "border": "CCCCCC",         # Gris claro
}

OUTPUT_DIR = Path(__file__).parent / "generados"
OUTPUT_DIR.mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════════════════════

def get_styles():
    """Retorna diccionario de estilos reutilizables."""
    thin_border = Border(
        left=Side(style='thin', color=COLORS["border"]),
        right=Side(style='thin', color=COLORS["border"]),
        top=Side(style='thin', color=COLORS["border"]),
        bottom=Side(style='thin', color=COLORS["border"])
    )
    
    return {
        "title": Font(bold=True, size=14, color=COLORS["header_bg"]),
        "subtitle": Font(bold=True, size=12, color=COLORS["header_bg"]),
        "header": Font(bold=True, size=10, color=COLORS["header_text"]),
        "header_fill": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
        "section": Font(bold=True, size=11),
        "section_fill": PatternFill(start_color=COLORS["section_bg"], end_color=COLORS["section_bg"], fill_type="solid"),
        "subtotal": Font(bold=True, size=10),
        "subtotal_fill": PatternFill(start_color=COLORS["subtotal_bg"], end_color=COLORS["subtotal_bg"], fill_type="solid"),
        "alert_fill": PatternFill(start_color=COLORS["alert_bg"], end_color=COLORS["alert_bg"], fill_type="solid"),
        "warning_fill": PatternFill(start_color=COLORS["warning_bg"], end_color=COLORS["warning_bg"], fill_type="solid"),
        "hyperlink": Font(color="0000FF", underline="single"),
        "border": thin_border,
        "number_format": '#,##0',
    }


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: RESUMEN PENDIENTES
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_resumen(writer, data: dict, empresa: str, fecha: str):
    """Crea la hoja Resumen Pendientes."""
    
    styles = get_styles()
    
    # Calcular métricas
    sii = data.get("pendientes_sii", {})
    contab = data.get("pendientes_contabilizar", {})
    concil = data.get("pendientes_conciliar", {})
    
    sii_docs = sii.get("documentos", [])
    contab_docs = contab.get("documentos", [])
    concil_movs = concil.get("movimientos", [])
    
    # Calcular días promedio y más antiguo
    def calcular_stats(items, campo_dias):
        if not items:
            return 0, None, 0
        dias_list = [i.get(campo_dias, 0) or 0 for i in items]
        promedio = sum(dias_list) / len(dias_list) if dias_list else 0
        max_dias = max(dias_list) if dias_list else 0
        # Buscar el más antiguo
        mas_antiguo = None
        for i in items:
            if i.get(campo_dias, 0) == max_dias:
                mas_antiguo = i.get("fecha")
                break
        return promedio, mas_antiguo, max_dias
    
    sii_prom, sii_antiguo, sii_max = calcular_stats(sii_docs, "dias_desde_recepcion")
    
    # Para contabilizar, calcular días desde fecha
    hoy = datetime.now().date()
    for doc in contab_docs:
        fecha_doc = doc.get("fecha")
        if fecha_doc:
            try:
                f = datetime.strptime(fecha_doc[:10], "%Y-%m-%d").date()
                doc["dias_pendiente"] = (hoy - f).days
            except:
                doc["dias_pendiente"] = 0
        else:
            doc["dias_pendiente"] = 0
    
    contab_prom, contab_antiguo, contab_max = calcular_stats(contab_docs, "dias_pendiente")
    
    # Para conciliar
    for mov in concil_movs:
        fecha_mov = mov.get("fecha")
        if fecha_mov:
            try:
                f = datetime.strptime(fecha_mov[:10], "%Y-%m-%d").date()
                mov["dias_pendiente"] = (hoy - f).days
            except:
                mov["dias_pendiente"] = 0
        else:
            mov["dias_pendiente"] = 0
    
    concil_prom, concil_antiguo, concil_max = calcular_stats(concil_movs, "dias_pendiente")
    
    # Construir datos
    rows = [
        ["← Navegación", "", "", "", "", ""],
        [f"RESUMEN DE PENDIENTES - {empresa}", "", "", "", "", ""],
        [f"Fecha consulta: {fecha}", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["Categoría", "Cantidad", "Monto Total", "Días Promedio", "Más Antiguo", "Ver Detalle"],
        [
            "Pendientes SII",
            sii.get("cantidad", 0),
            sii.get("total", 0),
            round(sii_prom, 1),
            f"{sii_antiguo} ({sii_max} días)" if sii_antiguo else "-",
            "→ Ver"
        ],
        [
            "Pendientes Contabilizar",
            contab.get("cantidad", 0),
            sum(d.get("monto", 0) or 0 for d in contab_docs),
            round(contab_prom, 1),
            f"{contab_antiguo} ({contab_max} días)" if contab_antiguo else "-",
            "→ Ver"
        ],
        [
            "Pendientes Conciliar",
            concil.get("cantidad", 0),
            concil.get("total_cargos", 0) + concil.get("total_abonos", 0),
            round(concil_prom, 1),
            f"{concil_antiguo} ({concil_max} días)" if concil_antiguo else "-",
            "→ Ver"
        ],
        ["", "", "", "", "", ""],
        ["", "Totales:", "", "", "", ""],
        ["", "Total Cargos (Conciliar):", concil.get("total_cargos", 0), "", "", ""],
        ["", "Total Abonos (Conciliar):", concil.get("total_abonos", 0), "", "", ""],
    ]
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Resumen Pendientes", index=False, header=False)
    
    ws = writer.sheets["Resumen Pendientes"]
    
    # Aplicar estilos
    # Título
    ws.cell(row=2, column=1).font = styles["title"]
    ws.cell(row=3, column=1).font = Font(italic=True, size=10)
    
    # Header de tabla
    for col in range(1, 7):
        cell = ws.cell(row=5, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    # Filas de datos
    for row in range(6, 9):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = styles["border"]
            if col in [2, 3, 4]:
                cell.number_format = styles["number_format"]
                cell.alignment = Alignment(horizontal='right')
    
    # Hipervínculos
    ws.cell(row=6, column=6).hyperlink = "#'Pendientes SII'!A1"
    ws.cell(row=6, column=6).font = styles["hyperlink"]
    ws.cell(row=7, column=6).hyperlink = "#'Pendientes Contabilizar'!A1"
    ws.cell(row=7, column=6).font = styles["hyperlink"]
    ws.cell(row=8, column=6).hyperlink = "#'Pendientes Conciliar'!A1"
    ws.cell(row=8, column=6).font = styles["hyperlink"]
    
    # Totales
    ws.cell(row=10, column=2).font = Font(bold=True)
    ws.cell(row=11, column=3).number_format = styles["number_format"]
    ws.cell(row=12, column=3).number_format = styles["number_format"]
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES SII
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_sii(writer, data: dict):
    """Crea la hoja Pendientes SII."""
    
    styles = get_styles()
    sii = data.get("pendientes_sii", {})
    docs = sii.get("documentos", [])
    
    # Header
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES SII - Documentos por Aceptar/Rechazar"],
        [""],
        ["Fecha", "Tipo", "Folio", "Proveedor RUT", "Proveedor", "Monto", "Días", "Estado"],
    ]
    
    for doc in docs:
        dias = doc.get("dias_desde_recepcion", 0) or 0
        estado = "Accionable" if dias <= 8 else "Tácito"
        rows.append([
            doc.get("fecha", ""),
            doc.get("tipo", ""),
            doc.get("folio", ""),
            doc.get("proveedor_rut", ""),
            doc.get("proveedor_nombre", ""),
            doc.get("monto", 0),
            dias,
            estado
        ])
    
    if not docs:
        rows.append(["", "Sin pendientes", "", "", "", "", "", ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes SII", index=False, header=False)
    
    ws = writer.sheets["Pendientes SII"]
    
    # Hipervínculo de retorno
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    
    # Título
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    # Header
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    # Datos con formato condicional
    for row_idx, doc in enumerate(docs, start=5):
        dias = doc.get("dias_desde_recepcion", 0) or 0
        
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            # Formato condicional
            if dias >= 6:
                cell.fill = styles["alert_fill"]
            elif dias >= 4:
                cell.fill = styles["warning_fill"]
            
            # Formato numérico
            if col == 6:
                cell.number_format = styles["number_format"]
    
    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES CONTABILIZAR
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_contabilizar(writer, data: dict):
    """Crea la hoja Pendientes Contabilizar."""
    
    styles = get_styles()
    contab = data.get("pendientes_contabilizar", {})
    docs = contab.get("documentos", [])
    
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES CONTABILIZAR - Documentos Aceptados sin Contabilizar"],
        [""],
        ["Fecha", "Tipo", "Folio", "Proveedor RUT", "Proveedor", "Monto", "Días Pendiente"],
    ]
    
    for doc in docs:
        rows.append([
            doc.get("fecha", ""),
            doc.get("tipo", ""),
            doc.get("folio", ""),
            doc.get("proveedor_rut", ""),
            doc.get("proveedor_nombre", ""),
            doc.get("monto", 0),
            doc.get("dias_pendiente", 0)
        ])
    
    if not docs:
        rows.append(["", "Sin pendientes", "", "", "", "", ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes Contabilizar", index=False, header=False)
    
    ws = writer.sheets["Pendientes Contabilizar"]
    
    # Hipervínculo
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    
    # Título
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    # Header
    for col in range(1, 8):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    # Datos
    for row_idx, doc in enumerate(docs, start=5):
        dias = doc.get("dias_pendiente", 0)
        
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            if dias >= 30:
                cell.fill = styles["alert_fill"]
            elif dias >= 15:
                cell.fill = styles["warning_fill"]
            
            if col == 6:
                cell.number_format = styles["number_format"]
    
    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PENDIENTES CONCILIAR
# ═══════════════════════════════════════════════════════════════════════════════

def crear_hoja_conciliar(writer, data: dict):
    """Crea la hoja Pendientes Conciliar."""
    
    styles = get_styles()
    concil = data.get("pendientes_conciliar", {})
    movs = concil.get("movimientos", [])
    
    rows = [
        ["← Volver al Resumen"],
        ["PENDIENTES CONCILIAR - Movimientos Bancarios sin Conciliar"],
        [""],
        ["Fecha", "Banco", "Código", "Descripción", "Número Doc", "Cargo", "Abono", "Días"],
    ]
    
    for mov in movs:
        monto = mov.get("monto", 0) or 0
        cargo = abs(monto) if monto < 0 else 0
        abono = monto if monto > 0 else 0
        
        rows.append([
            mov.get("fecha", "")[:10] if mov.get("fecha") else "",
            mov.get("banco", ""),
            mov.get("banco_codigo", ""),
            mov.get("descripcion", ""),
            mov.get("numero_doc", ""),
            cargo,
            abono,
            mov.get("dias_pendiente", 0)
        ])
    
    if not movs:
        rows.append(["", "Sin pendientes", "", "", "", "", "", ""])
    
    # Subtotales
    total_cargos = concil.get("total_cargos", 0)
    total_abonos = concil.get("total_abonos", 0)
    rows.append(["", "", "", "", "", "", "", ""])
    rows.append(["", "", "", "", "TOTALES:", total_cargos, total_abonos, ""])
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Pendientes Conciliar", index=False, header=False)
    
    ws = writer.sheets["Pendientes Conciliar"]
    
    # Hipervínculo
    ws.cell(row=1, column=1).hyperlink = "#'Resumen Pendientes'!A1"
    ws.cell(row=1, column=1).font = styles["hyperlink"]
    
    # Título
    ws.cell(row=2, column=1).font = styles["subtitle"]
    
    # Header
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.font = styles["header"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal='center')
        cell.border = styles["border"]
    
    # Datos
    for row_idx, mov in enumerate(movs, start=5):
        dias = mov.get("dias_pendiente", 0)
        
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = styles["border"]
            
            if dias >= 30:
                cell.fill = styles["alert_fill"]
            elif dias >= 15:
                cell.fill = styles["warning_fill"]
            
            if col in [6, 7]:
                cell.number_format = styles["number_format"]
    
    # Fila de totales
    total_row = 5 + len(movs) + 1
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = styles["number_format"]
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = styles["number_format"]
    
    # Anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 8


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def generar_reporte(empresa_id: str) -> str:
    """
    Genera el reporte Excel de pendientes para una empresa.
    
    Args:
        empresa_id: Alias (FIDI) o RUT (77285542-7)
    
    Returns:
        Path del archivo generado
    """
    print(f"\n📊 Obteniendo pendientes de {empresa_id}...")
    data = obtener_pendientes_empresa(empresa_id)
    
    empresa = data.get("empresa", empresa_id)
    fecha = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    filename = OUTPUT_DIR / f"Pendientes_{empresa_id}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 1. Resumen
        print("   📋 Generando Resumen...")
        crear_hoja_resumen(writer, data, empresa, fecha)
        
        # 2. Pendientes SII
        print("   📄 Generando Pendientes SII...")
        crear_hoja_sii(writer, data)
        
        # 3. Pendientes Contabilizar
        print("   📝 Generando Pendientes Contabilizar...")
        crear_hoja_contabilizar(writer, data)
        
        # 4. Pendientes Conciliar
        print("   🏦 Generando Pendientes Conciliar...")
        crear_hoja_conciliar(writer, data)
    
    print(f"\n✅ Guardado: {filename}")
    return str(filename)


def main():
    """CLI principal."""
    print("═" * 60)
    print("   ART-002: REPORTE DE PENDIENTES (Skualo)")
    print("═" * 60)
    
    empresa_id = sys.argv[1] if len(sys.argv) > 1 else None
    
    if empresa_id:
        generar_reporte(empresa_id)
    else:
        # Todas las empresas activas
        tenants = _get_tenants()
        for key, config in tenants.items():
            if config.get("activo", True):
                generar_reporte(key)
    
    print("\n" + "═" * 60)


if __name__ == "__main__":
    main()
