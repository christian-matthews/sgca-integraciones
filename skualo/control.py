"""
Skualo Control - Clase principal para control y reportes.

Esta clase encapsula todas las funciones de control y reportes
para facilitar la integración con bots y APIs.
"""

import os
import json
import requests
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, List
from dotenv import load_dotenv

from .config import cargar_config, guardar_config, config_existe

# Cargar variables de entorno
load_dotenv()


class SkualoControl:
    """
    Clase principal para interactuar con Skualo ERP.
    
    Ejemplo:
        ctrl = SkualoControl()
        ctrl.setup_empresa('77285542-7')
        resultado = ctrl.reporte_completo('77285542-7')
    """
    
    BASE_URL = 'https://api.skualo.cl'
    DIAS_ACEPTACION_TACITA = 8
    
    # Mapeo de tipos DTE del SII a tipos internos Skualo
    TIPO_DTE_A_INTERNO = {
        33: 'FACE',   # Factura Electrónica
        34: 'FXCE',   # Factura No Afecta o Exenta Electrónica
        61: 'NCCE',   # Nota de Crédito Electrónica
        56: 'NDCE',   # Nota de Débito Electrónica
        52: 'GDES',   # Guía de Despacho Electrónica
        110: 'FEXP',  # Factura de Exportación Electrónica
    }
    
    # Palabras clave para detectar cuentas bancarias
    PALABRAS_BANCO = [
        'banco', 'santander', 'chile', 'estado', 'bci', 'scotiabank', 
        'itau', 'itaú', 'security', 'bice', 'falabella', 'ripley',
        'consorcio', 'internacional', 'corpbanca', 'tapp', 'tenpo',
        'mercado pago', 'cuenta corriente', 'cta cte', 'cta. cte',
        'coopeuch', 'bancoestado', 'bco.', 'bco '
    ]
    
    def __init__(self, token: str = None):
        """
        Inicializa el controlador.
        
        Args:
            token: Token de API de Skualo. Si no se proporciona,
                   se lee de la variable de entorno SKUALO_API_TOKEN
        """
        self.token = token or os.getenv('SKUALO_API_TOKEN')
        if not self.token:
            raise ValueError("Token no proporcionado. Configure SKUALO_API_TOKEN en .env")
        
        self.output_dir = Path(__file__).parent.parent / 'generados'
        self.output_dir.mkdir(exist_ok=True)
    
    def _headers(self) -> dict:
        """Headers para las peticiones API."""
        return {
            'Authorization': f'Bearer {self.token}',
            'accept': 'application/json'
        }
    
    def _api_get(self, rut: str, endpoint: str, params: dict = None) -> Optional[Dict]:
        """Realiza una llamada GET a la API."""
        url = f'{self.BASE_URL}/{rut}{endpoint}'
        try:
            r = requests.get(url, headers=self._headers(), params=params, timeout=30)
            if r.ok:
                return r.json()
        except Exception as e:
            print(f'Error API: {e}')
        return None
    
    def _api_get_all(self, rut: str, endpoint: str, params: dict = None) -> List:
        """Obtiene todos los registros paginados de un endpoint."""
        all_items = []
        page = 1
        base_params = params or {}
        
        while True:
            paged_params = {**base_params, 'PageSize': 100, 'Page': page}
            data = self._api_get(rut, endpoint, paged_params)
            if not data:
                break
            items = data.get('items', data) if isinstance(data, dict) else data
            if isinstance(items, list):
                all_items.extend(items)
            if not data.get('next'):
                break
            page += 1
        
        return all_items
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SETUP DE EMPRESA
    # ═══════════════════════════════════════════════════════════════════════════
    
    def setup_empresa(self, rut: str, interactivo: bool = False) -> Optional[Dict]:
        """
        Configura una nueva empresa.
        
        Solo necesita el RUT, extrae todo lo demás automáticamente:
        - Nombre de la empresa
        - Cuentas bancarias
        - Cuenta de clientes y proveedores
        
        Args:
            rut: RUT de la empresa (ej: '77285542-7')
            interactivo: Si True, pide confirmación al usuario
        
        Returns:
            dict con la configuración guardada o None si falla
        """
        config = {
            'rut': rut,
            'configurado_el': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'cuentas_bancarias': [],
            'cuenta_clientes': None,
            'cuenta_proveedores': None,
        }
        
        # 1. Obtener información de la empresa
        empresa = self._api_get(rut, '/empresa')
        if not empresa:
            return None
        
        config['nombre'] = empresa.get('nombre', rut)
        config['razon_social'] = empresa.get('razonSocial', '')
        config['giro'] = empresa.get('giro', '')
        
        # 2. Detectar cuentas bancarias del balance
        periodo = datetime.now().strftime('%Y%m')
        balance = self._api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
        
        if not balance:
            # Intentar con el mes anterior
            fecha_ant = datetime.now().replace(day=1) - timedelta(days=1)
            periodo = fecha_ant.strftime('%Y%m')
            balance = self._api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
        
        if balance:
            # Detectar cuentas bancarias
            for cuenta in balance:
                codigo = cuenta.get('idCuenta', '')
                nombre = cuenta.get('cuenta', '').lower()
                
                es_banco = False
                if codigo.startswith('1102') or codigo.startswith('1103'):
                    es_banco = True
                elif any(p in nombre for p in self.PALABRAS_BANCO):
                    if codigo.startswith('1'):
                        es_banco = True
                
                if es_banco:
                    config['cuentas_bancarias'].append({
                        'codigo': codigo,
                        'nombre': cuenta.get('cuenta', ''),
                        'activa': True
                    })
            
            # Detectar cuenta de clientes y proveedores
            for cuenta in balance:
                codigo = cuenta.get('idCuenta', '')
                nombre = cuenta.get('cuenta', '').lower()
                
                if not config['cuenta_clientes']:
                    if codigo.startswith('1107') or codigo.startswith('1108'):
                        config['cuenta_clientes'] = codigo
                    elif 'cliente' in nombre or 'por cobrar' in nombre:
                        config['cuenta_clientes'] = codigo
                
                if not config['cuenta_proveedores']:
                    if codigo.startswith('2110') or codigo.startswith('2111'):
                        config['cuenta_proveedores'] = codigo
                    elif 'proveedor' in nombre or 'por pagar' in nombre:
                        config['cuenta_proveedores'] = codigo
        
        # 3. Verificar endpoints disponibles
        config['endpoints_disponibles'] = {
            '/sii/dte/recibidos': self._api_get(rut, '/sii/dte/recibidos', {'PageSize': 1}) is not None,
            '/sii/dte': self._api_get(rut, '/sii/dte', {'PageSize': 1}) is not None,
        }
        
        if config['cuentas_bancarias']:
            codigo_test = config['cuentas_bancarias'][0]['codigo']
            config['endpoints_disponibles']['/bancos'] = self._api_get(rut, f'/bancos/{codigo_test}', {'PageSize': 1}) is not None
        
        # 4. Guardar configuración
        guardar_config(rut, config)
        
        return config
    
    # ═══════════════════════════════════════════════════════════════════════════
    # FUNCIÓN 1: MOVIMIENTOS BANCARIOS PENDIENTES DE CONCILIAR
    # ═══════════════════════════════════════════════════════════════════════════
    
    def movimientos_bancarios_pendientes(self, rut: str) -> Optional[Dict]:
        """
        Obtiene movimientos bancarios pendientes de conciliar.
        
        Args:
            rut: RUT de la empresa
        
        Returns:
            dict con:
            - empresa: Nombre de la empresa
            - rut: RUT
            - fecha: Fecha del reporte
            - cuentas: Lista de cuentas con movimientos sin conciliar
            - total_sin_conciliar: Total de movimientos sin conciliar
        """
        config = cargar_config(rut)
        if not config:
            return None
        
        resultado = {
            'empresa': config['nombre'],
            'rut': rut,
            'fecha': datetime.now().isoformat(),
            'cuentas': [],
            'total_sin_conciliar': 0
        }
        
        if not config.get('cuentas_bancarias'):
            return resultado
        
        for cuenta in config['cuentas_bancarias']:
            if not cuenta.get('activa', True):
                continue
            
            codigo = cuenta['codigo']
            nombre = cuenta['nombre']
            
            movimientos = self._api_get_all(rut, f'/bancos/{codigo}')
            sin_conciliar = [m for m in movimientos if not m.get('conciliado', True)]
            
            cuenta_resultado = {
                'codigo': codigo,
                'nombre': nombre,
                'total_movimientos': len(movimientos),
                'sin_conciliar': len(sin_conciliar),
                'movimientos': sin_conciliar
            }
            resultado['cuentas'].append(cuenta_resultado)
            resultado['total_sin_conciliar'] += len(sin_conciliar)
        
        return resultado
    
    # ═══════════════════════════════════════════════════════════════════════════
    # FUNCIÓN 2: DOCUMENTOS PENDIENTES DE APROBAR EN SII
    # ═══════════════════════════════════════════════════════════════════════════
    
    def documentos_por_aprobar_sii(self, rut: str) -> Optional[Dict]:
        """
        Obtiene documentos pendientes de aprobar en el SII.
        
        Regla: DTEs recibidos con menos de 8 días sin respuesta.
        Después de 8 días se considera aceptación tácita.
        
        Args:
            rut: RUT de la empresa
        
        Returns:
            dict con:
            - empresa: Nombre de la empresa
            - pendientes: Lista de documentos pendientes
            - total_pendientes: Cantidad de documentos
            - monto_total: Monto total de los documentos
        """
        config = cargar_config(rut)
        if not config:
            return None
        
        resultado = {
            'empresa': config['nombre'],
            'rut': rut,
            'fecha': datetime.now().isoformat(),
            'pendientes': [],
            'total_pendientes': 0,
            'monto_total': 0
        }
        
        dtes = self._api_get_all(rut, '/sii/dte/recibidos')
        hoy = datetime.now()
        
        for dte in dtes:
            if dte.get('fechaRespuesta'):
                continue
            
            fecha_recep_str = dte.get('creadoEl', '')
            if not fecha_recep_str:
                continue
            
            try:
                fecha_recep = datetime.fromisoformat(fecha_recep_str.split('.')[0])
                dias = (hoy - fecha_recep).days
            except:
                continue
            
            if dias > self.DIAS_ACEPTACION_TACITA:
                continue
            
            doc = {
                'rut_emisor': dte.get('rutEmisor', ''),
                'emisor': dte.get('emisor', ''),
                'tipo_documento': dte.get('tipoDocumento', ''),
                'tipo_dte': dte.get('idTipoDocumento'),
                'folio': dte.get('folio'),
                'fecha_emision': str(dte.get('fechaEmision', ''))[:10],
                'monto': dte.get('montoTotal', 0),
                'dias_restantes': self.DIAS_ACEPTACION_TACITA - dias
            }
            resultado['pendientes'].append(doc)
            resultado['monto_total'] += doc['monto']
        
        resultado['total_pendientes'] = len(resultado['pendientes'])
        return resultado
    
    # ═══════════════════════════════════════════════════════════════════════════
    # FUNCIÓN 3: DOCUMENTOS PENDIENTES DE CONTABILIZAR
    # ═══════════════════════════════════════════════════════════════════════════
    
    def documentos_por_contabilizar(self, rut: str) -> Optional[Dict]:
        """
        Obtiene documentos pendientes de contabilizar.
        
        Son DTEs aceptados (> 8 días o con respuesta) que NO existen
        en el módulo de documentos (no han sido ingresados al sistema).
        
        Args:
            rut: RUT de la empresa
        
        Returns:
            dict con:
            - empresa: Nombre de la empresa
            - pendientes: Lista de documentos pendientes
            - ya_contabilizados: Cantidad de documentos ya contabilizados
            - total_pendientes: Cantidad de documentos pendientes
            - monto_total: Monto total de los documentos pendientes
        """
        config = cargar_config(rut)
        if not config:
            return None
        
        resultado = {
            'empresa': config['nombre'],
            'rut': rut,
            'fecha': datetime.now().isoformat(),
            'pendientes': [],
            'ya_contabilizados': 0,
            'total_pendientes': 0,
            'monto_total': 0
        }
        
        dtes = self._api_get_all(rut, '/sii/dte/recibidos')
        hoy = datetime.now()
        
        for dte in dtes:
            fecha_respuesta = dte.get('fechaRespuesta')
            
            if not fecha_respuesta:
                fecha_recep_str = dte.get('creadoEl', '')
                try:
                    fecha_recep = datetime.fromisoformat(fecha_recep_str.split('.')[0])
                    dias = (hoy - fecha_recep).days
                    if dias <= self.DIAS_ACEPTACION_TACITA:
                        continue
                except:
                    continue
            
            tipo_dte = dte.get('idTipoDocumento')
            folio = dte.get('folio')
            tipo_interno = self.TIPO_DTE_A_INTERNO.get(tipo_dte, 'FACE')
            
            doc = self._api_get(rut, f'/documentos/{tipo_interno}/{folio}')
            
            if doc:
                resultado['ya_contabilizados'] += 1
            else:
                pendiente = {
                    'rut_emisor': dte.get('rutEmisor', ''),
                    'emisor': dte.get('emisor', ''),
                    'tipo_documento': dte.get('tipoDocumento', ''),
                    'tipo_dte': tipo_dte,
                    'tipo_interno': tipo_interno,
                    'folio': folio,
                    'fecha_emision': str(dte.get('fechaEmision', ''))[:10],
                    'monto': dte.get('montoTotal', 0),
                }
                resultado['pendientes'].append(pendiente)
                resultado['monto_total'] += pendiente['monto']
        
        resultado['total_pendientes'] = len(resultado['pendientes'])
        return resultado
    
    # ═══════════════════════════════════════════════════════════════════════════
    # REPORTE COMPLETO
    # ═══════════════════════════════════════════════════════════════════════════
    
    def reporte_completo(self, rut: str) -> Optional[Dict]:
        """
        Genera un reporte completo con los 3 controles.
        
        Args:
            rut: RUT de la empresa
        
        Returns:
            dict con:
            - empresa: Nombre de la empresa
            - fecha: Fecha del reporte
            - bancos: Resultado de movimientos_bancarios_pendientes
            - aprobar: Resultado de documentos_por_aprobar_sii
            - contabilizar: Resultado de documentos_por_contabilizar
            - resumen: Resumen ejecutivo con totales
        """
        config = cargar_config(rut)
        if not config:
            return None
        
        bancos = self.movimientos_bancarios_pendientes(rut)
        aprobar = self.documentos_por_aprobar_sii(rut)
        contabilizar = self.documentos_por_contabilizar(rut)
        
        return {
            'empresa': config['nombre'],
            'rut': rut,
            'fecha': datetime.now().isoformat(),
            'bancos': bancos,
            'aprobar': aprobar,
            'contabilizar': contabilizar,
            'resumen': {
                'movimientos_sin_conciliar': bancos['total_sin_conciliar'] if bancos else 0,
                'documentos_por_aprobar': aprobar['total_pendientes'] if aprobar else 0,
                'documentos_por_contabilizar': contabilizar['total_pendientes'] if contabilizar else 0,
                'documentos_contabilizados': contabilizar['ya_contabilizados'] if contabilizar else 0,
            }
        }
    
    # ═══════════════════════════════════════════════════════════════════════════
    # GENERAR BALANCE EXCEL
    # ═══════════════════════════════════════════════════════════════════════════
    
    def generar_balance_excel(self, rut: str, periodo: str = None) -> Optional[str]:
        """
        Genera un Excel con el Balance Tributario y Análisis por Cuenta.
        
        Args:
            rut: RUT de la empresa
            periodo: Período en formato YYYYMM (ej: '202511'). 
                    Si no se proporciona, usa el mes actual.
        
        Returns:
            Ruta del archivo Excel generado o None si falla
        """
        try:
            import pandas as pd
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError("Requiere pandas y openpyxl: pip install pandas openpyxl")
        
        config = cargar_config(rut)
        if not config:
            return None
        
        if not periodo:
            periodo = datetime.now().strftime('%Y%m')
        
        # Calcular fecha de corte
        año = int(periodo[:4])
        mes = int(periodo[4:6])
        import calendar
        ultimo_dia = calendar.monthrange(año, mes)[1]
        fecha_corte = f'{año}-{mes:02d}-{ultimo_dia:02d}'
        
        # Obtener balance
        balance = self._api_get(rut, f'/contabilidad/reportes/balancetributario/{periodo}')
        if not balance:
            return None
        
        # Crear Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_empresa = config['nombre'].replace(' ', '_')[:20]
        filename = self.output_dir / f'Balance_{nombre_empresa}_{periodo}_{timestamp}.xlsx'
        
        # Estilos
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def sanitize_sheet_name(codigo, nombre):
            name = f"{codigo} {nombre}"
            for char in ['\\', '/', '*', '?', '[', ']', ':']:
                name = name.replace(char, '')
            return name[:31]
        
        # DataFrame del balance
        df_balance = pd.DataFrame(balance)
        columnas = {
            'idCuenta': 'Código',
            'cuenta': 'Cuenta',
            'tipo': 'Tipo',
            'debe': 'Debe',
            'haber': 'Haber',
            'saldo': 'Saldo'
        }
        df_balance = df_balance.rename(columns=columnas)
        cols_mostrar = [c for c in ['Código', 'Cuenta', 'Tipo', 'Debe', 'Haber', 'Saldo'] if c in df_balance.columns]
        df_balance = df_balance[cols_mostrar]
        df_balance['Ver Detalle'] = ''
        
        # Cuentas a procesar
        cuentas_con_mov = [c for c in balance if c.get('debe', 0) != 0 or c.get('haber', 0) != 0 or c.get('saldo', 0) != 0]
        cuentas_a_procesar = cuentas_con_mov if cuentas_con_mov else balance
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_balance.to_excel(writer, sheet_name='Balance Tributario', index=False)
            ws_balance = writer.sheets['Balance Tributario']
            
            # Formato encabezados
            for col in range(1, len(df_balance.columns) + 1):
                cell = ws_balance.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Formato datos
            for row in range(2, len(df_balance) + 2):
                for col in range(1, len(df_balance.columns) + 1):
                    cell = ws_balance.cell(row=row, column=col)
                    cell.border = border
                    if col >= 4 and col <= 6:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')
            
            # Ancho columnas
            ws_balance.column_dimensions['A'].width = 12
            ws_balance.column_dimensions['B'].width = 40
            ws_balance.column_dimensions['C'].width = 12
            ws_balance.column_dimensions['D'].width = 15
            ws_balance.column_dimensions['E'].width = 15
            ws_balance.column_dimensions['F'].width = 15
            ws_balance.column_dimensions['G'].width = 12
            
            # Procesar cuentas
            cuentas_procesadas = 0
            for cuenta in cuentas_a_procesar:
                codigo = cuenta.get('idCuenta', '')
                nombre = cuenta.get('cuenta', '')
                
                analisis = self._api_get(rut, f'/contabilidad/reportes/analisisporcuenta/{codigo}?fechaCorte={fecha_corte}&soloPendientes=false')
                if not analisis:
                    continue
                
                movimientos = [m for m in analisis if m.get('saldo', 0) != 0 or m.get('valor', 0) != 0]
                if not movimientos:
                    continue
                
                sheet_name = sanitize_sheet_name(codigo, nombre)
                df_cuenta = pd.DataFrame(movimientos)
                
                col_rename = {
                    'comprobante': 'Comp',
                    'idTipoDoc': 'Tipo',
                    'numDoc': 'N° Doc',
                    'auxiliar': 'Auxiliar',
                    'emision': 'Emisión',
                    'vencimiento': 'Vencimiento',
                    'valor': 'Valor',
                    'saldo': 'Saldo',
                    'glosa': 'Glosa'
                }
                df_cuenta = df_cuenta.rename(columns=col_rename)
                cols_cuenta = [c for c in ['Comp', 'Tipo', 'N° Doc', 'Auxiliar', 'Emisión', 'Vencimiento', 'Valor', 'Saldo', 'Glosa'] 
                              if c in df_cuenta.columns]
                df_cuenta = df_cuenta[cols_cuenta]
                
                for col in ['Emisión', 'Vencimiento']:
                    if col in df_cuenta.columns:
                        df_cuenta[col] = df_cuenta[col].apply(lambda x: str(x)[:10] if x else '')
                
                df_cuenta.to_excel(writer, sheet_name=sheet_name, index=False)
                
                ws_cuenta = writer.sheets[sheet_name]
                for col in range(1, len(df_cuenta.columns) + 1):
                    cell = ws_cuenta.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                
                # Hipervínculo
                for row in range(2, len(df_balance) + 2):
                    if ws_balance.cell(row=row, column=1).value == codigo:
                        cell_link = ws_balance.cell(row=row, column=7)
                        cell_link.value = 'Ver Detalle'
                        cell_link.hyperlink = f"#'{sheet_name}'!A1"
                        cell_link.font = Font(color='0563C1', underline='single')
                        break
                
                cuentas_procesadas += 1
        
        return str(filename)
    
    # ═══════════════════════════════════════════════════════════════════════════
    # MÉTODOS DE FORMATO PARA BOT
    # ═══════════════════════════════════════════════════════════════════════════
    
    def formato_reporte_telegram(self, rut: str) -> str:
        """
        Genera un reporte formateado para Telegram.
        
        Args:
            rut: RUT de la empresa
        
        Returns:
            String con el reporte formateado en Markdown
        """
        reporte = self.reporte_completo(rut)
        if not reporte:
            return f"❌ No hay configuración para {rut}"
        
        r = reporte['resumen']
        fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        texto = f"""
📊 *REPORTE DE CONTROL*
_{reporte['empresa']}_
_{fecha}_

🏦 *Movimientos sin conciliar:* {r['movimientos_sin_conciliar']}
📄 *Docs por aprobar SII:* {r['documentos_por_aprobar']}
📋 *Docs por contabilizar:* {r['documentos_por_contabilizar']}
✅ *Docs contabilizados:* {r['documentos_contabilizados']}
"""
        
        # Detalle de pendientes de aprobar
        if reporte['aprobar'] and reporte['aprobar']['pendientes']:
            texto += "\n*📄 Pendientes de aprobar:*\n"
            for doc in reporte['aprobar']['pendientes'][:5]:
                texto += f"  • {doc['emisor'][:20]} - ${doc['monto']:,.0f} ({doc['dias_restantes']}d)\n"
        
        # Detalle de pendientes de contabilizar
        if reporte['contabilizar'] and reporte['contabilizar']['pendientes']:
            texto += "\n*📋 Pendientes de contabilizar:*\n"
            for doc in reporte['contabilizar']['pendientes'][:5]:
                texto += f"  • {doc['emisor'][:20]} - ${doc['monto']:,.0f}\n"
            if len(reporte['contabilizar']['pendientes']) > 5:
                texto += f"  _...y {len(reporte['contabilizar']['pendientes']) - 5} más_\n"
        
        return texto

