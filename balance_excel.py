"""
Balance Tributario + Análisis por Cuenta a Excel
"""

import os
import json
import requests
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, Font, Alignment, PatternFill, Border, Side

# Carpeta para archivos generados
OUTPUT_DIR = "generados"

load_dotenv()

API_BASE = "https://api.skualo.cl"
TOKEN = os.getenv("SKUALO_API_TOKEN")

# Cargar tenants
with open("tenants.json", "r") as f:
    TENANTS = json.load(f)


def api_get(tenant_rut, path):
    """Llamada GET a la API"""
    url = f"{API_BASE}/{tenant_rut}{path}"
    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "accept": "application/json"
    }
    response = requests.get(url, headers=headers)
    if response.ok:
        return response.json()
    return None


def get_balance(tenant_rut, id_periodo):
    """Obtener Balance Tributario"""
    return api_get(tenant_rut, f"/contabilidad/reportes/balancetributario/{id_periodo}")


def get_analisis_cuenta(tenant_rut, id_cuenta, fecha_corte):
    """Obtener Análisis por Cuenta"""
    return api_get(tenant_rut, f"/contabilidad/reportes/analisisporcuenta/{id_cuenta}?fechaCorte={fecha_corte}&soloPendientes=false")


def sanitize_sheet_name(codigo, nombre):
    """Limpiar nombre para hoja Excel (max 31 chars)"""
    name = f"{codigo} {nombre}"
    for char in ['\\', '/', '*', '?', '[', ']', ':']:
        name = name.replace(char, '')
    return name[:31]


# Definición de agrupaciones para Estado de Resultados
AGRUPACIONES_EERR = {
    "Ingresos por Ventas": {
        "cuentas": ["4101001"],
        "descripcion": "Ventas Del Giro - Ingresos principales del negocio"
    },
    "Costo de Ventas": {
        "cuentas": ["5101001"],
        "descripcion": "Costo De Venta - Costo directo de los servicios/productos vendidos"
    },
    "Remuneraciones y Salarios": {
        "cuentas": ["5201001", "5201999"],
        "descripcion": "Remuneraciones + Otras Remuneraciones - Gastos de personal"
    },
    "Honorarios Profesionales": {
        "cuentas": ["5205002"],
        "descripcion": "Honorarios Profesionales - Servicios profesionales externos"
    },
    "Arriendo": {
        "cuentas": ["5304001"],
        "descripcion": "Arriendo Inmueble Del Giro - Gastos de arriendo de oficinas"
    },
    "Marketing y Publicidad": {
        "cuentas": ["5307001", "5307002"],
        "descripcion": "Publicidad Y Promociones + Medios - Gastos de marketing"
    },
    "Gastos Legales y Contables": {
        "cuentas": ["5308001", "5308002", "5308003"],
        "descripcion": "Gastos Notariales + Asesoría Legal + Asesoría Contable"
    },
    "Otros Gastos Operativos": {
        "cuentas": ["5309016", "5310999"],
        "descripcion": "Suscripciones + Gastos Generales"
    },
    "Multas y Sanciones": {
        "cuentas": ["5314005"],
        "descripcion": "Multas Al Fisco - Gastos no operativos por sanciones"
    },
    "Gastos Financieros": {
        "cuentas": ["5403001", "5403003", "5403005"],
        "descripcion": "Intereses Financieros + Comisiones Financieras + Gastos Factoring"
    }
}


def crear_estado_resultados(balance, writer):
    """Crea la hoja de Estado de Resultados"""
    
    # Crear diccionario de cuentas para búsqueda rápida
    cuentas_dict = {c["idCuenta"]: c for c in balance}
    
    # Calcular valores por agrupación
    def get_valor(cuenta_id):
        if cuenta_id in cuentas_dict:
            c = cuentas_dict[cuenta_id]
            # Ganancias son positivas (ingresos), Pérdidas son negativas (gastos)
            return c.get("ganancias", 0) - c.get("perdidas", 0)
        return 0
    
    def suma_grupo(cuentas_list):
        return sum(get_valor(c) for c in cuentas_list)
    
    # Calcular cada línea
    ingresos = suma_grupo(AGRUPACIONES_EERR["Ingresos por Ventas"]["cuentas"])
    costo_ventas = suma_grupo(AGRUPACIONES_EERR["Costo de Ventas"]["cuentas"])
    utilidad_bruta = ingresos + costo_ventas  # costo_ventas es negativo
    
    remuneraciones = suma_grupo(AGRUPACIONES_EERR["Remuneraciones y Salarios"]["cuentas"])
    honorarios = suma_grupo(AGRUPACIONES_EERR["Honorarios Profesionales"]["cuentas"])
    arriendo = suma_grupo(AGRUPACIONES_EERR["Arriendo"]["cuentas"])
    marketing = suma_grupo(AGRUPACIONES_EERR["Marketing y Publicidad"]["cuentas"])
    legales = suma_grupo(AGRUPACIONES_EERR["Gastos Legales y Contables"]["cuentas"])
    otros_gastos = suma_grupo(AGRUPACIONES_EERR["Otros Gastos Operativos"]["cuentas"])
    
    total_gastos_op = remuneraciones + honorarios + arriendo + marketing + legales + otros_gastos
    
    resultado_operacional = utilidad_bruta + total_gastos_op
    
    multas = suma_grupo(AGRUPACIONES_EERR["Multas y Sanciones"]["cuentas"])
    gastos_financieros = suma_grupo(AGRUPACIONES_EERR["Gastos Financieros"]["cuentas"])
    
    resultado_antes_impuestos = resultado_operacional + multas + gastos_financieros
    
    # Construir DataFrame
    data = [
        ("", "", ""),
        ("ESTADO DE RESULTADOS", "", "Monto ($)"),
        ("", "", ""),
        ("Ingresos por Ventas", "", ingresos),
        ("Costo de Ventas", "", costo_ventas),
        ("", "", ""),
        ("UTILIDAD BRUTA", "", utilidad_bruta),
        ("", "", ""),
        ("Gastos Operacionales:", "", ""),
        ("  Remuneraciones y Salarios", "", remuneraciones),
        ("  Honorarios Profesionales", "", honorarios),
        ("  Arriendo", "", arriendo),
        ("  Marketing y Publicidad", "", marketing),
        ("  Gastos Legales y Contables", "", legales),
        ("  Otros Gastos Operativos", "", otros_gastos),
        ("Total Gastos Operacionales", "", total_gastos_op),
        ("", "", ""),
        ("RESULTADO OPERACIONAL (EBIT)", "", resultado_operacional),
        ("", "", ""),
        ("Otros Ingresos / Gastos:", "", ""),
        ("  Multas y Sanciones", "", multas),
        ("  Gastos Financieros", "", gastos_financieros),
        ("", "", ""),
        ("RESULTADO ANTES DE IMPUESTOS", "", resultado_antes_impuestos),
        ("", "", ""),
        ("Impuesto a la Renta (estimado 27%)", "", -resultado_antes_impuestos * 0.27 if resultado_antes_impuestos > 0 else 0),
        ("", "", ""),
        ("RESULTADO NETO DEL EJERCICIO", "", resultado_antes_impuestos * 0.73 if resultado_antes_impuestos > 0 else resultado_antes_impuestos),
    ]
    
    df = pd.DataFrame(data, columns=["Concepto", "", "Valor"])
    df.to_excel(writer, sheet_name="Estado de Resultados", index=False, header=False)
    
    # Aplicar formato
    ws = writer.sheets["Estado de Resultados"]
    formato_miles = '#,##0'
    
    # Estilos
    font_titulo = Font(bold=True, size=14, color="006400")
    font_subtotal = Font(bold=True, size=11)
    font_total = Font(bold=True, size=12, color="FFFFFF")
    fill_total = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    
    # Aplicar formatos
    for row in range(1, len(data) + 1):
        cell_concepto = ws.cell(row=row, column=1)
        cell_valor = ws.cell(row=row, column=3)
        
        concepto = data[row-1][0]
        
        if concepto == "ESTADO DE RESULTADOS":
            cell_concepto.font = font_titulo
        elif concepto in ["UTILIDAD BRUTA", "Total Gastos Operacionales", "RESULTADO OPERACIONAL (EBIT)"]:
            cell_concepto.font = font_subtotal
            cell_valor.font = font_subtotal
        elif concepto in ["RESULTADO ANTES DE IMPUESTOS", "RESULTADO NETO DEL EJERCICIO"]:
            cell_concepto.font = font_total
            cell_concepto.fill = fill_total
            cell_valor.font = font_total
            cell_valor.fill = fill_total
        
        if isinstance(data[row-1][2], (int, float)) and data[row-1][2] != "":
            cell_valor.number_format = formato_miles
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['C'].width = 20
    
    # Retornar valores para KPIs
    return {
        "ingresos": ingresos,
        "costo_ventas": costo_ventas,
        "utilidad_bruta": utilidad_bruta,
        "gastos_operacionales": total_gastos_op,
        "resultado_operacional": resultado_operacional,
        "gastos_financieros": gastos_financieros,
        "resultado_antes_impuestos": resultado_antes_impuestos,
        "resultado_neto": resultado_antes_impuestos * 0.73 if resultado_antes_impuestos > 0 else resultado_antes_impuestos
    }


def crear_kpis(valores, balance, writer):
    """Crea la hoja de KPIs"""
    
    ingresos = valores["ingresos"] if valores["ingresos"] != 0 else 1  # Evitar división por 0
    
    # Calcular activos y pasivos totales del balance
    total_activos = sum(c.get("activos", 0) for c in balance)
    total_pasivos = sum(c.get("pasivos", 0) for c in balance)
    patrimonio = total_activos - total_pasivos
    
    # KPIs
    margen_bruto = (valores["utilidad_bruta"] / valores["ingresos"] * 100) if valores["ingresos"] != 0 else 0
    margen_operacional = (valores["resultado_operacional"] / valores["ingresos"] * 100) if valores["ingresos"] != 0 else 0
    margen_neto = (valores["resultado_neto"] / valores["ingresos"] * 100) if valores["ingresos"] != 0 else 0
    
    # EBITDA (sin depreciación, usamos EBIT como proxy)
    ebitda = valores["resultado_operacional"]
    margen_ebitda = (ebitda / valores["ingresos"] * 100) if valores["ingresos"] != 0 else 0
    
    # ROA y ROE
    roa = (valores["resultado_neto"] / total_activos * 100) if total_activos != 0 else 0
    roe = (valores["resultado_neto"] / patrimonio * 100) if patrimonio != 0 else 0
    
    # Ratio de endeudamiento
    ratio_deuda = (total_pasivos / total_activos * 100) if total_activos != 0 else 0
    
    data = [
        ("", "", ""),
        ("INDICADORES FINANCIEROS (KPIs)", "", ""),
        ("", "", ""),
        ("═" * 50, "", ""),
        ("MÁRGENES DE RENTABILIDAD", "", ""),
        ("═" * 50, "", ""),
        ("", "", ""),
        ("Margen Bruto", f"{margen_bruto:.1f}%", "(Utilidad Bruta / Ingresos) × 100"),
        ("Margen Operacional (EBIT)", f"{margen_operacional:.1f}%", "(Resultado Operacional / Ingresos) × 100"),
        ("Margen EBITDA", f"{margen_ebitda:.1f}%", "(EBITDA / Ingresos) × 100"),
        ("Margen Neto", f"{margen_neto:.1f}%", "(Resultado Neto / Ingresos) × 100"),
        ("", "", ""),
        ("═" * 50, "", ""),
        ("RENTABILIDAD SOBRE INVERSIÓN", "", ""),
        ("═" * 50, "", ""),
        ("", "", ""),
        ("ROA (Return on Assets)", f"{roa:.1f}%", "(Resultado Neto / Total Activos) × 100"),
        ("ROE (Return on Equity)", f"{roe:.1f}%", "(Resultado Neto / Patrimonio) × 100"),
        ("", "", ""),
        ("═" * 50, "", ""),
        ("ESTRUCTURA FINANCIERA", "", ""),
        ("═" * 50, "", ""),
        ("", "", ""),
        ("Ratio de Endeudamiento", f"{ratio_deuda:.1f}%", "(Total Pasivos / Total Activos) × 100"),
        ("", "", ""),
        ("═" * 50, "", ""),
        ("VALORES ABSOLUTOS", "", ""),
        ("═" * 50, "", ""),
        ("", "", ""),
        ("Ingresos Totales", f"${valores['ingresos']:,.0f}", ""),
        ("Utilidad Bruta", f"${valores['utilidad_bruta']:,.0f}", ""),
        ("EBITDA", f"${ebitda:,.0f}", ""),
        ("Resultado Neto", f"${valores['resultado_neto']:,.0f}", ""),
        ("", "", ""),
        ("Total Activos", f"${total_activos:,.0f}", ""),
        ("Total Pasivos", f"${total_pasivos:,.0f}", ""),
        ("Patrimonio", f"${patrimonio:,.0f}", ""),
    ]
    
    df = pd.DataFrame(data, columns=["Indicador", "Valor", "Fórmula"])
    df.to_excel(writer, sheet_name="KPIs", index=False, header=False)
    
    # Aplicar formato
    ws = writer.sheets["KPIs"]
    font_titulo = Font(bold=True, size=14, color="006400")
    font_seccion = Font(bold=True, size=11)
    
    for row in range(1, len(data) + 1):
        concepto = data[row-1][0]
        if concepto == "INDICADORES FINANCIEROS (KPIs)":
            ws.cell(row=row, column=1).font = font_titulo
        elif concepto in ["MÁRGENES DE RENTABILIDAD", "RENTABILIDAD SOBRE INVERSIÓN", 
                          "ESTRUCTURA FINANCIERA", "VALORES ABSOLUTOS"]:
            ws.cell(row=row, column=1).font = font_seccion
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45


def crear_documentacion(writer):
    """Crea la hoja de documentación"""
    
    data = [
        ("← Volver al Resumen", "", ""),  # Fila 1: Navegación
        ("", "", ""),
        ("DOCUMENTACIÓN - AGRUPACIONES Y CÁLCULOS", "", ""),
        ("", "", ""),
        ("═" * 80, "", ""),
        ("1. AGRUPACIONES DE CUENTAS EN ESTADO DE RESULTADOS", "", ""),
        ("═" * 80, "", ""),
        ("", "", ""),
    ]
    
    for nombre, info in AGRUPACIONES_EERR.items():
        cuentas_str = ", ".join(info["cuentas"])
        data.append((nombre, cuentas_str, info["descripcion"]))
    
    data.extend([
        ("", "", ""),
        ("═" * 80, "", ""),
        ("2. CÁLCULOS DEL ESTADO DE RESULTADOS", "", ""),
        ("═" * 80, "", ""),
        ("", "", ""),
        ("Utilidad Bruta", "=", "Ingresos por Ventas - Costo de Ventas"),
        ("Total Gastos Operacionales", "=", "Suma de todos los gastos operativos"),
        ("Resultado Operacional (EBIT)", "=", "Utilidad Bruta - Total Gastos Operacionales"),
        ("Resultado Antes de Impuestos", "=", "EBIT + Otros Ingresos - Otros Gastos"),
        ("Impuesto a la Renta", "=", "27% del Resultado Antes de Impuestos (si es positivo)"),
        ("Resultado Neto", "=", "Resultado Antes de Impuestos - Impuesto a la Renta"),
        ("", "", ""),
        ("═" * 80, "", ""),
        ("3. DEFINICIÓN DE KPIs", "", ""),
        ("═" * 80, "", ""),
        ("", "", ""),
        ("Margen Bruto", "=", "(Utilidad Bruta / Ingresos) × 100 - Mide eficiencia en costos directos"),
        ("Margen Operacional", "=", "(EBIT / Ingresos) × 100 - Mide rentabilidad operativa"),
        ("Margen EBITDA", "=", "(EBITDA / Ingresos) × 100 - Rentabilidad antes de D&A"),
        ("Margen Neto", "=", "(Resultado Neto / Ingresos) × 100 - Rentabilidad final"),
        ("ROA", "=", "(Resultado Neto / Total Activos) × 100 - Retorno sobre activos"),
        ("ROE", "=", "(Resultado Neto / Patrimonio) × 100 - Retorno sobre patrimonio"),
        ("Ratio Endeudamiento", "=", "(Total Pasivos / Total Activos) × 100 - Nivel de deuda"),
        ("", "", ""),
        ("═" * 80, "", ""),
        ("4. AGRUPACIONES DEL BALANCE CLASIFICADO", "", ""),
        ("═" * 80, "", ""),
        ("", "", ""),
        ("ACTIVO CORRIENTE", "", "Cuentas 11xxxxx (Caja, Bancos, Cuentas por Cobrar, etc.) excepto Intangibles"),
        ("ACTIVO NO CORRIENTE", "", "Cuentas 12xxxxx y 13xxxxx (Activo Fijo) excepto Intangibles"),
        ("INTANGIBLES", "1109009, 1301001", "Depósito + Inversiones En Empresas Relacionada"),
        ("PASIVO CORRIENTE", "", "Cuentas 21xxxxx (Proveedores, Cuentas por Pagar CP)"),
        ("PASIVO NO CORRIENTE", "", "Cuentas 22xxxxx (Deudas LP, Notas Convertibles)"),
        ("PATRIMONIO", "", "Cuentas 31xxxxx (Capital, Utilidades Acumuladas)"),
        ("", "", ""),
        ("═" * 80, "", ""),
        ("5. NOTAS", "", ""),
        ("═" * 80, "", ""),
        ("", "", ""),
        ("• EBITDA se aproxima al EBIT ya que no hay cuentas de depreciación identificadas", "", ""),
        ("• El impuesto estimado usa tasa del 27% (tasa corporativa Chile)", "", ""),
        ("• Los valores negativos en gastos indican pérdidas/gastos", "", ""),
        ("• Los valores positivos en ingresos indican ganancias", "", ""),
    ])
    
    df = pd.DataFrame(data, columns=["Concepto", "Operador", "Descripción"])
    df.to_excel(writer, sheet_name="Documentación", index=False, header=False)
    
    ws = writer.sheets["Documentación"]
    font_titulo = Font(bold=True, size=14, color="006400")
    font_seccion = Font(bold=True, size=11)
    
    # Fila 1: Hipervínculo de navegación
    ws.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    
    # Fila 3: Título
    ws.cell(row=3, column=1).font = font_titulo
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 70


def crear_resumen(balance, writer, tenant_name, periodo):
    """Crea la hoja Resumen con Balance Clasificado, Estado de Resultados y KPIs"""
    
    # Crear diccionario de cuentas
    cuentas_dict = {c["idCuenta"]: c for c in balance}
    
    # ═══════════════════════════════════════════════════════════
    # BALANCE CLASIFICADO
    # ═══════════════════════════════════════════════════════════
    
    # Clasificar cuentas por tipo
    # Cuentas especiales que van en Intangibles
    CUENTAS_INTANGIBLES = ["1109009", "1301001"]  # Depósito + Inversiones Empresas Relacionadas
    
    activo_corriente = []
    activo_no_corriente = []
    intangibles = []
    pasivo_corriente = []
    pasivo_no_corriente = []
    patrimonio = []
    
    for c in balance:
        codigo = c["idCuenta"]
        
        # Primero verificar si es cuenta de intangibles
        if codigo in CUENTAS_INTANGIBLES:
            if c["activos"] != 0:
                intangibles.append((c["cuenta"], c["activos"]))
        elif codigo.startswith("11"):
            if c["activos"] != 0:
                activo_corriente.append((c["cuenta"], c["activos"]))
        elif codigo.startswith("12") or codigo.startswith("13"):
            if c["activos"] != 0:
                activo_no_corriente.append((c["cuenta"], c["activos"]))
        elif codigo.startswith("21"):
            if c["pasivos"] != 0:
                pasivo_corriente.append((c["cuenta"], c["pasivos"]))
        elif codigo.startswith("22"):
            if c["pasivos"] != 0:
                pasivo_no_corriente.append((c["cuenta"], c["pasivos"]))
        elif codigo.startswith("31"):
            # Patrimonio: pasivos con código 31 o activos negativos
            valor = c["pasivos"] if c["pasivos"] != 0 else -c["activos"]
            if valor != 0:
                patrimonio.append((c["cuenta"], valor))
    
    total_activo_corriente = sum(v for _, v in activo_corriente)
    total_activo_no_corriente = sum(v for _, v in activo_no_corriente)
    total_intangibles = sum(v for _, v in intangibles)
    total_activos = total_activo_corriente + total_activo_no_corriente + total_intangibles
    
    total_pasivo_corriente = sum(v for _, v in pasivo_corriente)
    total_pasivo_no_corriente = sum(v for _, v in pasivo_no_corriente)
    total_pasivos = total_pasivo_corriente + total_pasivo_no_corriente
    
    total_patrimonio = sum(v for _, v in patrimonio)
    
    # ═══════════════════════════════════════════════════════════
    # ESTADO DE RESULTADOS
    # ═══════════════════════════════════════════════════════════
    
    def get_valor(cuenta_id):
        if cuenta_id in cuentas_dict:
            c = cuentas_dict[cuenta_id]
            return c.get("ganancias", 0) - c.get("perdidas", 0)
        return 0
    
    def suma_grupo(cuentas_list):
        return sum(get_valor(c) for c in cuentas_list)
    
    ingresos = suma_grupo(AGRUPACIONES_EERR["Ingresos por Ventas"]["cuentas"])
    costo_ventas = suma_grupo(AGRUPACIONES_EERR["Costo de Ventas"]["cuentas"])
    utilidad_bruta = ingresos + costo_ventas
    
    remuneraciones = suma_grupo(AGRUPACIONES_EERR["Remuneraciones y Salarios"]["cuentas"])
    honorarios = suma_grupo(AGRUPACIONES_EERR["Honorarios Profesionales"]["cuentas"])
    arriendo = suma_grupo(AGRUPACIONES_EERR["Arriendo"]["cuentas"])
    marketing = suma_grupo(AGRUPACIONES_EERR["Marketing y Publicidad"]["cuentas"])
    legales = suma_grupo(AGRUPACIONES_EERR["Gastos Legales y Contables"]["cuentas"])
    otros_gastos = suma_grupo(AGRUPACIONES_EERR["Otros Gastos Operativos"]["cuentas"])
    total_gastos_op = remuneraciones + honorarios + arriendo + marketing + legales + otros_gastos
    
    resultado_operacional = utilidad_bruta + total_gastos_op
    
    multas = suma_grupo(AGRUPACIONES_EERR["Multas y Sanciones"]["cuentas"])
    gastos_financieros = suma_grupo(AGRUPACIONES_EERR["Gastos Financieros"]["cuentas"])
    
    resultado_antes_impuestos = resultado_operacional + multas + gastos_financieros
    impuesto = -resultado_antes_impuestos * 0.27 if resultado_antes_impuestos > 0 else 0
    resultado_neto = resultado_antes_impuestos + impuesto
    
    # ═══════════════════════════════════════════════════════════
    # KPIs
    # ═══════════════════════════════════════════════════════════
    
    margen_bruto = (utilidad_bruta / ingresos * 100) if ingresos != 0 else 0
    margen_operacional = (resultado_operacional / ingresos * 100) if ingresos != 0 else 0
    margen_neto = (resultado_neto / ingresos * 100) if ingresos != 0 else 0
    roa = (resultado_neto / total_activos * 100) if total_activos != 0 else 0
    roe = (resultado_neto / total_patrimonio * 100) if total_patrimonio != 0 else 0
    ratio_deuda = (total_pasivos / total_activos * 100) if total_activos != 0 else 0
    
    # ═══════════════════════════════════════════════════════════
    # CONSTRUIR HOJA
    # ═══════════════════════════════════════════════════════════
    
    rows = []
    row_types = []  # Para trackear el tipo de cada fila para formato
    
    # Fila 1: Navegación (se dejará vacía, el hipervínculo se agrega después)
    rows.append(["", "", "", "", ""])
    row_types.append("nav_row")
    
    # Header
    periodo_texto = f"{periodo[:4]}-{periodo[4:]}"
    rows.append([f"RESUMEN FINANCIERO - {tenant_name}", "", "", "", f"Período: {periodo_texto}"])
    row_types.append("header_main")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # BALANCE CLASIFICADO
    # ─────────────────────────────────────────────────────────
    rows.append(["BALANCE CLASIFICADO", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    # ACTIVOS
    rows.append(["ACTIVOS", "", "", "", ""])
    row_types.append("category")
    
    rows.append(["  Activo Corriente", "", "", "", ""])
    row_types.append("subcategory")
    for nombre, valor in activo_corriente[:5]:  # Top 5
        rows.append([f"    {nombre}", valor, "", "", ""])
        row_types.append("item")
    if len(activo_corriente) > 5:
        otros = sum(v for _, v in activo_corriente[5:])
        rows.append(["    Otros activos corrientes", otros, "", "", ""])
        row_types.append("item")
    rows.append(["  Total Activo Corriente", total_activo_corriente, "", "", ""])
    row_types.append("subtotal")
    
    rows.append(["  Activo No Corriente", "", "", "", ""])
    row_types.append("subcategory")
    for nombre, valor in activo_no_corriente:
        rows.append([f"    {nombre}", valor, "", "", ""])
        row_types.append("item")
    rows.append(["  Total Activo No Corriente", total_activo_no_corriente, "", "", ""])
    row_types.append("subtotal")
    
    rows.append(["  Intangibles", "", "", "", ""])
    row_types.append("subcategory")
    for nombre, valor in intangibles:
        rows.append([f"    {nombre}", valor, "", "", ""])
        row_types.append("item")
    rows.append(["  Total Intangibles", total_intangibles, "", "", ""])
    row_types.append("subtotal")
    
    rows.append(["TOTAL ACTIVOS", total_activos, "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # PASIVOS
    rows.append(["PASIVOS", "", "", "", ""])
    row_types.append("category")
    
    rows.append(["  Pasivo Corriente", "", "", "", ""])
    row_types.append("subcategory")
    for nombre, valor in pasivo_corriente[:5]:
        rows.append([f"    {nombre}", valor, "", "", ""])
        row_types.append("item")
    if len(pasivo_corriente) > 5:
        otros = sum(v for _, v in pasivo_corriente[5:])
        rows.append(["    Otros pasivos corrientes", otros, "", "", ""])
        row_types.append("item")
    rows.append(["  Total Pasivo Corriente", total_pasivo_corriente, "", "", ""])
    row_types.append("subtotal")
    
    rows.append(["  Pasivo No Corriente", "", "", "", ""])
    row_types.append("subcategory")
    for nombre, valor in pasivo_no_corriente:
        rows.append([f"    {nombre}", valor, "", "", ""])
        row_types.append("item")
    rows.append(["  Total Pasivo No Corriente", total_pasivo_no_corriente, "", "", ""])
    row_types.append("subtotal")
    
    rows.append(["TOTAL PASIVOS", total_pasivos, "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # PATRIMONIO
    rows.append(["PATRIMONIO", "", "", "", ""])
    row_types.append("category")
    for nombre, valor in patrimonio:
        rows.append([f"  {nombre}", valor, "", "", ""])
        row_types.append("item")
    rows.append(["TOTAL PATRIMONIO", total_patrimonio, "", "", ""])
    row_types.append("total")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["TOTAL PASIVOS + PATRIMONIO", total_pasivos + total_patrimonio, "", "", ""])
    row_types.append("total_final")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # ESTADO DE RESULTADOS
    # ─────────────────────────────────────────────────────────
    eerr_start_row = len(rows) + 1
    
    rows.append(["ESTADO DE RESULTADOS", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    rows.append(["Ingresos por Ventas", ingresos, "", "", ""])
    row_types.append("item")
    rows.append(["Costo de Ventas", costo_ventas, "", "", ""])
    row_types.append("item")
    rows.append(["UTILIDAD BRUTA", utilidad_bruta, "", "", ""])
    row_types.append("subtotal")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["Gastos Operacionales:", "", "", "", ""])
    row_types.append("subcategory")
    rows.append(["  Remuneraciones y Salarios", remuneraciones, "", "", ""])
    row_types.append("item")
    rows.append(["  Honorarios Profesionales", honorarios, "", "", ""])
    row_types.append("item")
    rows.append(["  Arriendo", arriendo, "", "", ""])
    row_types.append("item")
    rows.append(["  Marketing y Publicidad", marketing, "", "", ""])
    row_types.append("item")
    rows.append(["  Gastos Legales y Contables", legales, "", "", ""])
    row_types.append("item")
    rows.append(["  Otros Gastos Operativos", otros_gastos, "", "", ""])
    row_types.append("item")
    rows.append(["Total Gastos Operacionales", total_gastos_op, "", "", ""])
    row_types.append("subtotal")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["RESULTADO OPERACIONAL (EBIT)", resultado_operacional, "", "", ""])
    row_types.append("total")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["Otros Gastos:", "", "", "", ""])
    row_types.append("subcategory")
    rows.append(["  Multas y Sanciones", multas, "", "", ""])
    row_types.append("item")
    rows.append(["  Gastos Financieros", gastos_financieros, "", "", ""])
    row_types.append("item")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    rows.append(["RESULTADO ANTES DE IMPUESTOS", resultado_antes_impuestos, "", "", ""])
    row_types.append("total")
    rows.append(["Impuesto a la Renta (27%)", impuesto, "", "", ""])
    row_types.append("item")
    rows.append(["RESULTADO NETO", resultado_neto, "", "", ""])
    row_types.append("total_final")
    
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    rows.append(["", "", "", "", ""])
    row_types.append("empty")
    
    # ─────────────────────────────────────────────────────────
    # KPIs
    # ─────────────────────────────────────────────────────────
    kpi_start_row = len(rows) + 1
    
    rows.append(["INDICADORES FINANCIEROS (KPIs)", "", "", "", "Ver Documentación →"])
    row_types.append("section_title")
    rows.append(["─" * 40, "", "", "", ""])
    row_types.append("separator")
    
    rows.append(["", "Valor", "Interpretación", "", ""])
    row_types.append("header_kpi")
    
    rows.append(["Margen Bruto", f"{margen_bruto:.1f}%", "Eficiencia en costos directos", "", ""])
    row_types.append("kpi")
    rows.append(["Margen Operacional", f"{margen_operacional:.1f}%", "Rentabilidad operativa", "", ""])
    row_types.append("kpi")
    rows.append(["Margen Neto", f"{margen_neto:.1f}%", "Rentabilidad final", "", ""])
    row_types.append("kpi")
    rows.append(["ROA", f"{roa:.1f}%", "Retorno sobre activos", "", ""])
    row_types.append("kpi")
    rows.append(["ROE", f"{roe:.1f}%", "Retorno sobre patrimonio", "", ""])
    row_types.append("kpi")
    rows.append(["Ratio de Endeudamiento", f"{ratio_deuda:.1f}%", "Nivel de apalancamiento", "", ""])
    row_types.append("kpi")
    
    # Escribir a Excel
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Resumen", index=False, header=False)
    
    ws = writer.sheets["Resumen"]
    
    # Estilos
    font_header = Font(bold=True, size=16, color="006400")
    font_section = Font(bold=True, size=13, color="006400")
    font_category = Font(bold=True, size=11)
    font_subtotal = Font(bold=True, size=10)
    font_total = Font(bold=True, size=11)
    font_total_final = Font(bold=True, size=12, color="FFFFFF")
    fill_total_final = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    fill_section = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    formato_miles = '#,##0'
    
    # Aplicar formatos
    for row_idx, row_type in enumerate(row_types, start=1):
        cell_a = ws.cell(row=row_idx, column=1)
        cell_b = ws.cell(row=row_idx, column=2)
        cell_e = ws.cell(row=row_idx, column=5)
        
        if row_type == "header_main":
            cell_a.font = font_header
        elif row_type == "section_title":
            cell_a.font = font_section
            cell_a.fill = fill_section
            # Hipervínculo a documentación
            cell_e.hyperlink = "#'Documentación'!A1"
            cell_e.style = "Hyperlink"
        elif row_type == "category":
            cell_a.font = font_category
        elif row_type == "subcategory":
            cell_a.font = Font(bold=True, italic=True)
        elif row_type == "subtotal":
            cell_a.font = font_subtotal
            cell_b.font = font_subtotal
        elif row_type == "total":
            cell_a.font = font_total
            cell_b.font = font_total
        elif row_type == "total_final":
            cell_a.font = font_total_final
            cell_a.fill = fill_total_final
            cell_b.font = font_total_final
            cell_b.fill = fill_total_final
        elif row_type == "header_kpi":
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
            ws.cell(row=row_idx, column=3).font = Font(bold=True)
        elif row_type == "kpi":
            cell_b.font = Font(bold=True, color="006400")
        
        # Formato de miles para columna B si es número
        if isinstance(rows[row_idx-1][1], (int, float)):
            cell_b.number_format = formato_miles
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['E'].width = 22
    
    # Agregar texto de navegación en A1 (es la hoja principal)
    ws.cell(row=1, column=1).value = "📊 DASHBOARD PRINCIPAL"
    ws.cell(row=1, column=1).font = Font(bold=True, size=10, color="666666")
    
    # Retornar valores para uso externo
    return {
        "ingresos": ingresos,
        "utilidad_bruta": utilidad_bruta,
        "resultado_operacional": resultado_operacional,
        "resultado_neto": resultado_neto,
        "total_activos": total_activos,
        "total_pasivos": total_pasivos,
        "total_patrimonio": total_patrimonio
    }


def crear_estados_financieros_comparativos(tenant_rut, periodos, writer):
    """
    Crea una hoja con Estados Financieros Comparativos:
    - Balance General
    - Estado de Resultados
    - Análisis de Tendencia
    - KPIs
    """
    
    # Obtener balances de todos los períodos
    balances = {}
    for id_periodo, nombre in periodos:
        print(f"   📊 Obteniendo balance {nombre}...")
        balance = get_balance(tenant_rut, id_periodo)
        if balance:
            balances[nombre] = {c["idCuenta"]: c for c in balance}
    
    if not balances:
        return
    
    # Obtener todas las cuentas únicas
    todas_cuentas = set()
    for balance_dict in balances.values():
        todas_cuentas.update(balance_dict.keys())
    cuentas_ordenadas = sorted(todas_cuentas)
    
    # Definir categorías especiales
    CUENTAS_INTANGIBLES = ["1109009", "1301001"]
    
    # Construir datos
    rows = []
    row_types = []
    
    # Calcular número de columnas: Cuenta + períodos + variaciones
    num_periodos = len(periodos)
    nombres_periodos = [nombre for _, nombre in periodos]
    
    # Fila 1: Navegación
    nav_row = ["← Volver al Resumen", ""] + [""] * len(periodos)
    rows.append(nav_row)
    row_types.append("nav_row")
    
    # Fila 2: Título
    titulo_row = ["ESTADOS FINANCIEROS COMPARATIVOS", ""] + [""] * len(periodos)
    rows.append(titulo_row)
    row_types.append("titulo_principal")
    
    # Fila vacía
    empty_row = ["", ""] + [""] * len(periodos)
    rows.append(empty_row)
    row_types.append("empty")
    
    # Header (sin variaciones)
    header = ["Código", "Cuenta"] + [nombre for _, nombre in periodos]
    rows.append(header)
    row_types.append("header")
    
    # Variables para totales por período
    totales = {nombre: {
        "activos": 0, "pasivos": 0, "patrimonio": 0,
        "activo_corriente": 0, "activo_no_corriente": 0, "intangibles": 0,
        "pasivo_corriente": 0, "pasivo_no_corriente": 0
    } for _, nombre in periodos}
    
    # ═══════════════════════════════════════════════════════════
    # BALANCE GENERAL
    # ═══════════════════════════════════════════════════════════
    rows.append(["", "BALANCE GENERAL"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ACTIVOS
    rows.append(["", "ACTIVOS"] + [""] * len(periodos))
    row_types.append("category")
    
    # Activo Corriente
    rows.append(["", "  Activo Corriente"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for codigo in cuentas_ordenadas:
        if codigo.startswith("11") and codigo not in CUENTAS_INTANGIBLES:
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["activos"]
                    valores.append(val)
                    totales[nombre_periodo]["activo_corriente"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    # Subtotal Activo Corriente
    subtotal_row = ["", "  Total Activo Corriente"] + [totales[nombre]["activo_corriente"] for _, nombre in periodos]
    rows.append(subtotal_row)
    row_types.append("subtotal")
    
    # Activo No Corriente
    rows.append(["", "  Activo No Corriente"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for codigo in cuentas_ordenadas:
        if (codigo.startswith("12") or codigo.startswith("13")) and codigo not in CUENTAS_INTANGIBLES:
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["activos"]
                    valores.append(val)
                    totales[nombre_periodo]["activo_no_corriente"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    rows.append(["", "  Total Activo No Corriente"] + [totales[nombre]["activo_no_corriente"] for _, nombre in periodos])
    row_types.append("subtotal")
    
    # Intangibles
    rows.append(["", "  Intangibles"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for codigo in CUENTAS_INTANGIBLES:
        if codigo in cuentas_ordenadas:
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["activos"]
                    valores.append(val)
                    totales[nombre_periodo]["intangibles"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    rows.append(["", "  Total Intangibles"] + [totales[nombre]["intangibles"] for _, nombre in periodos])
    row_types.append("subtotal")
    
    # Total Activos
    for _, nombre in periodos:
        totales[nombre]["activos"] = (totales[nombre]["activo_corriente"] + 
                                       totales[nombre]["activo_no_corriente"] + 
                                       totales[nombre]["intangibles"])
    
    rows.append(["", "TOTAL ACTIVOS"] + [totales[nombre]["activos"] for _, nombre in periodos])
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # PASIVOS
    rows.append(["", "PASIVOS"] + [""] * len(periodos))
    row_types.append("category")
    
    # Pasivo Corriente
    rows.append(["", "  Pasivo Corriente"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for codigo in cuentas_ordenadas:
        if codigo.startswith("21"):
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["pasivos"]
                    valores.append(val)
                    totales[nombre_periodo]["pasivo_corriente"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    rows.append(["", "  Total Pasivo Corriente"] + [totales[nombre]["pasivo_corriente"] for _, nombre in periodos])
    row_types.append("subtotal")
    
    # Pasivo No Corriente
    rows.append(["", "  Pasivo No Corriente"] + [""] * len(periodos))
    row_types.append("subcategory")
    
    for codigo in cuentas_ordenadas:
        if codigo.startswith("22"):
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    val = balances[nombre_periodo][codigo]["pasivos"]
                    valores.append(val)
                    totales[nombre_periodo]["pasivo_no_corriente"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"    {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    rows.append(["", "  Total Pasivo No Corriente"] + [totales[nombre]["pasivo_no_corriente"] for _, nombre in periodos])
    row_types.append("subtotal")
    
    # Total Pasivos
    for _, nombre in periodos:
        totales[nombre]["pasivos"] = totales[nombre]["pasivo_corriente"] + totales[nombre]["pasivo_no_corriente"]
    
    rows.append(["", "TOTAL PASIVOS"] + [totales[nombre]["pasivos"] for _, nombre in periodos])
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # PATRIMONIO
    rows.append(["", "PATRIMONIO"] + [""] * len(periodos))
    row_types.append("category")
    
    for codigo in cuentas_ordenadas:
        if codigo.startswith("31"):
            nombre_cuenta = ""
            for balance_dict in balances.values():
                if codigo in balance_dict:
                    nombre_cuenta = balance_dict[codigo]["cuenta"]
                    break
            
            valores = []
            tiene_valor = False
            for i, (_, nombre_periodo) in enumerate(periodos):
                if nombre_periodo in balances and codigo in balances[nombre_periodo]:
                    c = balances[nombre_periodo][codigo]
                    val = c["pasivos"] if c["pasivos"] != 0 else -c["activos"]
                    valores.append(val)
                    totales[nombre_periodo]["patrimonio"] += val
                    if val != 0:
                        tiene_valor = True
                else:
                    valores.append(0)
            
            if tiene_valor:
                rows.append([codigo, f"  {nombre_cuenta}"] + valores)
                row_types.append("item")
    
    rows.append(["", "TOTAL PATRIMONIO"] + [totales[nombre]["patrimonio"] for _, nombre in periodos])
    row_types.append("total")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Total Pasivos + Patrimonio
    rows.append(["", "TOTAL PASIVOS + PATRIMONIO"] + [totales[nombre]["pasivos"] + totales[nombre]["patrimonio"] for _, nombre in periodos])
    row_types.append("total_final")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ═══════════════════════════════════════════════════════════
    # ESTADO DE RESULTADOS
    # ═══════════════════════════════════════════════════════════
    rows.append(["", "ESTADO DE RESULTADOS"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Calcular valores del EERR por período
    eerr = {nombre: {} for _, nombre in periodos}
    
    for _, nombre_periodo in periodos:
        if nombre_periodo not in balances:
            continue
        balance_dict = balances[nombre_periodo]
        
        def get_eerr_valor(cuenta_id):
            if cuenta_id in balance_dict:
                c = balance_dict[cuenta_id]
                return c.get("ganancias", 0) - c.get("perdidas", 0)
            return 0
        
        def suma_eerr(cuentas_list):
            return sum(get_eerr_valor(c) for c in cuentas_list)
        
        ingresos = suma_eerr(AGRUPACIONES_EERR["Ingresos por Ventas"]["cuentas"])
        costo_ventas = suma_eerr(AGRUPACIONES_EERR["Costo de Ventas"]["cuentas"])
        utilidad_bruta = ingresos + costo_ventas
        
        remuneraciones = suma_eerr(AGRUPACIONES_EERR["Remuneraciones y Salarios"]["cuentas"])
        honorarios = suma_eerr(AGRUPACIONES_EERR["Honorarios Profesionales"]["cuentas"])
        arriendo = suma_eerr(AGRUPACIONES_EERR["Arriendo"]["cuentas"])
        marketing = suma_eerr(AGRUPACIONES_EERR["Marketing y Publicidad"]["cuentas"])
        legales = suma_eerr(AGRUPACIONES_EERR["Gastos Legales y Contables"]["cuentas"])
        otros_gastos = suma_eerr(AGRUPACIONES_EERR["Otros Gastos Operativos"]["cuentas"])
        total_gastos_op = remuneraciones + honorarios + arriendo + marketing + legales + otros_gastos
        
        resultado_operacional = utilidad_bruta + total_gastos_op
        
        multas = suma_eerr(AGRUPACIONES_EERR["Multas y Sanciones"]["cuentas"])
        gastos_financieros = suma_eerr(AGRUPACIONES_EERR["Gastos Financieros"]["cuentas"])
        
        resultado_antes_impuestos = resultado_operacional + multas + gastos_financieros
        impuesto = -resultado_antes_impuestos * 0.27 if resultado_antes_impuestos > 0 else 0
        resultado_neto = resultado_antes_impuestos + impuesto
        
        eerr[nombre_periodo] = {
            "ingresos": ingresos,
            "costo_ventas": costo_ventas,
            "utilidad_bruta": utilidad_bruta,
            "remuneraciones": remuneraciones,
            "honorarios": honorarios,
            "arriendo": arriendo,
            "marketing": marketing,
            "legales": legales,
            "otros_gastos": otros_gastos,
            "total_gastos_op": total_gastos_op,
            "resultado_operacional": resultado_operacional,
            "multas": multas,
            "gastos_financieros": gastos_financieros,
            "resultado_antes_impuestos": resultado_antes_impuestos,
            "impuesto": impuesto,
            "resultado_neto": resultado_neto
        }
    
    # Helper para construir filas del EERR
    def eerr_row(concepto, campo, tipo="item"):
        row = ["", concepto] + [eerr.get(nombre, {}).get(campo, 0) for _, nombre in periodos]
        rows.append(row)
        row_types.append(tipo)
    
    eerr_row("Ingresos por Ventas", "ingresos")
    eerr_row("Costo de Ventas", "costo_ventas")
    eerr_row("UTILIDAD BRUTA", "utilidad_bruta", "subtotal")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Gastos Operacionales:"] + [""] * len(periodos))
    row_types.append("subcategory")
    eerr_row("  Remuneraciones y Salarios", "remuneraciones")
    eerr_row("  Honorarios Profesionales", "honorarios")
    eerr_row("  Arriendo", "arriendo")
    eerr_row("  Marketing y Publicidad", "marketing")
    eerr_row("  Gastos Legales y Contables", "legales")
    eerr_row("  Otros Gastos Operativos", "otros_gastos")
    eerr_row("Total Gastos Operacionales", "total_gastos_op", "subtotal")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    eerr_row("RESULTADO OPERACIONAL (EBIT)", "resultado_operacional", "total")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Otros Gastos:"] + [""] * len(periodos))
    row_types.append("subcategory")
    eerr_row("  Multas y Sanciones", "multas")
    eerr_row("  Gastos Financieros", "gastos_financieros")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    eerr_row("RESULTADO ANTES DE IMPUESTOS", "resultado_antes_impuestos", "total")
    eerr_row("Impuesto a la Renta (27%)", "impuesto")
    eerr_row("RESULTADO NETO", "resultado_neto", "total_final")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # ═══════════════════════════════════════════════════════════
    # KPIs COMPARATIVOS
    # ═══════════════════════════════════════════════════════════
    rows.append(["", "INDICADORES FINANCIEROS (KPIs)"] + [""] * len(periodos))
    row_types.append("section_title")
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    # Calcular KPIs por período
    kpis = {}
    for _, nombre_periodo in periodos:
        if nombre_periodo not in eerr or nombre_periodo not in totales:
            continue
        
        e = eerr[nombre_periodo]
        t = totales[nombre_periodo]
        
        ingresos = e.get("ingresos", 0) or 1
        total_activos = t["activos"] or 1
        patrimonio = t["patrimonio"] or 1
        
        kpis[nombre_periodo] = {
            "margen_bruto": (e.get("utilidad_bruta", 0) / ingresos * 100) if e.get("ingresos", 0) else 0,
            "margen_operacional": (e.get("resultado_operacional", 0) / ingresos * 100) if e.get("ingresos", 0) else 0,
            "margen_neto": (e.get("resultado_neto", 0) / ingresos * 100) if e.get("ingresos", 0) else 0,
            "roa": (e.get("resultado_neto", 0) / total_activos * 100),
            "roe": (e.get("resultado_neto", 0) / patrimonio * 100),
            "ratio_deuda": (t["pasivos"] / total_activos * 100),
            "razon_corriente": t["activo_corriente"] / t["pasivo_corriente"] if t["pasivo_corriente"] else 0
        }
    
    def kpi_row(nombre_kpi, campo, formato="{:.1f}%"):
        row = ["", nombre_kpi] + [formato.format(kpis.get(nombre, {}).get(campo, 0)) for _, nombre in periodos]
        rows.append(row)
        row_types.append("kpi")
    
    rows.append(["", "Márgenes de Rentabilidad"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  Margen Bruto", "margen_bruto")
    kpi_row("  Margen Operacional", "margen_operacional")
    kpi_row("  Margen Neto", "margen_neto")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Rentabilidad sobre Inversión"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  ROA (Return on Assets)", "roa")
    kpi_row("  ROE (Return on Equity)", "roe")
    
    rows.append(["", ""] + [""] * len(periodos))
    row_types.append("empty")
    
    rows.append(["", "Estructura Financiera"] + [""] * len(periodos))
    row_types.append("subcategory")
    kpi_row("  Ratio de Endeudamiento", "ratio_deuda")
    kpi_row("  Razón Corriente", "razon_corriente", "{:.2f}x")
    
    # ═══════════════════════════════════════════════════════════
    # ESCRIBIR A EXCEL Y APLICAR FORMATO
    # ═══════════════════════════════════════════════════════════
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="EEFF Comparativos", index=False, header=False)
    
    ws = writer.sheets["EEFF Comparativos"]
    
    # Estilos
    font_header = Font(bold=True, size=11, color="FFFFFF")
    fill_header = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    font_section = Font(bold=True, size=13, color="006400")
    fill_section = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    font_category = Font(bold=True, size=11)
    font_subcategory = Font(bold=True, italic=True, size=10)
    font_subtotal = Font(bold=True, size=10)
    font_total = Font(bold=True, size=11)
    font_total_final = Font(bold=True, size=12, color="FFFFFF")
    fill_total_final = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    fill_subtotal = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    font_kpi = Font(bold=True, color="006400")
    fill_var_positive = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    fill_var_negative = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    formato_miles = '#,##0'
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))
    
    num_cols = len(header)
    
    for row_idx, row_type in enumerate(row_types, start=1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if row_type == "nav_row":
                if col_idx == 1:
                    cell.hyperlink = "#'Resumen'!A1"
                    cell.style = "Hyperlink"
            elif row_type == "titulo_principal":
                if col_idx == 1:
                    cell.font = Font(bold=True, size=14, color="006400")
            elif row_type == "header":
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center')
            elif row_type == "section_title":
                cell.font = font_section
                cell.fill = fill_section
            elif row_type == "category":
                cell.font = font_category
            elif row_type == "subcategory":
                cell.font = font_subcategory
            elif row_type == "subtotal":
                cell.font = font_subtotal
                cell.fill = fill_subtotal
                cell.border = thin_border
            elif row_type == "total":
                cell.font = font_total
                cell.border = thin_border
            elif row_type == "total_final":
                cell.font = font_total_final
                cell.fill = fill_total_final
            elif row_type == "kpi":
                if col_idx == 2:
                    pass
                elif col_idx >= 3:
                    cell.font = font_kpi
                    cell.alignment = Alignment(horizontal='center')
            
            # Formato de miles para columnas numéricas
            if col_idx >= 3 and row_idx > 1:
                val = rows[row_idx-1][col_idx-1] if col_idx-1 < len(rows[row_idx-1]) else None
                if isinstance(val, (int, float)):
                    cell.number_format = formato_miles
                # Color para variaciones
                if isinstance(val, str) and "%" in val and row_type not in ["kpi", "header"]:
                    cell.alignment = Alignment(horizontal='center')
                    if val.startswith("-"):
                        cell.fill = fill_var_negative
                    elif val not in ["N/A", "∞"] and not val.startswith("0"):
                        cell.fill = fill_var_positive
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    col_idx = 3
    for i, (_, nombre) in enumerate(periodos):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16
        col_idx += 1
        if i > 0:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 10
            col_idx += 1


def main():
    print("═" * 60)
    print("   BALANCE + ANÁLISIS POR CUENTA A EXCEL")
    print("═" * 60)

    # Configuración
    tenant_key = "FIDI"
    tenant = TENANTS[tenant_key]
    id_periodo = "202511"
    fecha_corte = "2025-11-30"

    # 1. Obtener Balance
    print("\n📊 Obteniendo Balance Tributario...")
    balance = get_balance(tenant["rut"], id_periodo)
    
    if not balance:
        print("   ❌ Error obteniendo balance")
        return

    # Filtrar cuentas con saldo != 0
    balance_filtrado = [
        c for c in balance
        if c["deudor"] != 0 or c["acreedor"] != 0 or
           c["activos"] != 0 or c["pasivos"] != 0 or
           c["perdidas"] != 0 or c["ganancias"] != 0
    ]
    print(f"   ✅ {len(balance_filtrado)} cuentas (de {len(balance)} totales)")

    # Crear DataFrame del balance
    df_balance = pd.DataFrame(balance_filtrado)
    df_balance = df_balance.rename(columns={
        "idCuenta": "Código",
        "cuenta": "Cuenta",
        "debitos": "Débitos",
        "creditos": "Créditos",
        "deudor": "Saldo Deudor",
        "acreedor": "Saldo Acreedor",
        "activos": "Activos",
        "pasivos": "Pasivos",
        "perdidas": "Pérdidas",
        "ganancias": "Ganancias"
    })

    # Agregar columna para hipervínculo
    df_balance["Ver Detalle"] = ""
    
    # 2. Crear Excel con múltiples hojas
    # Crear carpeta si no existe
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Nombre con timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{OUTPUT_DIR}/Balance_PorCuenta_{tenant_key}_{id_periodo}_{timestamp}.xlsx"
    
    # Mapeo de código de cuenta -> nombre de hoja
    cuenta_a_hoja = {}
    
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # 3. Crear hoja Resumen primero (será la primera hoja)
        print("\n📈 Generando Resumen (Balance + EERR + KPIs)...")
        crear_resumen(balance, writer, tenant["nombre"], id_periodo)
        
        # 4. Crear Estados Financieros Comparativos (Marzo, Junio, Noviembre)
        print("\n📊 Generando Estados Financieros Comparativos...")
        periodos_comparar = [
            ("202503", "Mar 2025"),
            ("202506", "Jun 2025"),
            ("202511", "Nov 2025")
        ]
        crear_estados_financieros_comparativos(tenant["rut"], periodos_comparar, writer)
        
        # 5. Crear Documentación
        print("\n📝 Generando Documentación...")
        crear_documentacion(writer)
        
        # 6. Hoja de Balance Tributario detallado (con fila 1 libre para navegación)
        # Escribir con startrow=2 para dejar filas 1 y 2 libres
        df_balance.to_excel(writer, sheet_name="Balance Tributario", index=False, startrow=2)
        
        ws_balance = writer.sheets["Balance Tributario"]
        
        # A1: Hipervínculo de retorno
        ws_balance.cell(row=1, column=1).value = "← Volver al Resumen"
        ws_balance.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
        ws_balance.cell(row=1, column=1).style = "Hyperlink"
        
        # A2: Período del documento
        periodo_texto = f"Período: {id_periodo[:4]}-{id_periodo[4:]}"
        ws_balance.cell(row=2, column=1).value = f"BALANCE TRIBUTARIO - {tenant['nombre']} - {periodo_texto}"
        ws_balance.cell(row=2, column=1).font = Font(bold=True, size=12, color="006400")
        
        # 7. Para cada cuenta, obtener análisis
        print("\n📋 Procesando cuentas con movimientos...")
        cuentas_con_datos = 0
        sheet_names = {"Balance Tributario", "Resumen", "EEFF Comparativos", "Documentación"}

        for cuenta in balance_filtrado:
            analisis = get_analisis_cuenta(tenant["rut"], cuenta["idCuenta"], fecha_corte)
            
            if analisis and len(analisis) > 0:
                # Crear DataFrame y filtrar saldos 0
                df_cuenta = pd.DataFrame(analisis)
                if "saldo" in df_cuenta.columns:
                    df_cuenta = df_cuenta[df_cuenta["saldo"] != 0]
                
                if len(df_cuenta) == 0:
                    continue
                
                # Seleccionar y renombrar solo columnas relevantes
                cols_deseadas = ["fecha", "numero", "tipo", "glosa", "debe", "haber", "saldo"]
                cols_disponibles = [c for c in cols_deseadas if c in df_cuenta.columns]
                
                # Si tiene auxiliar, agregarlo
                if "auxiliar" in df_cuenta.columns:
                    cols_disponibles.insert(3, "auxiliar")
                elif "idAuxiliar" in df_cuenta.columns:
                    df_cuenta["auxiliar"] = df_cuenta["idAuxiliar"]
                    cols_disponibles.insert(3, "auxiliar")
                
                df_cuenta = df_cuenta[cols_disponibles]
                
                # Renombrar columnas
                cols_rename = {
                    "fecha": "Fecha",
                    "numero": "Comprobante",
                    "tipo": "Tipo",
                    "auxiliar": "Auxiliar",
                    "glosa": "Glosa",
                    "debe": "Debe",
                    "haber": "Haber",
                    "saldo": "Saldo"
                }
                df_cuenta = df_cuenta.rename(columns=cols_rename)
                
                # Formatear fecha
                if "Fecha" in df_cuenta.columns:
                    df_cuenta["Fecha"] = df_cuenta["Fecha"].str[:10]

                # Nombre de hoja único
                sheet_name = sanitize_sheet_name(cuenta["idCuenta"], cuenta["cuenta"])
                suffix = 1
                original = sheet_name
                while sheet_name in sheet_names:
                    sheet_name = f"{original[:28]}_{suffix}"
                    suffix += 1
                
                sheet_names.add(sheet_name)
                
                # Escribir con startrow=2 para dejar filas 1 y 2 libres
                df_cuenta.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                cuenta_a_hoja[cuenta["idCuenta"]] = sheet_name
                cuentas_con_datos += 1
                print(f"   ✅ {cuenta['idCuenta']}: {len(df_cuenta)} movimientos")
                
                # Formatear hoja de análisis
                ws_cuenta = writer.sheets[sheet_name]
                
                # A1: Hipervínculo de retorno
                ws_cuenta.cell(row=1, column=1).value = "← Volver al Resumen"
                ws_cuenta.cell(row=1, column=1).hyperlink = "#'Resumen'!A1"
                ws_cuenta.cell(row=1, column=1).style = "Hyperlink"
                
                # A2: Título con cuenta y período
                ws_cuenta.cell(row=2, column=1).value = f"ANÁLISIS DE CUENTA: {cuenta['idCuenta']} - {cuenta['cuenta']} | {periodo_texto}"
                ws_cuenta.cell(row=2, column=1).font = Font(bold=True, size=11, color="006400")

        print(f"\n   📊 Cuentas con datos: {cuentas_con_datos}")
        
        # 8. Aplicar formato profesional a todas las hojas
        print("\n💰 Aplicando formato...")
        formato_miles = '#,##0'
        
        # Estilos comunes
        font_header = Font(bold=True, size=10, color="FFFFFF")
        fill_header = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Formato en Balance Tributario
        ws_balance = writer.sheets["Balance Tributario"]
        
        # Header en fila 3
        for col in range(1, 12):
            cell = ws_balance.cell(row=3, column=col)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Datos con formato (empiezan en fila 4)
        for row in range(4, len(balance_filtrado) + 4):
            for col in range(1, 12):
                cell = ws_balance.cell(row=row, column=col)
                cell.border = thin_border
                if col >= 3 and col <= 10:  # Columnas numéricas
                    cell.number_format = formato_miles
        
        # Anchos de columna
        ws_balance.column_dimensions['A'].width = 12
        ws_balance.column_dimensions['B'].width = 35
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_balance.column_dimensions[col].width = 14
        ws_balance.column_dimensions['K'].width = 10
        
        # Formato en hojas de cuentas
        for codigo, sheet_name in cuenta_a_hoja.items():
            ws = writer.sheets[sheet_name]
            num_cols = ws.max_column
            
            # Header en fila 3
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=3, column=col)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Datos con formato (empiezan en fila 4)
            for row in range(4, ws.max_row + 1):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    # Aplicar formato miles a las últimas 3 columnas (Debe, Haber, Saldo)
                    if col >= num_cols - 2:
                        cell.number_format = formato_miles
            
            # Anchos de columna dinámicos
            ws.column_dimensions['A'].width = 12  # Fecha
            ws.column_dimensions['B'].width = 12  # Comprobante
            ws.column_dimensions['C'].width = 8   # Tipo
            if num_cols == 8:  # Con Auxiliar
                ws.column_dimensions['D'].width = 18  # Auxiliar
                ws.column_dimensions['E'].width = 35  # Glosa
                ws.column_dimensions['F'].width = 14  # Debe
                ws.column_dimensions['G'].width = 14  # Haber
                ws.column_dimensions['H'].width = 14  # Saldo
            else:  # Sin Auxiliar
                ws.column_dimensions['D'].width = 35  # Glosa
                ws.column_dimensions['E'].width = 14  # Debe
                ws.column_dimensions['F'].width = 14  # Haber
                ws.column_dimensions['G'].width = 14  # Saldo
        
        # 9. Agregar hipervínculos en columna K del Balance Tributario
        print("🔗 Agregando hipervínculos...")
        ws_balance = writer.sheets["Balance Tributario"]
        
        for row_idx, cuenta in enumerate(balance_filtrado, start=4):  # start=4 por filas libres + header
            codigo = cuenta["idCuenta"]
            if codigo in cuenta_a_hoja:
                sheet_name = cuenta_a_hoja[codigo]
                cell = ws_balance.cell(row=row_idx, column=11)
                cell.value = "→ Ver"
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.style = "Hyperlink"
        
    print(f"\n💾 Guardado: {filename}")
    print("\n" + "═" * 60)


if __name__ == "__main__":
    main()

